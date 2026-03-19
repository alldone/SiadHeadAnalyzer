from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
import json
import os
from pathlib import Path
import subprocess
import sys
from tempfile import TemporaryDirectory
from typing import Any, Iterable
import re
import threading
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import xmlschema
except ModuleNotFoundError:
    xmlschema = None


MONTH_CODES = {
    "A": 1,
    "B": 2,
    "C": 3,
    "D": 4,
    "E": 5,
    "H": 6,
    "L": 7,
    "M": 8,
    "P": 9,
    "R": 10,
    "S": 11,
    "T": 12,
}

FEMALE_DAY_OFFSET = 40
CF_PATTERN = re.compile(r"^[A-Z]{6}\d{2}[ABCDEHLMPRST]\d{2}[A-Z]\d{3}[A-Z]$")
XML_NS_RE = re.compile(r"^\{(?P<ns>[^}]+)\}(?P<name>.+)$")
QUARTER_PATH_RE = re.compile(r"_(20\d{2})_([1-4])_")


def local_name(tag: str) -> str:
    match = XML_NS_RE.match(tag)
    return match.group("name") if match else tag


def parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def quarter_from_path(relative_path: str) -> str:
    match = QUARTER_PATH_RE.search(relative_path)
    if match:
        return f"T{match.group(2)}"
    return ""


def next_available_report_path(xml_dir: Path) -> Path:
    base_dir = xml_dir.parent
    base_name = "report_siad"
    candidate = base_dir / f"{base_name}.xlsx"
    if not candidate.exists():
        return candidate

    suffix_pattern = re.compile(rf"^{re.escape(base_name)}_(\d+)\.xlsx$")
    max_suffix = 1
    for existing_file in base_dir.glob(f"{base_name}*.xlsx"):
        match = suffix_pattern.match(existing_file.name)
        if match:
            max_suffix = max(max_suffix, int(match.group(1)))
        elif existing_file.name == f"{base_name}.xlsx":
            max_suffix = max(max_suffix, 1)

    return base_dir / f"{base_name}_{max_suffix + 1}.xlsx"


@dataclass
class XmlFileInfo:
    path: Path
    relative_path: str
    track: int
    root_name: str
    line_count: int
    size_mb: float


@dataclass
class RecordDetail:
    sede: str
    track: int
    source_file: str
    quarter: str
    presa_in_carico_date: str
    id_rec: str
    codice_fiscale: str
    anno_nascita_xml: int | None
    anno_nascita_usato: int | None
    eta_al_31_12: int | None
    is_over_65: bool
    cf_ambiguo: bool
    included_in_report: bool
    note: str


@dataclass
class PatientInfo:
    codice_fiscale: str
    guessed_year: int | None
    resolved_year: int | None = None
    resolved_from_track1: bool = False
    ambiguous: bool = False


def parse_xsd_root_name(xsd_path: Path) -> str:
    tree = ET.parse(xsd_path)
    root = tree.getroot()
    for child in root:
        if local_name(child.tag) == "element":
            return child.attrib["name"]
    raise ValueError(f"Impossibile determinare l'elemento radice da {xsd_path}")


def classify_xml_file(xml_path: Path, root_to_track: dict[str, int]) -> tuple[int, str]:
    root = ET.parse(xml_path).getroot()
    root_name = local_name(root.tag)
    track = root_to_track.get(root_name)
    if track is None:
        raise ValueError(f"Root XML non riconosciuta per {xml_path}: {root_name}")
    return track, root_name


def iter_assistenza(root: ET.Element) -> Iterable[ET.Element]:
    for child in root:
        if local_name(child.tag) == "Assistenza":
            yield child


def find_child(parent: ET.Element, name: str) -> ET.Element | None:
    for child in parent:
        if local_name(child.tag) == name:
            return child
    return None


def find_text(parent: ET.Element | None, path: list[str]) -> str | None:
    current = parent
    for name in path:
        if current is None:
            return None
        current = find_child(current, name)
    return current.text.strip() if current is not None and current.text is not None else None


def extract_cf_from_id_rec(id_rec: str) -> str | None:
    if not id_rec:
        return None
    candidate = id_rec.strip()[-16:].upper()
    return candidate if CF_PATTERN.match(candidate) else None


def infer_birth_year_from_cf(cf: str) -> int | None:
    yy = int(cf[6:8])
    month = MONTH_CODES.get(cf[8])
    day_code = int(cf[9:11])
    if month is None:
        return None
    day = day_code - FEMALE_DAY_OFFSET if day_code > FEMALE_DAY_OFFSET else day_code
    try:
        date(2000 + yy, month, day)
    except ValueError:
        return None
    return yy


def full_year_from_two_digits(two_digits: int, century: int) -> int:
    return century + two_digits


def age_on_reference_date(cf: str, birth_year: int, reference_date: date) -> int | None:
    yy = int(cf[6:8])
    month = MONTH_CODES.get(cf[8])
    day_code = int(cf[9:11])
    if month is None:
        return None
    day = day_code - FEMALE_DAY_OFFSET if day_code > FEMALE_DAY_OFFSET else day_code
    if birth_year % 100 != yy:
        return None
    try:
        born = date(birth_year, month, day)
    except ValueError:
        return None
    years = reference_date.year - born.year
    if (reference_date.month, reference_date.day) < (born.month, born.day):
        years -= 1
    return years


def resolve_patients_birth_years(track1_rows: list[dict], track2_rows: list[dict], analysis_year: int) -> dict[str, PatientInfo]:
    patients: dict[str, PatientInfo] = {}
    for row in track1_rows + track2_rows:
        cf = row["codice_fiscale"]
        if cf not in patients:
            patients[cf] = PatientInfo(codice_fiscale=cf, guessed_year=infer_birth_year_from_cf(cf))

    for row in track1_rows:
        cf = row["codice_fiscale"]
        anno = row["anno_nascita"]
        patient = patients[cf]
        if anno is not None:
            patient.resolved_year = anno
            patient.resolved_from_track1 = True

    for patient in patients.values():
        if patient.resolved_year is not None:
            patient.ambiguous = False
            continue
        yy = patient.guessed_year
        if yy is None:
            patient.ambiguous = True
            continue
        if yy <= analysis_year % 100:
            patient.ambiguous = True
        else:
            patient.resolved_year = full_year_from_two_digits(yy, 1900)
            patient.ambiguous = False

    return patients


def parse_track1_assistenza(assistenza: ET.Element, source_file: str) -> dict | None:
    sede = find_text(assistenza, ["Erogatore", "CodiceASL"])
    id_rec = find_text(assistenza, ["Eventi", "PresaInCarico", "Id_Rec"])
    presa_date = parse_iso_date(find_child(find_child(assistenza, "Eventi"), "PresaInCarico").attrib.get("data") if find_child(find_child(assistenza, "Eventi"), "PresaInCarico") is not None else None)
    anno_nascita_text = find_text(assistenza, ["Assistito", "DatiAnagrafici", "AnnoNascita"])
    cf = extract_cf_from_id_rec(id_rec or "")
    if not sede or not id_rec or not cf or presa_date is None:
        return None
    return {
        "track": 1,
        "sede": sede,
        "source_file": source_file,
        "presa_in_carico_date": presa_date,
        "id_rec": id_rec,
        "codice_fiscale": cf,
        "anno_nascita": int(anno_nascita_text) if anno_nascita_text and anno_nascita_text.isdigit() else None,
    }


def parse_track2_assistenza(assistenza: ET.Element, source_file: str) -> dict | None:
    sede = find_text(assistenza, ["Erogatore", "CodiceASL"])
    eventi = find_child(assistenza, "Eventi")
    presa = find_child(eventi, "PresaInCarico") if eventi is not None else None
    id_rec = find_text(assistenza, ["Eventi", "PresaInCarico", "Id_Rec"])
    presa_date = parse_iso_date(presa.attrib.get("data") if presa is not None else None)
    cf = extract_cf_from_id_rec(id_rec or "")
    if not sede or not id_rec or not cf or presa_date is None:
        return None
    return {
        "track": 2,
        "sede": sede,
        "source_file": source_file,
        "presa_in_carico_date": presa_date,
        "id_rec": id_rec,
        "codice_fiscale": cf,
        "anno_nascita": None,
    }


def scan_xml_files(base_dir: Path, root_to_track: dict[str, int], filename_token: str = "SIAD") -> list[XmlFileInfo]:
    results: list[XmlFileInfo] = []
    token = filename_token.strip()
    pattern = f"{token}*.xml" if token else "*.xml"
    for path in sorted(base_dir.rglob(pattern)):
        if not path.is_file():
            continue
        track, root_name = classify_xml_file(path, root_to_track)
        results.append(
            XmlFileInfo(
                path=path,
                relative_path=path.relative_to(base_dir).as_posix(),
                track=track,
                root_name=root_name,
                line_count=count_file_lines(path),
                size_mb=file_size_mb(path),
            )
        )
    return results


def count_file_lines(path: Path) -> int:
    line_count = 0
    with path.open("r", encoding="utf-8", errors="ignore", newline=None) as file_obj:
        for _ in file_obj:
            line_count += 1
    if line_count > 0:
        return line_count
    return 1 if path.stat().st_size > 0 else 0


def file_size_mb(path: Path) -> float:
    return path.stat().st_size / (1024 * 1024)


def xml_namespace(tag: str) -> str | None:
    match = XML_NS_RE.match(tag)
    return match.group("ns") if match else None


def strip_namespaces(element: ET.Element) -> None:
    element.tag = local_name(element.tag)
    for child in element:
        strip_namespaces(child)


def collect_validation_errors(schema: Any, xml_source: Path) -> list[dict[str, str]]:
    errors: list[dict[str, str]] = []
    for error in schema.iter_errors(xml_source):
        errors.append(
            {
                "Path": getattr(error, "path", "") or "",
                "Messaggio": getattr(error, "reason", None) or str(error),
            }
        )
    return errors


def is_namespace_mismatch_error(errors: list[dict[str, str]]) -> bool:
    mismatch_markers = (
        "is not an element of the schema",
        "not an element of the schema",
    )
    return any(any(marker in error["Messaggio"] for marker in mismatch_markers) for error in errors)


def validate_xml_against_xsd(xml_path: Path, xsd_path: Path) -> list[dict[str, str]]:
    if xmlschema is None:
        raise RuntimeError(
            "La validazione XSD richiede il pacchetto 'xmlschema'. "
            "Installa le dipendenze aggiornate con: python -m pip install -r requirements.txt"
        )

    xml_root = ET.parse(xml_path).getroot()
    xml_target_namespace = xml_namespace(xml_root.tag)
    xsd_tree = ET.parse(xsd_path)
    xsd_root = xsd_tree.getroot()
    xsd_target_namespace = xsd_root.attrib.get("targetNamespace")

    try:
        schema = xmlschema.XMLSchema(xsd_path)
        direct_errors = collect_validation_errors(schema, xml_path)
        if not direct_errors or not xml_target_namespace or xsd_target_namespace or not is_namespace_mismatch_error(direct_errors):
            return direct_errors
    except Exception as first_error:
        if not xml_target_namespace or xsd_target_namespace:
            raise first_error

    with TemporaryDirectory() as temp_dir:
        temp_dir_path = Path(temp_dir)

        namespaced_xsd_tree = ET.parse(xsd_path)
        namespaced_xsd_root = namespaced_xsd_tree.getroot()
        namespaced_xsd_root.set("targetNamespace", xml_target_namespace)
        namespaced_xsd_root.set("xmlns", xml_target_namespace)
        namespaced_xsd_path = temp_dir_path / xsd_path.name
        namespaced_xsd_tree.write(namespaced_xsd_path, encoding="utf-8", xml_declaration=True)

        try:
            schema = xmlschema.XMLSchema(namespaced_xsd_path)
            namespaced_errors = collect_validation_errors(schema, xml_path)
            if not namespaced_errors or not is_namespace_mismatch_error(namespaced_errors):
                return namespaced_errors
        except Exception:
            pass

        stripped_xml_tree = ET.parse(xml_path)
        stripped_xml_root = stripped_xml_tree.getroot()
        strip_namespaces(stripped_xml_root)
        stripped_xml_path = temp_dir_path / xml_path.name
        stripped_xml_tree.write(stripped_xml_path, encoding="utf-8", xml_declaration=True)

        schema = xmlschema.XMLSchema(xsd_path)
        return collect_validation_errors(schema, stripped_xml_path)


def build_report(xml_files: list[XmlFileInfo], analysis_year: int) -> tuple[list[dict], list[RecordDetail], list[dict], int]:
    track1_rows: list[dict] = []
    track2_rows: list[dict] = []

    for file_info in xml_files:
        root = ET.parse(file_info.path).getroot()
        for assistenza in iter_assistenza(root):
            row = (
                parse_track1_assistenza(assistenza, file_info.relative_path)
                if file_info.track == 1
                else parse_track2_assistenza(assistenza, file_info.relative_path)
            )
            if row is None:
                continue
            if row["track"] == 1:
                track1_rows.append(row)
            else:
                track2_rows.append(row)

    patients = resolve_patients_birth_years(track1_rows, track2_rows, analysis_year)
    reference_date = date(analysis_year, 12, 31)
    previous_active_date = date(analysis_year, 1, 1)

    by_sede_track1: dict[str, list[dict]] = defaultdict(list)
    by_sede_track2_prev: dict[str, list[dict]] = defaultdict(list)
    by_sede_track2_late: dict[str, list[dict]] = defaultdict(list)
    details: list[RecordDetail] = []
    seen_track_keys: set[tuple[int, str, str]] = set()
    counted_unique_cfs_by_sede: dict[str, set[str]] = defaultdict(set)

    track2_late_candidates: list[dict] = []

    def patient_age_and_flag(cf: str) -> tuple[int | None, bool]:
        patient = patients[cf]
        age = age_on_reference_date(cf, patient.resolved_year, reference_date) if patient.resolved_year else None
        return age, bool(age is not None and age >= 65)

    for row in track1_rows:
        if row["presa_in_carico_date"].year == analysis_year:
            active_key = (1, row["sede"], row["id_rec"])
            if active_key not in seen_track_keys:
                by_sede_track1[row["sede"]].append(row)
                seen_track_keys.add(active_key)
                counted_unique_cfs_by_sede[row["sede"]].add(row["codice_fiscale"])
                included = True
                note = "Nuova presa in carico dell'anno di analisi."
            else:
                included = False
                note = "Escluso: duplicato dello stesso Id_Rec gia' contato."
        else:
            included = False
            note = "Escluso: presa in carico tracciato 1 fuori anno di analisi."
        patient = patients[row["codice_fiscale"]]
        age, is_over_65 = patient_age_and_flag(row["codice_fiscale"])
        details.append(
            RecordDetail(
                sede=row["sede"],
                track=1,
                source_file=row["source_file"],
                quarter=quarter_from_path(row["source_file"]),
                presa_in_carico_date=row["presa_in_carico_date"].isoformat(),
                id_rec=row["id_rec"],
                codice_fiscale=row["codice_fiscale"],
                anno_nascita_xml=row["anno_nascita"],
                anno_nascita_usato=patient.resolved_year,
                eta_al_31_12=age,
                is_over_65=is_over_65,
                cf_ambiguo=patient.ambiguous,
                included_in_report=included,
                note=note,
            )
        )

    for row in track2_rows:
        if row["presa_in_carico_date"] < previous_active_date:
            active_key = (2, row["sede"], row["id_rec"])
            if active_key not in seen_track_keys:
                by_sede_track2_prev[row["sede"]].append(row)
                seen_track_keys.add(active_key)
                counted_unique_cfs_by_sede[row["sede"]].add(row["codice_fiscale"])
                included = True
                note = "Presa in carico precedente considerata attiva al 01/01 dell'anno di analisi."
            else:
                included = False
                note = "Escluso: duplicato dello stesso Id_Rec gia' contato."
        else:
            track2_late_candidates.append(row)
            included = False
            note = "Da valutare: presa in carico tracciato 2 nell'anno di analisi."
        patient = patients[row["codice_fiscale"]]
        age, is_over_65 = patient_age_and_flag(row["codice_fiscale"])
        details.append(
            RecordDetail(
                sede=row["sede"],
                track=2,
                source_file=row["source_file"],
                quarter=quarter_from_path(row["source_file"]),
                presa_in_carico_date=row["presa_in_carico_date"].isoformat(),
                id_rec=row["id_rec"],
                codice_fiscale=row["codice_fiscale"],
                anno_nascita_xml=None,
                anno_nascita_usato=patient.resolved_year,
                eta_al_31_12=age,
                is_over_65=is_over_65,
                cf_ambiguo=patient.ambiguous,
                included_in_report=included,
                note=note,
            )
        )

    detail_index_by_key = {
        (detail.track, detail.sede, detail.id_rec, detail.source_file): idx for idx, detail in enumerate(details)
    }

    for row in track2_late_candidates:
        detail_key = (2, row["sede"], row["id_rec"], row["source_file"])
        detail_idx = detail_index_by_key[detail_key]
        active_key = (2, row["sede"], row["id_rec"])
        if active_key in seen_track_keys:
            details[detail_idx].included_in_report = False
            details[detail_idx].note = "Escluso: duplicato dello stesso Id_Rec gia' contato."
            continue
        if row["codice_fiscale"] in counted_unique_cfs_by_sede[row["sede"]]:
            details[detail_idx].included_in_report = False
            details[detail_idx].note = (
                "Escluso dal conteggio aggiuntivo: CF gia' presente tra i CF univoci conteggiati per la stessa azienda nei tracciati 1 e 2."
            )
            continue
        by_sede_track2_late[row["sede"]].append(row)
        seen_track_keys.add(active_key)
        counted_unique_cfs_by_sede[row["sede"]].add(row["codice_fiscale"])
        details[detail_idx].included_in_report = True
        details[detail_idx].note = (
            "Incluso: presa in carico tracciato 2 nell'anno con CF non ancora presente tra i CF univoci conteggiati per la stessa azienda."
        )

    summary_rows: list[dict] = []
    sedi = sorted(set(by_sede_track1) | set(by_sede_track2_prev) | set(by_sede_track2_late))
    active_unique_cfs_by_sede: dict[str, set[str]] = {}

    for sede in sedi:
        prev_rows = by_sede_track2_prev.get(sede, [])
        new_rows = by_sede_track1.get(sede, [])
        late_rows = by_sede_track2_late.get(sede, [])
        active_rows = prev_rows + new_rows + late_rows
        unique_patients = {row["codice_fiscale"] for row in active_rows}
        active_unique_cfs_by_sede[sede] = unique_patients

    cfs_to_sedi: dict[str, set[str]] = defaultdict(set)
    for sede, unique_cfs in active_unique_cfs_by_sede.items():
        for cf in unique_cfs:
            cfs_to_sedi[cf].add(sede)

    global_single_heads = len(cfs_to_sedi)

    def cf_stats(cf_set: set[str]) -> tuple[int, int, int]:
        over_65 = 0
        ambiguous_count = 0
        for cf in cf_set:
            patient = patients[cf]
            if patient.ambiguous:
                ambiguous_count += 1
            age, is_over_65 = patient_age_and_flag(cf)
            if is_over_65:
                over_65 += 1
        return len(cf_set), over_65, ambiguous_count

    for sede in sedi:
        prev_rows = by_sede_track2_prev.get(sede, [])
        new_rows = by_sede_track1.get(sede, [])
        late_rows = by_sede_track2_late.get(sede, [])
        active_rows = prev_rows + new_rows + late_rows
        unique_patients = active_unique_cfs_by_sede[sede]
        shared_cfs = {cf for cf in unique_patients if len(cfs_to_sedi[cf]) > 1}
        exclusive_cfs = unique_patients - shared_cfs
        total_unique_count, total_over_65, total_ambiguous = cf_stats(unique_patients)
        exclusive_count, exclusive_over_65, exclusive_ambiguous = cf_stats(exclusive_cfs)
        shared_count, shared_over_65, shared_ambiguous = cf_stats(shared_cfs)
        summary_rows.append(
            {
                "SEDE": sede,
                f"Prese in carico precedenti ancora attive al 01/01/{analysis_year}": len(prev_rows),
                f"Nuove Prese in carico {analysis_year}": len(new_rows),
                f"Prese in carico T2 {analysis_year} con CF non ancora presente": len(late_rows),
                f"TOT. PRESE IN CARICO attive nel {analysis_year}": len(active_rows),
                f"TOT. CF non univoci attivi nel {analysis_year}": len(active_rows),
                f"[CF per azienda] TOT. PAZIENTI* attivi nel {analysis_year}": total_unique_count,
                "[CF per azienda] di cui >= 65 anni": total_over_65,
                "[CF per azienda] numero cf ambigui": total_ambiguous,
                "[Teste singole globali] CF esclusivi azienda": exclusive_count,
                "[Teste singole globali] di cui >= 65 anni": exclusive_over_65,
                "[Teste singole globali] numero cf ambigui": exclusive_ambiguous,
                "[Differenze] CF condivisi con altre aziende": shared_count,
                "[Differenze] di cui >= 65 anni": shared_over_65,
                "[Differenze] numero cf ambigui": shared_ambiguous,
            }
        )

    details.sort(key=lambda item: (item.sede, item.track, item.source_file, item.id_rec))
    unique_cf_rows: list[dict] = []
    rows_for_cf_sheet = track1_rows + track2_rows
    rows_by_cf_sede: dict[tuple[str, str], list[dict]] = defaultdict(list)
    included_occurrences_by_cf_sede: dict[tuple[str, str], int] = defaultdict(int)
    for row in rows_for_cf_sheet:
        rows_by_cf_sede[(row["sede"], row["codice_fiscale"])].append(row)
    for detail in details:
        if detail.included_in_report:
            included_occurrences_by_cf_sede[(detail.sede, detail.codice_fiscale)] += 1

    for sede_cf in sorted(rows_by_cf_sede):
        sede, cf = sede_cf
        cf_rows = rows_by_cf_sede[sede_cf]
        patient = patients[cf]
        quarters = sorted({quarter_from_path(row["source_file"]) for row in cf_rows if quarter_from_path(row["source_file"])})
        tracks = sorted({str(row["track"]) for row in cf_rows})
        unique_cf_rows.append(
            {
                "Azienda di competenza": sede,
                "Codice fiscale": cf,
                "Tracciati": ", ".join(tracks),
                "Trimestri riscontrati": ", ".join(quarters),
                "Numero trimestri": len(quarters),
                "Numero occorrenze": len(cf_rows),
                "Numero occorrenze incluse": included_occurrences_by_cf_sede.get(sede_cf, 0),
                "Anno nascita usato": patient.resolved_year,
                "Eta al 31/12": age_on_reference_date(cf, patient.resolved_year, reference_date) if patient.resolved_year else None,
                "CF ambiguo": "SI" if patient.ambiguous else "NO",
            }
        )

    return summary_rows, details, unique_cf_rows, global_single_heads


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)


def add_total_row(summary_rows: list[dict], global_single_heads: int | None = None) -> list[dict]:
    if not summary_rows:
        return []
    headers = list(summary_rows[0].keys())
    total_row: dict[str, object] = {headers[0]: "TOTALE"}
    for header in headers[1:]:
        total_row[header] = sum(int(row[header]) for row in summary_rows)
    return [*summary_rows, total_row]


def detail_headers() -> list[str]:
    return [
        "Azienda di competenza",
        "SEDE",
        "Tracciato",
        "Trimestre",
        "File XML",
        "Data presa in carico",
        "Id_Rec",
        "Codice fiscale",
        "AnnoNascita XML",
        "Anno nascita usato",
        "Eta al 31/12",
        "IsOver65",
        "CF ambiguo",
        "Incluso nel report",
        "Note",
    ]


def detail_to_row(detail: RecordDetail) -> list[object]:
    return [
        detail.sede,
        detail.sede,
        detail.track,
        detail.quarter,
        detail.source_file,
        detail.presa_in_carico_date,
        detail.id_rec,
        detail.codice_fiscale,
        detail.anno_nascita_xml,
        detail.anno_nascita_usato,
        detail.eta_al_31_12,
        "SI" if detail.is_over_65 else "NO",
        "SI" if detail.cf_ambiguo else "NO",
        "SI" if detail.included_in_report else "NO",
        detail.note,
    ]


def make_sheet_title(prefix: str, sede: str) -> str:
    raw = f"{prefix}_{sede}"
    sanitized = re.sub(r"[\[\]\*\?/\\:]", "_", raw)
    return sanitized[:31]


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL_ODD = PatternFill(fill_type="solid", fgColor="F7FBFF")
ALT_FILL_EVEN = PatternFill(fill_type="solid", fgColor="EAF2F8")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
TOTAL_FONT = Font(bold=True, color="1F1F1F")
CENTER_ALIGNMENT = Alignment(vertical="center")
THICK_SIDE = Side(style="thick", color="1F1F1F")
GROUP_FILLS = {
    "attivita": PatternFill(fill_type="solid", fgColor="3D6D99"),
    "cf_azienda": PatternFill(fill_type="solid", fgColor="2E8B57"),
    "teste_globali": PatternFill(fill_type="solid", fgColor="8C6D1F"),
    "differenze": PatternFill(fill_type="solid", fgColor="7A3E9D"),
}


def summary_column_groups(headers: list[str]) -> list[tuple[int, int, str]]:
    groups: list[tuple[int, int, str]] = []
    current_group = "attivita"
    start = 1
    for idx, header in enumerate(headers, start=1):
        if header.startswith("[CF per azienda]"):
            group = "cf_azienda"
        elif header.startswith("[Teste singole globali]"):
            group = "teste_globali"
        elif header.startswith("[Differenze]"):
            group = "differenze"
        else:
            group = "attivita"
        if idx == 1:
            current_group = group
            start = idx
            continue
        if group != current_group:
            groups.append((start, idx - 1, current_group))
            start = idx
            current_group = group
    groups.append((start, len(headers), current_group))
    return groups


def apply_summary_group_style(ws) -> None:
    headers = [cell.value for cell in ws[1]]
    groups = summary_column_groups(headers)

    for start_col, end_col, group_name in groups:
        fill = GROUP_FILLS[group_name]
        for col_idx in range(start_col, end_col + 1):
            ws.cell(row=1, column=col_idx).fill = fill
        for row_idx in range(1, ws.max_row + 1):
            left_cell = ws.cell(row=row_idx, column=start_col)
            right_cell = ws.cell(row=row_idx, column=end_col)
            left_cell.border = Border(
                left=THICK_SIDE,
                right=left_cell.border.right,
                top=left_cell.border.top,
                bottom=left_cell.border.bottom,
            )
            right_cell.border = Border(
                left=right_cell.border.left,
                right=THICK_SIDE,
                top=right_cell.border.top,
                bottom=right_cell.border.bottom,
            )


def style_worksheet(ws, has_total_row: bool = False) -> None:
    max_row = ws.max_row
    max_col = ws.max_column

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT

    data_end_row = max_row - 1 if has_total_row and max_row >= 2 else max_row
    for row_idx in range(2, data_end_row + 1):
        fill = ALT_FILL_ODD if row_idx % 2 == 0 else ALT_FILL_EVEN
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = CENTER_ALIGNMENT

    if has_total_row and max_row >= 2:
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=max_row, column=col_idx)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.alignment = CENTER_ALIGNMENT

    autosize_columns(ws)


def write_table_sheet(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def build_excluded_cf_rows(details: list[RecordDetail]) -> list[dict]:
    details_by_key: dict[tuple[str, str], list[RecordDetail]] = defaultdict(list)
    for detail in details:
        key = (detail.sede, detail.codice_fiscale)
        details_by_key[key].append(detail)

    rows: list[dict] = []
    for key in sorted(details_by_key):
        grouped_details = details_by_key[key]
        if any(detail.included_in_report for detail in grouped_details):
            continue

        non_duplicate_details = [
            detail for detail in grouped_details if "duplicato dello stesso Id_Rec" not in detail.note
        ]
        if not non_duplicate_details:
            continue

        sede, cf = key
        rows.append(
            {
                "Azienda di competenza": sede,
                "Codice fiscale": cf,
                "Tracciati": ", ".join(sorted({str(detail.track) for detail in non_duplicate_details})),
                "Trimestri riscontrati": ", ".join(sorted({detail.quarter for detail in non_duplicate_details if detail.quarter})),
                "Numero occorrenze escluse": len(non_duplicate_details),
                "Motivi di esclusione": " | ".join(sorted({detail.note for detail in non_duplicate_details})),
            }
        )
    return rows


def save_workbook(
    output_path: Path,
    summary_rows: list[dict],
    details: list[RecordDetail],
    unique_cf_rows: list[dict],
    global_single_heads: int,
) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Report"

    summary_with_total = add_total_row(summary_rows, global_single_heads)
    if summary_with_total:
        headers = list(summary_with_total[0].keys())
    else:
        headers = [
            "SEDE",
            "Prese in carico precedenti ancora attive al 01/01/ANNO",
            "Nuove Prese in carico ANNO",
            "Prese in carico T2 ANNO con CF non ancora presente",
            "TOT. PRESE IN CARICO attive nel ANNO",
            "TOT. CF non univoci attivi nel ANNO",
            "[CF per azienda] TOT. PAZIENTI* attivi nel ANNO",
            "[CF per azienda] di cui >= 65 anni",
            "[CF per azienda] numero cf ambigui",
            "[Teste singole globali] CF esclusivi azienda",
            "[Teste singole globali] di cui >= 65 anni",
            "[Teste singole globali] numero cf ambigui",
            "[Differenze] CF condivisi con altre aziende",
            "[Differenze] di cui >= 65 anni",
            "[Differenze] numero cf ambigui",
        ]

    write_table_sheet(ws_summary, headers, [[row[h] for h in headers] for row in summary_with_total])

    ws_details = wb.create_sheet("Dettaglio")
    detail_header_row = detail_headers()
    write_table_sheet(ws_details, detail_header_row, [detail_to_row(detail) for detail in details])

    details_by_sede: dict[str, list[RecordDetail]] = defaultdict(list)
    for detail in details:
        details_by_sede[detail.sede].append(detail)

    detail_sheets = [ws_details]
    for sede in sorted(details_by_sede):
        ws_sede = wb.create_sheet(make_sheet_title("Dettaglio", sede))
        write_table_sheet(ws_sede, detail_header_row, [detail_to_row(detail) for detail in details_by_sede[sede]])
        detail_sheets.append(ws_sede)

    ws_unique_cf = wb.create_sheet("CF_Univoci")
    if unique_cf_rows:
        cf_headers = list(unique_cf_rows[0].keys())
        cf_rows = [[row[h] for h in cf_headers] for row in unique_cf_rows]
    else:
        cf_headers = [
            "Azienda di competenza",
            "Codice fiscale",
            "Tracciati",
            "Trimestri riscontrati",
            "Numero trimestri",
            "Numero occorrenze",
            "Numero occorrenze incluse",
            "Anno nascita usato",
            "Eta al 31/12",
            "CF ambiguo",
        ]
        cf_rows = []
    write_table_sheet(ws_unique_cf, cf_headers, cf_rows)

    excluded_cf_rows = build_excluded_cf_rows(details)
    ws_excluded_cf = wb.create_sheet("CF_Esclusi")
    if excluded_cf_rows:
        excluded_headers = list(excluded_cf_rows[0].keys())
        excluded_rows = [[row[h] for h in excluded_headers] for row in excluded_cf_rows]
    else:
        excluded_headers = [
            "Azienda di competenza",
            "Codice fiscale",
            "Tracciati",
            "Trimestri riscontrati",
            "Numero occorrenze escluse",
            "Motivi di esclusione",
        ]
        excluded_rows = []
    write_table_sheet(ws_excluded_cf, excluded_headers, excluded_rows)

    style_worksheet(ws_summary, has_total_row=bool(summary_rows))
    apply_summary_group_style(ws_summary)
    for ws in [*detail_sheets, ws_unique_cf, ws_excluded_cf]:
        style_worksheet(ws)

    wb.save(output_path)


class SiadReportApp:
    SPINNER_FRAMES = ["[...]", "[ ..]", "[  .]", "[ ..]"]
    CHECKED_SYMBOL = "[x]"
    UNCHECKED_SYMBOL = "[ ]"
    STATE_FILE_NAME = ".siad_report_gui_state.json"

    def __init__(
        self,
        root: Any,
        tk_module: Any,
        ttk_module: Any,
        filedialog_module: Any,
        messagebox_module: Any,
        *,
        parent: Any | None = None,
        embed_mode: bool = False,
    ) -> None:
        self.root = root
        self.parent = parent or root
        self.embed_mode = embed_mode
        self.tk = tk_module
        self.ttk = ttk_module
        self.filedialog = filedialog_module
        self.messagebox = messagebox_module
        if not self.embed_mode:
            self.root.title("Report SIAD")
            self.root.geometry("1100x700")

        self.track1_xsd_var = self.tk.StringVar()
        self.track2_xsd_var = self.tk.StringVar()
        self.xml_dir_var = self.tk.StringVar()
        self.output_var = self.tk.StringVar()
        self.analysis_year_var = self.tk.StringVar(value="2025")
        self.filename_token_var = self.tk.StringVar(value="SIAD")
        self.status_var = self.tk.StringVar(value="Seleziona i file XSD, la cartella XML e il file di output.")
        self.spinner_var = self.tk.StringVar(value="")
        self.select_all_var = self.tk.BooleanVar(value=False)

        self.xml_files: list[XmlFileInfo] = []
        self.summary_rows: list[dict] = []
        self.validation_rows: list[dict[str, str]] = []
        self.checked_files: dict[tuple[int, str], bool] = {}
        self.latest_report_path: Path | None = None
        self._busy_count = 0
        self._spinner_job: str | None = None
        self._spinner_index = 0
        self._syncing_select_all = False

        self._build_ui()
        self.root.after(0, self.prompt_restore_saved_paths)

    def _build_ui(self) -> None:
        frame = self.ttk.Frame(self.parent, padding=12)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(5, weight=1)

        self.validate_track1_button = self._add_file_row(
            frame,
            0,
            "TRACCIATO 1 XSD",
            self.track1_xsd_var,
            self.choose_track1_xsd,
            extra_button_text="Valida XML...",
            extra_button_command=self.validate_xml_with_track1_xsd,
        )
        self.validate_track2_button = self._add_file_row(
            frame,
            1,
            "TRACCIATO 2 XSD",
            self.track2_xsd_var,
            self.choose_track2_xsd,
            extra_button_text="Valida XML...",
            extra_button_command=self.validate_xml_with_track2_xsd,
        )
        self._add_file_row(frame, 2, "Cartella XML", self.xml_dir_var, self.choose_xml_dir, folder=True)
        self._add_file_row(frame, 3, "Output XLSX", self.output_var, self.choose_output)
        self.validate_track1_button.grid_remove()
        self.validate_track2_button.grid_remove()

        options_row = self.ttk.Frame(frame)
        options_row.grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 8))
        self.ttk.Label(options_row, text="Anno di analisi").pack(side="left")
        self.ttk.Entry(options_row, textvariable=self.analysis_year_var, width=12).pack(side="left", padx=(8, 24))
        self.ttk.Label(options_row, text="Token file XML").pack(side="left")
        token_combo = self.ttk.Combobox(
            options_row,
            textvariable=self.filename_token_var,
            values=("SIAD",),
            width=12,
            state="readonly",
        )
        token_combo.pack(side="left", padx=(8, 0))

        buttons = self.ttk.Frame(frame)
        buttons.grid(row=4, column=2, sticky="e", pady=(8, 8))
        self.ttk.Button(buttons, text="Scansiona XML", command=self.scan_files).pack(side="left", padx=(0, 8))
        self.ttk.Button(buttons, text="Genera report", command=self.generate_report).pack(side="left")

        self.notebook = self.ttk.Notebook(frame)
        self.notebook.grid(row=5, column=0, columnspan=4, sticky="nsew")

        files_tab = self.ttk.Frame(self.notebook, padding=8)
        summary_tab = self.ttk.Frame(self.notebook, padding=8)
        validation_tab = self.ttk.Frame(self.notebook, padding=8)
        self.notebook.add(files_tab, text="File XML")
        self.notebook.add(summary_tab, text="Riepilogo")
        self.notebook.add(validation_tab, text="Validazione")

        files_tab.columnconfigure(0, weight=1)
        files_tab.rowconfigure(1, weight=1)
        summary_tab.columnconfigure(0, weight=1)
        summary_tab.rowconfigure(0, weight=1)
        validation_tab.columnconfigure(0, weight=1)
        validation_tab.rowconfigure(0, weight=1)

        self.ttk.Checkbutton(
            files_tab,
            text="Seleziona tutti",
            variable=self.select_all_var,
            command=self.toggle_all_files,
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        file_columns = ("row_number", "checked", "track", "line_count", "size_mb", "relative_path")
        self.file_tree = self.ttk.Treeview(files_tab, columns=file_columns, show="headings", height=20)
        self.file_tree.heading("row_number", text="#")
        self.file_tree.heading("checked", text="Sel")
        self.file_tree.heading("track", text="Tracciato")
        self.file_tree.heading("line_count", text="Righe")
        self.file_tree.heading("size_mb", text="MB")
        self.file_tree.heading("relative_path", text="File XML")
        self.file_tree.column("row_number", width=60, anchor="center", stretch=False)
        self.file_tree.column("checked", width=60, anchor="center", stretch=False)
        self.file_tree.column("track", width=100, anchor="center", stretch=False)
        self.file_tree.column("line_count", width=90, anchor="e", stretch=False)
        self.file_tree.column("size_mb", width=90, anchor="e", stretch=False)
        self.file_tree.column("relative_path", width=600, anchor="w")
        self.file_tree.grid(row=1, column=0, sticky="nsew")
        self.file_tree.bind("<ButtonRelease-1>", self.on_file_clicked)

        file_scrollbar = self.ttk.Scrollbar(files_tab, orient="vertical", command=self.file_tree.yview)
        file_scrollbar.grid(row=1, column=1, sticky="ns")
        self.file_tree.configure(yscrollcommand=file_scrollbar.set)

        self.summary_tree = self.ttk.Treeview(summary_tab, show="headings", height=20)
        self.summary_tree.grid(row=0, column=0, sticky="nsew")
        summary_scrollbar = self.ttk.Scrollbar(summary_tab, orient="vertical", command=self.summary_tree.yview)
        summary_scrollbar.grid(row=0, column=1, sticky="ns")
        summary_x_scrollbar = self.ttk.Scrollbar(summary_tab, orient="horizontal", command=self.summary_tree.xview)
        summary_x_scrollbar.grid(row=1, column=0, sticky="ew")
        self.summary_tree.configure(yscrollcommand=summary_scrollbar.set, xscrollcommand=summary_x_scrollbar.set)

        validation_columns = ("file_xml", "path", "messaggio")
        self.validation_tree = self.ttk.Treeview(validation_tab, columns=validation_columns, show="headings", height=20)
        self.validation_tree.heading("file_xml", text="File XML")
        self.validation_tree.heading("path", text="Path")
        self.validation_tree.heading("messaggio", text="Errore")
        self.validation_tree.column("file_xml", width=320, anchor="w")
        self.validation_tree.column("path", width=220, anchor="w")
        self.validation_tree.column("messaggio", width=700, anchor="w")
        self.validation_tree.grid(row=0, column=0, sticky="nsew")

        validation_scrollbar = self.ttk.Scrollbar(validation_tab, orient="vertical", command=self.validation_tree.yview)
        validation_scrollbar.grid(row=0, column=1, sticky="ns")
        self.validation_tree.configure(yscrollcommand=validation_scrollbar.set)

        status_bar = self.ttk.Frame(frame)
        status_bar.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(10, 0))
        status_bar.columnconfigure(1, weight=1)
        self.ttk.Label(status_bar, textvariable=self.spinner_var, width=5, anchor="w").grid(row=0, column=0, sticky="w")
        self.status_label = self.ttk.Label(status_bar, textvariable=self.status_var)
        self.status_label.grid(row=0, column=1, sticky="w")
        self.status_label.bind("<Button-1>", self.on_status_label_clicked)

    def _add_file_row(
        self,
        parent,
        row: int,
        label: str,
        variable: Any,
        command,
        folder: bool = False,
        extra_button_text: str | None = None,
        extra_button_command=None,
    ):
        self.ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=4)
        self.ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=(8, 8), pady=4)
        self.ttk.Button(parent, text="Sfoglia...", command=command).grid(row=row, column=2, sticky="e", pady=4)
        extra_button = None
        if extra_button_text and extra_button_command:
            extra_button = self.ttk.Button(parent, text=extra_button_text, command=extra_button_command)
            extra_button.grid(row=row, column=3, sticky="e", pady=4, padx=(8, 0))
        return extra_button

    def choose_track1_xsd(self) -> None:
        path = self.filedialog.askopenfilename(filetypes=[("XSD files", "*.xsd"), ("All files", "*.*")])
        if path:
            self.track1_xsd_var.set(path)
            self.validate_track1_button.grid()
            self.save_current_paths()

    def choose_track2_xsd(self) -> None:
        path = self.filedialog.askopenfilename(filetypes=[("XSD files", "*.xsd"), ("All files", "*.*")])
        if path:
            self.track2_xsd_var.set(path)
            self.validate_track2_button.grid()
            self.save_current_paths()

    def choose_xml_dir(self) -> None:
        path = self.filedialog.askdirectory()
        if path:
            self.xml_dir_var.set(path)
            self.refresh_output_from_xml_dir()
            self.save_current_paths()

    def choose_output(self) -> None:
        path = self.filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile="report_siad.xlsx",
            confirmoverwrite=False,
        )
        if path:
            self.output_var.set(path)

    def validate_inputs(self) -> tuple[Path, Path, Path, Path, int] | None:
        try:
            year = int(self.analysis_year_var.get().strip())
            if year < 2000 or year > 2100:
                raise ValueError
        except ValueError:
            self.messagebox.showerror("Errore", "L'anno di analisi deve essere un intero valido.")
            return None

        paths = [
            Path(self.track1_xsd_var.get().strip()),
            Path(self.track2_xsd_var.get().strip()),
            Path(self.xml_dir_var.get().strip()),
            Path(self.output_var.get().strip()),
        ]
        if any(not str(path) for path in paths):
            self.messagebox.showerror("Errore", "Compila tutti i campi richiesti.")
            return None

        if not paths[0].is_file() or not paths[1].is_file():
            self.messagebox.showerror("Errore", "I file XSD selezionati non esistono.")
            return None
        if not paths[2].is_dir():
            self.messagebox.showerror("Errore", "La cartella XML selezionata non esiste.")
            return None

        if not self.filename_token_var.get().strip():
            self.messagebox.showerror("Errore", "Seleziona un token valido per il riconoscimento dei file XML.")
            return None

        return paths[0], paths[1], paths[2], paths[3], year

    def start_busy_indicator(self) -> None:
        self._busy_count += 1
        if self._spinner_job is None:
            self._spinner_index = 0
            self._animate_spinner()

    def stop_busy_indicator(self) -> None:
        self._busy_count = max(0, self._busy_count - 1)
        if self._busy_count == 0:
            if self._spinner_job is not None:
                self.root.after_cancel(self._spinner_job)
                self._spinner_job = None
            self.spinner_var.set("")

    def _animate_spinner(self) -> None:
        self.spinner_var.set(self.SPINNER_FRAMES[self._spinner_index % len(self.SPINNER_FRAMES)])
        self._spinner_index += 1
        self._spinner_job = self.root.after(250, self._animate_spinner)

    def state_file_path(self) -> Path:
        return Path(__file__).with_name(self.STATE_FILE_NAME)

    def load_saved_paths(self) -> dict[str, str]:
        state_path = self.state_file_path()
        if not state_path.is_file():
            return {}
        try:
            data = json.loads(state_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        return data if isinstance(data, dict) else {}

    def save_current_paths(self) -> None:
        state = {
            "track1_xsd": self.track1_xsd_var.get().strip(),
            "track2_xsd": self.track2_xsd_var.get().strip(),
            "xml_dir": self.xml_dir_var.get().strip(),
        }
        try:
            self.state_file_path().write_text(json.dumps(state, indent=2), encoding="utf-8")
        except OSError:
            pass

    def prompt_restore_saved_paths(self) -> None:
        state = self.load_saved_paths()
        self.prompt_restore_single_path(
            "VUoi ricaricare il path del TRACCIATO1?",
            state.get("track1_xsd", ""),
            self.track1_xsd_var,
            is_file=True,
            on_accept=lambda: self.validate_track1_button.grid(),
            on_decline=lambda: self._clear_track_path(self.track1_xsd_var, self.validate_track1_button),
        )
        self.prompt_restore_single_path(
            "VUoi ricaricare il path del TRACCIATO2?",
            state.get("track2_xsd", ""),
            self.track2_xsd_var,
            is_file=True,
            on_accept=lambda: self.validate_track2_button.grid(),
            on_decline=lambda: self._clear_track_path(self.track2_xsd_var, self.validate_track2_button),
        )
        self.prompt_restore_single_path(
            "VUoi ricaricare il path della CARTELLA?",
            state.get("xml_dir", ""),
            self.xml_dir_var,
            is_file=False,
            on_accept=self.refresh_output_from_xml_dir,
            on_decline=self._clear_xml_dir_path,
        )

    def prompt_restore_single_path(
        self,
        question: str,
        saved_path: str,
        variable: Any,
        *,
        is_file: bool,
        on_accept=None,
        on_decline=None,
    ) -> None:
        if not saved_path:
            return
        candidate = Path(saved_path)
        exists = candidate.is_file() if is_file else candidate.is_dir()
        if not exists:
            return
        should_restore = self.messagebox.askyesno("Ripristino path", f"{question}\n\n{saved_path}")
        if should_restore:
            variable.set(saved_path)
            if on_accept is not None:
                on_accept()
        else:
            variable.set("")
            if on_decline is not None:
                on_decline()
        self.save_current_paths()

    def _clear_track_path(self, variable: Any, button: Any) -> None:
        variable.set("")
        button.grid_remove()

    def _clear_xml_dir_path(self) -> None:
        self.xml_dir_var.set("")
        self.output_var.set("")

    def refresh_output_from_xml_dir(self) -> None:
        xml_dir_value = self.xml_dir_var.get().strip()
        if not xml_dir_value:
            self.output_var.set("")
            return
        self.output_var.set(str(next_available_report_path(Path(xml_dir_value))))

    def validate_xml_with_track1_xsd(self) -> None:
        self.validate_xml_with_selected_xsd(1, Path(self.track1_xsd_var.get().strip()))

    def validate_xml_with_track2_xsd(self) -> None:
        self.validate_xml_with_selected_xsd(2, Path(self.track2_xsd_var.get().strip()))

    def validate_xml_with_selected_xsd(self, track: int, xsd_path: Path) -> None:
        if not xsd_path.is_file():
            self.messagebox.showerror("Errore", f"XSD del tracciato {track} non valido o non selezionato.")
            return
        xml_path = self.filedialog.askopenfilename(
            title=f"Seleziona XML da validare con XSD tracciato {track}",
            filetypes=[("XML files", "*.xml"), ("All files", "*.*")],
        )
        if not xml_path:
            return
        xml_file_info = XmlFileInfo(
            path=Path(xml_path),
            relative_path=Path(xml_path).name,
            track=track,
            root_name="",
        )
        self.set_status_message(f"Validazione in corso: {xml_file_info.relative_path}")
        self.start_busy_indicator()
        threading.Thread(
            target=self._validate_xml_worker,
            args=(xml_file_info, xsd_path),
            daemon=True,
        ).start()

    def set_status_message(self, message: str, clickable_path: Path | None = None) -> None:
        self.status_var.set(message)
        self.latest_report_path = clickable_path
        if clickable_path is not None:
            self.status_label.configure(cursor="hand2", foreground="#0B5ED7")
        else:
            self.status_label.configure(cursor="", foreground="")

    def on_status_label_clicked(self, _event) -> None:
        if self.latest_report_path is None:
            return
        try:
            self.open_path(self.latest_report_path)
        except Exception as exc:
            self.messagebox.showerror("Errore", f"Impossibile aprire il report:\n{exc}")

    def open_path(self, path: Path) -> None:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
            return
        if sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
            return
        subprocess.Popen(["xdg-open", str(path)])

    def file_key(self, xml_file_info: XmlFileInfo) -> tuple[int, str]:
        return (xml_file_info.track, xml_file_info.relative_path)

    def is_file_checked(self, xml_file_info: XmlFileInfo) -> bool:
        return self.checked_files.get(self.file_key(xml_file_info), False)

    def set_file_checked(self, xml_file_info: XmlFileInfo, checked: bool) -> None:
        self.checked_files[self.file_key(xml_file_info)] = checked

    def populate_file_tree(self) -> None:
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        for index, info in enumerate(self.xml_files, start=1):
            checked_symbol = self.CHECKED_SYMBOL if self.is_file_checked(info) else self.UNCHECKED_SYMBOL
            self.file_tree.insert(
                "",
                "end",
                values=(index, checked_symbol, info.track, info.line_count, f"{info.size_mb:.2f}", info.relative_path),
            )
        self.update_select_all_state()

    def update_select_all_state(self) -> None:
        self._syncing_select_all = True
        all_selected = bool(self.xml_files) and all(self.is_file_checked(info) for info in self.xml_files)
        self.select_all_var.set(all_selected)
        self._syncing_select_all = False

    def toggle_all_files(self) -> None:
        if self._syncing_select_all or not self.xml_files:
            return
        select_all = not all(self.is_file_checked(info) for info in self.xml_files)
        for info in self.xml_files:
            self.set_file_checked(info, select_all)
        self.select_all_var.set(select_all)
        self.populate_file_tree()

    def on_file_clicked(self, event) -> None:
        item_id = self.file_tree.identify_row(event.y)
        if not item_id:
            return
        if self.file_tree.identify_region(event.x, event.y) != "cell":
            return

        column = self.file_tree.identify_column(event.x)
        values = self.file_tree.item(item_id, "values")
        if len(values) != 6:
            return

        track = int(values[2])
        relative_path = values[5]
        xml_file_info = next((info for info in self.xml_files if info.relative_path == relative_path and info.track == track), None)
        if xml_file_info is None:
            return

        if column == "#2":
            self.set_file_checked(xml_file_info, not self.is_file_checked(xml_file_info))
            self.populate_file_tree()
            return
        if column != "#6":
            return

        xsd_path = Path(self.track1_xsd_var.get().strip()) if track == 1 else Path(self.track2_xsd_var.get().strip())
        if not xsd_path.is_file():
            self.messagebox.showerror("Errore", f"XSD del tracciato {track} non valido o non selezionato.")
            return

        should_validate = self.messagebox.askyesno(
            "Validazione XML",
            f"Hai selezionato un file del tracciato {track}.\n\n"
            f"Vuoi validare questo XML con il relativo XSD?\n\n{xml_file_info.relative_path}",
        )
        if not should_validate:
            return

        self.set_status_message(f"Validazione in corso: {xml_file_info.relative_path}")
        self.start_busy_indicator()
        threading.Thread(
            target=self._validate_xml_worker,
            args=(xml_file_info, xsd_path),
            daemon=True,
        ).start()

    def _validate_xml_worker(self, xml_file_info: XmlFileInfo, xsd_path: Path) -> None:
        try:
            errors = validate_xml_against_xsd(xml_file_info.path, xsd_path)
        except Exception as exc:
            self.root.after(0, lambda: self.messagebox.showerror("Errore", str(exc)))
            self.root.after(0, lambda: self.set_status_message("Validazione fallita."))
            self.root.after(0, self.stop_busy_indicator)
            return

        def done() -> None:
            rows = [
                {
                    "file_xml": xml_file_info.relative_path,
                    "path": error["Path"],
                    "messaggio": error["Messaggio"],
                }
                for error in errors
            ]
            self.populate_validation_tree(rows)
            self.notebook.select(2)
            if errors:
                self.set_status_message(f"Validazione completata con {len(errors)} errore/i.")
                self.messagebox.showerror(
                    "Validazione non superata",
                    f"Il file {xml_file_info.relative_path} contiene {len(errors)} errore/i di validazione.\n"
                    "Controlla la tab 'Validazione' per il dettaglio.",
                )
            else:
                self.set_status_message("Validazione completata senza errori.")
                self.messagebox.showinfo(
                    "Validazione OK",
                    f"Il file {xml_file_info.relative_path} e' conforme allo XSD del tracciato {xml_file_info.track}.",
                )
            self.stop_busy_indicator()

        self.root.after(0, done)

    def scan_files(self) -> None:
        validated = self.validate_inputs()
        if validated is None:
            return
        track1_xsd, track2_xsd, xml_dir, _, _ = validated
        filename_token = self.filename_token_var.get().strip()
        self.set_status_message("Scansione XML in corso...")
        self.start_busy_indicator()
        threading.Thread(
            target=self._scan_files_worker,
            args=(track1_xsd, track2_xsd, xml_dir, filename_token),
            daemon=True,
        ).start()

    def _scan_files_worker(self, track1_xsd: Path, track2_xsd: Path, xml_dir: Path, filename_token: str) -> None:
        try:
            root_to_track = {
                parse_xsd_root_name(track1_xsd): 1,
                parse_xsd_root_name(track2_xsd): 2,
            }
            xml_files = scan_xml_files(xml_dir, root_to_track, filename_token=filename_token)
        except Exception as exc:
            self.root.after(0, lambda: self.messagebox.showerror("Errore", str(exc)))
            self.root.after(0, lambda: self.set_status_message("Scansione fallita."))
            self.root.after(0, self.stop_busy_indicator)
            return

        def update_ui() -> None:
            self.xml_files = xml_files
            self.checked_files = {self.file_key(info): True for info in xml_files}
            self.populate_file_tree()
            self.set_status_message(f"Trovati {len(xml_files)} file XML SIAD.")
            self.stop_busy_indicator()

        self.root.after(0, update_ui)

    def generate_report(self) -> None:
        validated = self.validate_inputs()
        if validated is None:
            return
        _, _, xml_dir, output_path, analysis_year = validated
        filename_token = self.filename_token_var.get().strip()
        if not self.xml_files:
            try:
                root_to_track = {
                    parse_xsd_root_name(Path(self.track1_xsd_var.get().strip())): 1,
                    parse_xsd_root_name(Path(self.track2_xsd_var.get().strip())): 2,
                }
                self.xml_files = scan_xml_files(xml_dir, root_to_track, filename_token=filename_token)
                self.checked_files = {self.file_key(info): True for info in self.xml_files}
                self.root.after(0, self.populate_file_tree)
            except Exception as exc:
                self.messagebox.showerror("Errore", str(exc))
                return
        selected_xml_files = [info for info in self.xml_files if self.is_file_checked(info)]
        if not selected_xml_files:
            self.messagebox.showerror("Errore", "Seleziona almeno un file XML da includere nel report.")
            return
        deselected_count = len(self.xml_files) - len(selected_xml_files)
        if deselected_count > 0:
            self.messagebox.showinfo(
                "Selezione parziale",
                f"Il report verra' generato usando {len(selected_xml_files)} file selezionati "
                f"e ignorando {deselected_count} file non selezionati.",
            )
        self.set_status_message("Generazione report in corso...")
        self.start_busy_indicator()
        threading.Thread(
            target=self._generate_report_worker,
            args=(selected_xml_files, output_path, analysis_year),
            daemon=True,
        ).start()

    def _generate_report_worker(self, xml_files: list[XmlFileInfo], output_path: Path, analysis_year: int) -> None:
        try:
            summary_rows, details, unique_cf_rows, global_single_heads = build_report(xml_files, analysis_year)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            if output_path.exists():
                output_path.unlink()
            save_workbook(output_path, summary_rows, details, unique_cf_rows, global_single_heads)
        except Exception as exc:
            self.root.after(0, lambda: self.messagebox.showerror("Errore", str(exc)))
            self.root.after(0, lambda: self.set_status_message("Generazione report fallita."))
            self.root.after(0, self.stop_busy_indicator)
            return

        def done() -> None:
            self.summary_rows = add_total_row(summary_rows, global_single_heads)
            self.populate_summary_tree(self.summary_rows)
            self.set_status_message(f"Report creato: {output_path}", clickable_path=output_path)
            self.stop_busy_indicator()
            self.messagebox.showinfo("Completato", f"Report salvato in:\n{output_path}")

        self.root.after(0, done)

    def populate_summary_tree(self, rows: list[dict]) -> None:
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        if not rows:
            self.summary_tree["columns"] = ()
            return

        headers = list(rows[0].keys())
        self.summary_tree["columns"] = headers
        for header in headers:
            self.summary_tree.heading(header, text=header)
            width = 120 if header == "SEDE" else 190
            anchor = "center" if header == "SEDE" else "w"
            self.summary_tree.column(header, width=width, anchor=anchor)
        for row in rows:
            self.summary_tree.insert("", "end", values=[row[h] for h in headers])

    def populate_validation_tree(self, rows: list[dict[str, str]]) -> None:
        self.validation_rows = rows
        for item in self.validation_tree.get_children():
            self.validation_tree.delete(item)

        for row in rows:
            self.validation_tree.insert("", "end", values=(row["file_xml"], row["path"], row["messaggio"]))


def main() -> None:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Tkinter non disponibile in questa installazione Python. "
            "Per usare la GUI installa una build Python con supporto Tk."
        ) from exc

    root = tk.Tk()
    app = SiadReportApp(root, tk, ttk, filedialog, messagebox)
    root.mainloop()


if __name__ == "__main__":
    main()
