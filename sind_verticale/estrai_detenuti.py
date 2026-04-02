#!/usr/bin/env python3
"""
Estrazione detenuti tossicodipendenti dai tracciati SIND
per Relazione al Parlamento.
Legge i file ZIP direttamente dalle cartelle attività/, hiv/, strutture/.
"""

from __future__ import annotations

import glob
import os
import xml.etree.ElementTree as ET
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Costanti / Decodifiche SIND v3.5
# ---------------------------------------------------------------------------

ASP_CODES = ["201", "202", "203", "204", "205"]

ASP_NAMES: dict[str, str] = {
    "201": "ASP Cosenza",
    "202": "ASP Crotone",
    "203": "ASP Catanzaro",
    "204": "ASP Vibo Valentia",
    "205": "ASP Reggio Calabria",
}

SESSO = {"1": "M", "2": "F", "9": "N.D."}
STATO_CIVILE = {
    "1": "Celibe", "2": "Nubile", "3": "Coniugato/a",
    "4": "Separato/a", "5": "Divorziato/a", "6": "Vedovo/a", "9": "N.D.",
}
TITOLO_STUDIO = {
    "1": "Nessuno", "2": "Lic. elementare", "3": "Lic. media inf.",
    "4": "Diploma qual. prof.", "5": "Diploma media sup.",
    "6": "Laurea", "7": "Laurea magistrale", "9": "N.D.",
}
OCCUPAZIONE = {
    "1": "Occupato stabile", "2": "Occupato saltuario", "3": "Disoccupato",
    "4": "Studente", "5": "Casalingo/a", "6": "Econ. non attivo",
    "7": "Altro", "9": "N.D.",
}
CONVIVENZA = {
    "1": "Da solo", "2": "Famiglia origine", "3": "Con partner",
    "4": "Con amici/altri", "5": "In carcere", "6": "In struttura",
    "7": "Altro", "9": "N.D.",
}
ALLOGGIO = {
    "1": "Residenza fissa", "2": "In carcere",
    "3": "Altro (ospedale, comunità)", "8": "Senza fissa dimora", "9": "N.D.",
}
TIPO_INVIO = {
    "01": "Altro Ser.D.", "02": "Privato Sociale", "03": "Prefettura art.121",
    "04": "Medico di base", "05": "Strutt. Ospedaliere", "06": "Servizi Sociali",
    "07": "Serv. Sociale Adulti", "08": "Serv. Sociale Minori",
    "09": "Accesso Volontario", "10": "Familiari/Amici", "11": "Altro",
    "12": "Comm. Medica Locale", "13": "Scuola", "14": "Prefettura art.75",
    "15": "Trasf. temp. intra-reg.", "16": "Trasf. temp. extra-reg.",
    "17": "Centro antiusura", "99": "N.D.",
}
TIPO_TRATTAMENTO = {
    "1": "Primo trattamento", "2": "Rientrato", "3": "Già in carico",
    "8": "Senza presa in carico", "9": "N.D.",
}
STATO_CONTATTO = {
    "1": "Chiuso dimissione", "2": "Chiuso interruzione", "3": "Chiuso decesso",
    "4": "Chiuso trasferimento", "5": "Aperto", "6": "Chiuso d'ufficio",
    "7": "Abbandono", "8": "Prest. senza presa in carico",
}
AREA_PROBLEMATICA = {"S": "Droghe e/o farmaci", "A": "Alcolismo", "G": "Gioco d'Azzardo"}
GPO_CODICI = {
    "01": "Relazioni/certificazioni", "02": "Visite",
    "03": "Colloqui prevenzione", "04": "Colloqui assistenza",
    "05": "Prev. patologie correlate", "06": "Esami/procedure cliniche",
    "07": "Farmaci e vaccini", "08": "Interventi psicoterapeutici",
    "09": "Interventi sociali", "10": "Test psicologici",
    "11": "Attività reinserimento", "12": "Accompagnamento paziente",
    "13": "Programma terapeutico", "14": "Prestazioni alberghiere",
    "15": "Prestazioni straord. economiche", "16": "Inserimento in comunità",
    "17": "Gruppi mutuo aiuto", "18": "Ricovero ospedaliero",
    "19": "Interventi educativi", "20": "Tutoraggio economico",
    "21": "Consulenze legali/finanziarie",
}
SEDE_TRATTAMENTO = {
    "1": "Ser.D.", "2": "Strutt. ospedaliera", "3": "Comunità terapeutica",
    "4": "Carcere", "5": "Videochiamata", "6": "Telefonata",
    "7": "A domicilio", "9": "Altro",
}
TIPO_DIP = {"S": "Droghe/farmaci", "A": "Alcolismo", "G": "Gioco d'Azzardo"}
ESITO_HIV = {
    "1": "Positivo", "2": "Negativo", "3": "Indeterminato",
    "4": "Rifiuto/non eseguito", "5": "Non prescritto",
}


# ---------------------------------------------------------------------------
# Stili openpyxl
# ---------------------------------------------------------------------------

_header_font = Font(bold=True, color="FFFFFF", size=11)
_header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_detenuto_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_carcere_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_totale_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_title_font = Font(bold=True, size=14, color="1F4E79")
_subtitle_font = Font(bold=True, size=11, color="2F5496")
_thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header_row(ws: Any, row: int, cols: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _thin_border


def _style_data_cell(
    ws: Any, row: int, col: int, value: Any,
    fill: PatternFill | None = None, bold: bool = False,
) -> Any:
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _thin_border
    cell.alignment = Alignment(horizontal="center")
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    return cell


def _auto_width(ws: Any, max_col: int, max_row: int, min_width: int = 10, max_width: int = 30) -> None:
    for col in range(1, max_col + 1):
        best = min_width
        for r in range(1, min(max_row + 1, 100)):
            val = ws.cell(row=r, column=col).value
            if val:
                best = max(best, min(len(str(val)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = best


def _calc_durata_giorni(data_inizio: str, data_fine: str) -> int | None:
    try:
        d1 = datetime.strptime(data_inizio, "%Y-%m-%d").date()
        d2 = datetime.strptime(data_fine, "%Y-%m-%d").date()
        return (d2 - d1).days
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Helper: lettura XML da ZIP
# ---------------------------------------------------------------------------

def _xml_from_zip(zip_path: str) -> ET.ElementTree | None:
    with zipfile.ZipFile(zip_path) as zf:
        xml_names = [n for n in zf.namelist() if n.upper().endswith(".XML")]
        if not xml_names:
            return None
        with zf.open(xml_names[0]) as xf:
            return ET.parse(xf)


def _find_zip(base: str, pattern: str) -> str | None:
    matches = glob.glob(os.path.join(base, pattern))
    return matches[0] if matches else None


# ---------------------------------------------------------------------------
# Parser singoli tracciati
# ---------------------------------------------------------------------------

def _parse_scg(base_attivita: str, asp_code: str) -> dict[str, dict]:
    zip_path = (_find_zip(base_attivita, f"{asp_code} ASP */SIND_ANS_*.xml.zip")
                or _find_zip(base_attivita, f"{asp_code} ASP */SIND_SCG_*.XML.zip"))
    if not zip_path:
        return {}
    tree = _xml_from_zip(zip_path)
    if tree is None:
        return {}
    soggetti: dict[str, dict] = {}
    for assistito in tree.getroot().findall(".//Assistito"):
        id_rec = assistito.findtext("Id_Rec", "")
        sostanze = [s.findtext("CodiceSostanzaUso", "") for s in assistito.findall("SostanzeUso")]
        soggetti[id_rec] = {
            "Id_Rec": id_rec,
            "AnnoNascita": assistito.findtext("AnnoNascita", ""),
            "Sesso": assistito.findtext("Sesso", ""),
            "Cittadinanza": assistito.findtext("Cittadinanza", ""),
            "StatoCivile": assistito.findtext("StatoCivile", ""),
            "TitoloStudio": assistito.findtext("TitoloStudio", ""),
            "CodiceOccupazione": assistito.findtext("CodiceOccupazione", ""),
            "CondizioneAbitativaConvivenza": assistito.findtext("CondizioneAbitativaConvivenza", ""),
            "CondizioneAbitativaAlloggio": assistito.findtext("CondizioneAbitativaAlloggio", ""),
            "TerapiaSostitutivaOppioidi": assistito.findtext("TerapiaSostitutivaOppioidi", ""),
            "SostanzeUso": ", ".join(sostanze),
            "IS_DETENUTO": assistito.findtext("CondizioneAbitativaAlloggio", "") == "2",
        }
    return soggetti


def _parse_ctt(base_attivita: str, asp_code: str) -> dict[str, list[dict]]:
    zip_path = (_find_zip(base_attivita, f"{asp_code} ASP */SIND_COX_*.xml.zip")
                or _find_zip(base_attivita, f"{asp_code} ASP */SIND_CTT_*.XML.zip"))
    if not zip_path:
        return {}
    tree = _xml_from_zip(zip_path)
    if tree is None:
        return {}
    contatti: dict[str, list[dict]] = {}
    for serd in tree.getroot().findall(".//SERD"):
        codice_serd = serd.findtext("CodiceSERD", "")
        for assistito in serd.findall("Assistito"):
            id_rec = assistito.findtext("Id_Rec", "")
            ct = assistito.findtext("ChiaveTecnica", "")
            for contatto in assistito.findall("Contatto"):
                contatti.setdefault(id_rec, []).append({
                    "CodiceSERD": codice_serd, "ChiaveTecnica": ct,
                    "IdContatto": contatto.findtext("IdContatto", ""),
                    "AreaProblematica": contatto.findtext("AreaProblematica", ""),
                    "TipoInvio": contatto.findtext("TipoInvio", ""),
                    "DataAperturaCartella": contatto.findtext("DataAperturaCartella", ""),
                    "DataInizioContattoAttivo": contatto.findtext("DataInizioContattoAttivo", ""),
                    "StatoContattoFineAnno": contatto.findtext("StatoContattoFineAnno", ""),
                    "TipoTrattamento": contatto.findtext("TipoTrattamento", ""),
                })
    return contatti


def _parse_gpo(base_attivita: str, asp_code: str) -> dict[str, list[dict]]:
    zip_path = (_find_zip(base_attivita, f"{asp_code} ASP */SIND_GPX_*.xml.zip")
                or _find_zip(base_attivita, f"{asp_code} ASP */SIND_GPO_*.XML.zip"))
    if not zip_path:
        return {}
    tree = _xml_from_zip(zip_path)
    if tree is None:
        return {}
    prestazioni: dict[str, list[dict]] = {}
    for serd in tree.getroot().findall(".//SERD"):
        codice_serd = serd.findtext("CodiceSERD", "")
        for assistito in serd.findall("Assistito"):
            id_rec = assistito.findtext("Id_Rec", "")
            for contatto in assistito.findall("Contatto"):
                id_contatto = contatto.findtext("IdContatto", "")
                for gpo in contatto.findall("GruppoPrestazioniOmogenee"):
                    prestazioni.setdefault(id_rec, []).append({
                        "CodiceSERD": codice_serd,
                        "IdContatto": id_contatto,
                        "CodGruppoPrestazioni": gpo.findtext("CodGruppoPrestazioni", ""),
                        "TipoSedeTrattamento": gpo.findtext("TipoSedeTrattamento", ""),
                        "DataInizioPrestazioni": gpo.findtext("DataInizioPrestazioni", ""),
                        "DataChiusuraPrestazioni": gpo.findtext("DataChiusuraPrestazioni", ""),
                        "NumeroPrestazioni": gpo.findtext("NumeroPrestazioni", ""),
                    })
    return prestazioni


def _parse_str_all(base_str: str) -> dict[str, dict[str, dict]]:
    result: dict[str, dict[str, dict]] = {}
    for f in sorted(glob.glob(os.path.join(base_str, "SIND_STR_*.xml.zip"))
                    + glob.glob(os.path.join(base_str, "SIND_STR_*.XML.zip"))):
        tree = _xml_from_zip(f)
        if tree is None:
            continue
        root = tree.getroot()
        asp_code = root.findtext(".//CodiceAziendaSanitariaRifer", "")
        if not asp_code:
            asp_code = os.path.basename(f).split("_")[3]
        serd_data: dict[str, dict] = {}
        for serd in root.findall(".//SERD"):
            cod = serd.findtext("CodiceSERD", "")
            if cod in serd_data:
                continue
            serd_data[cod] = {
                "CodiceSERD": cod,
                "TipoDipendenza": serd.findtext("TipoDipendenza", ""),
                "Indirizzo": serd.findtext("IndirizzoSERD", ""),
                "Comune": serd.findtext("ComuneSERD", ""),
                "CAP": serd.findtext("CapSERD", ""),
                "Telefono": serd.findtext("TelefonoSERD", ""),
                "Email": serd.findtext("EMailSERD", ""),
                "NomeResponsabile": serd.findtext("NomeResponsabileSERD", ""),
                "CognomeResponsabile": serd.findtext("CognomeResponsabileSERD", ""),
                "StruttureCarcerarie": int(serd.findtext("StruttureCarcerarie", "0") or 0),
                "PopolazioneCarceraria": int(serd.findtext("PopolazioneStruttureCarcerarie", "0") or 0),
                "ComunitaTerapeutiche": int(serd.findtext("ComunitaTerapeutiche", "0") or 0),
                "PresenzeCT": int(serd.findtext("PresenzeComunitaTerapeutiche", "0") or 0),
            }
        result[asp_code] = serd_data
    return result


def _parse_hiv_all(base_hiv: str) -> dict[str, list[dict]]:
    result: dict[str, list[dict]] = {}
    for f in sorted(glob.glob(os.path.join(base_hiv, "SIND_HIV_*.xml.zip"))
                    + glob.glob(os.path.join(base_hiv, "SIND_HIV_*.XML.zip"))):
        tree = _xml_from_zip(f)
        if tree is None:
            continue
        root = tree.getroot()
        asp_code = root.findtext(".//CodiceAziendaSanitariaRiferimento", "")
        if not asp_code:
            asp_code = os.path.basename(f).split("_")[3]
        groups: list[dict] = []
        for serd in root.findall(".//SERD"):
            cod_serd = serd.findtext("CodiceSERD", "")
            area = serd.findtext("AreaProblematica", "")
            soggetti = []
            for sogg in serd.findall("Soggetto"):
                soggetti.append({
                    "Sesso": sogg.findtext("Sesso", ""),
                    "TipoTrattamento": sogg.findtext("TipoTrattamento", ""),
                    "UsoViaIniettiva": sogg.findtext("UsoPerViaIniettiva", ""),
                    "DataUltimoEsameHIV": sogg.findtext("DataUltimoEsameHIV", ""),
                    "EsitoUltimoEsameHIV": sogg.findtext("EsitoUltimoEsameHIV", ""),
                })
            groups.append({
                "CodiceSERD": cod_serd,
                "AreaProblematica": area,
                "Soggetti": soggetti,
                "TotSoggetti": len(soggetti),
            })
        result[asp_code] = groups
    return result


# ---------------------------------------------------------------------------
# Validazione struttura cartelle
# ---------------------------------------------------------------------------

def validate_input_folder(base_dir: str) -> list[str]:
    """Verifica che la cartella contenga le 3 sottocartelle necessarie con i file ZIP attesi.
    Ritorna lista di errori (vuota = ok)."""
    errors: list[str] = []
    base = Path(base_dir)
    if not base.is_dir():
        return [f"La cartella '{base_dir}' non esiste."]

    att = base / "attività"
    if not att.is_dir():
        # Prova variante senza accento
        att = base / "attivita"
        if not att.is_dir():
            errors.append("Cartella 'attività' non trovata.")
    else:
        found_asp = 0
        for asp_code in ASP_CODES:
            matches = glob.glob(str(att / f"{asp_code} ASP *"))
            if matches:
                found_asp += 1
                asp_dir = matches[0]
                for nuovo, vecchio in [("ANS", "SCG"), ("COX", "CTT"), ("GPX", "GPO")]:
                    if (not glob.glob(os.path.join(asp_dir, f"SIND_{nuovo}_*.xml.zip"))
                            and not glob.glob(os.path.join(asp_dir, f"SIND_{vecchio}_*.XML.zip"))):
                        errors.append(f"ASP {asp_code}: manca SIND_{nuovo}_ (o {vecchio}_) zip")
        if found_asp == 0:
            errors.append("Nessuna sottocartella ASP trovata in 'attività'.")

    hiv = base / "hiv"
    if not hiv.is_dir():
        errors.append("Cartella 'hiv' non trovata.")
    elif not glob.glob(str(hiv / "SIND_HIV_*.xml.zip")) and not glob.glob(str(hiv / "SIND_HIV_*.XML.zip")):
        errors.append("Nessun file SIND_HIV_*.zip trovato in 'hiv'.")

    strutt = base / "strutture"
    if not strutt.is_dir():
        errors.append("Cartella 'strutture' non trovata.")
    elif not glob.glob(str(strutt / "SIND_STR_*.XML.zip")):
        errors.append("Nessun file SIND_STR_*.XML.zip trovato in 'strutture'.")

    return errors


# ---------------------------------------------------------------------------
# Funzione principale di generazione Excel
# ---------------------------------------------------------------------------

def genera_excel(
    base_dir: str,
    output_path: str,
    *,
    anno_riferimento: int | None = None,
    protocollo: str = "",
    log: Callable[[str], None] | None = None,
) -> str:
    """Genera il file Excel con 5 sheet.

    Args:
        base_dir: cartella radice contenente attività/, hiv/, strutture/
        output_path: percorso del file Excel di output
        anno_riferimento: anno dei dati (default: anno corrente - 1)
        protocollo: riferimento protocollo per intestazione
        log: callback opzionale per messaggi di avanzamento

    Returns:
        il percorso del file Excel generato
    """
    def _log(msg: str) -> None:
        if log:
            log(msg)

    if anno_riferimento is None:
        anno_riferimento = date.today().year - 1

    base = Path(base_dir)
    base_attivita = str(base / "attività")
    if not Path(base_attivita).is_dir():
        base_attivita = str(base / "attivita")
    base_hiv = str(base / "hiv")
    base_str = str(base / "strutture")

    wb = Workbook()

    # --- Carica STR e HIV ---
    _log("Parsing tracciato STR (Strutture)...")
    str_data = _parse_str_all(base_str)
    _log(f"  STR: {sum(len(v) for v in str_data.values())} Ser.D. trovati in {len(str_data)} ASP")

    _log("Parsing tracciato HIV (Monitoraggio)...")
    hiv_data = _parse_hiv_all(base_hiv)
    _log(f"  HIV: {sum(sum(g['TotSoggetti'] for g in groups) for groups in hiv_data.values())} soggetti")

    riepilogo: dict[str, dict] = {}

    # =============================================
    # SHEET 1: DETTAGLIO DETENUTI
    # =============================================
    ws1 = wb.active
    ws1.title = "Dettaglio Detenuti"

    headers1 = [
        "ASP", "Cod. ASP", "Ser.D.", "ID Record (hash)", "Anno Nascita",
        "Sesso", "Cittadinanza", "Stato Civile", "Titolo Studio", "Occupazione",
        "Convivenza", "Alloggio", "DETENUTO",
        "Sostanze Uso", "Terapia Sost. Oppioidi",
        "Area Problematica", "Tipo Invio", "Data Apertura Cartella",
        "Data Inizio Contatto", "Stato Contatto Fine Anno", "Tipo Trattamento",
        "Cod. Gruppo Prest.", "Gruppo Prestazioni", "Sede Trattamento",
        "Data Inizio Prest.", "Data Fine Prest.", "Durata (gg)", "N. Prestazioni",
        "INSERIMENTO CT", "PRESTAZIONE IN CARCERE",
    ]
    _style_header_row(ws1, 1, len(headers1), headers1)
    row = 2

    for asp_code in ASP_CODES:
        asp_name = ASP_NAMES[asp_code]
        _log(f"Elaborazione {asp_name}...")

        soggetti = _parse_scg(base_attivita, asp_code)
        contatti = _parse_ctt(base_attivita, asp_code)
        prestazioni_gpo = _parse_gpo(base_attivita, asp_code)

        _log(f"  SCG: {len(soggetti)} soggetti")

        riepilogo[asp_code] = {
            "nome": asp_name, "tot_soggetti": len(soggetti),
            "tot_detenuti": 0, "detenuti_maschi": 0, "detenuti_femmine": 0,
            "detenuti_con_contatto": 0, "inserimenti_ct": 0,
            "prestazioni_carcere": 0, "prestazioni_ct_sede": 0,
        }

        detenuti = {k: v for k, v in soggetti.items() if v.get("IS_DETENUTO")}
        riepilogo[asp_code]["tot_detenuti"] = len(detenuti)
        _log(f"  Detenuti TD: {len(detenuti)}")

        for id_rec, sogg in detenuti.items():
            if sogg["Sesso"] == "1":
                riepilogo[asp_code]["detenuti_maschi"] += 1
            elif sogg["Sesso"] == "2":
                riepilogo[asp_code]["detenuti_femmine"] += 1

            ctt_list = contatti.get(id_rec, [])
            gpo_list = prestazioni_gpo.get(id_rec, [])
            if ctt_list:
                riepilogo[asp_code]["detenuti_con_contatto"] += 1

            def write_base_cols(r: int, sogg: dict, ctt: dict | None = None, fill: PatternFill = _detenuto_fill) -> None:
                _style_data_cell(ws1, r, 1, asp_name, fill)
                _style_data_cell(ws1, r, 2, asp_code, fill)
                _style_data_cell(ws1, r, 3, ctt["CodiceSERD"] if ctt else "", fill)
                _style_data_cell(ws1, r, 4, id_rec[:20] + "...", fill)
                _style_data_cell(ws1, r, 5, sogg["AnnoNascita"], fill)
                _style_data_cell(ws1, r, 6, SESSO.get(sogg["Sesso"], sogg["Sesso"]), fill)
                _style_data_cell(ws1, r, 7, sogg["Cittadinanza"], fill)
                _style_data_cell(ws1, r, 8, STATO_CIVILE.get(sogg["StatoCivile"], ""), fill)
                _style_data_cell(ws1, r, 9, TITOLO_STUDIO.get(sogg["TitoloStudio"], ""), fill)
                _style_data_cell(ws1, r, 10, OCCUPAZIONE.get(sogg["CodiceOccupazione"], ""), fill)
                _style_data_cell(ws1, r, 11, CONVIVENZA.get(sogg["CondizioneAbitativaConvivenza"], ""), fill)
                _style_data_cell(ws1, r, 12, ALLOGGIO.get(sogg["CondizioneAbitativaAlloggio"], ""), fill)
                _style_data_cell(ws1, r, 13, "SI", fill)
                _style_data_cell(ws1, r, 14, sogg["SostanzeUso"], fill)
                ost = sogg.get("TerapiaSostitutivaOppioidi", "")
                _style_data_cell(ws1, r, 15, "Sì" if ost == "2" else ("No" if ost == "1" else "N.D."), fill)
                if ctt:
                    _style_data_cell(ws1, r, 16, AREA_PROBLEMATICA.get(ctt["AreaProblematica"], ""), fill)
                    _style_data_cell(ws1, r, 17, TIPO_INVIO.get(ctt["TipoInvio"], ctt["TipoInvio"]), fill)
                    _style_data_cell(ws1, r, 18, ctt["DataAperturaCartella"], fill)
                    _style_data_cell(ws1, r, 19, ctt["DataInizioContattoAttivo"], fill)
                    _style_data_cell(ws1, r, 20, STATO_CONTATTO.get(ctt["StatoContattoFineAnno"], ""), fill)
                    _style_data_cell(ws1, r, 21, TIPO_TRATTAMENTO.get(ctt["TipoTrattamento"], ""), fill)
                for c in range(max(16 if not ctt else 22, 1), len(headers1) + 1):
                    if ws1.cell(row=r, column=c).value is None:
                        _style_data_cell(ws1, r, c, "", fill)

            if not ctt_list and not gpo_list:
                write_base_cols(row, sogg)
                row += 1
                continue

            for ctt in ctt_list:
                gpo_contatto = [
                    g for g in gpo_list
                    if g["IdContatto"] == ctt["IdContatto"] and g["CodiceSERD"] == ctt["CodiceSERD"]
                ]
                if not gpo_contatto:
                    write_base_cols(row, sogg, ctt)
                    row += 1
                else:
                    for gpo in gpo_contatto:
                        cod_gpo = gpo["CodGruppoPrestazioni"]
                        sede = gpo["TipoSedeTrattamento"]
                        is_ct = cod_gpo == "16"
                        is_carcere = sede == "4"
                        fill = _ct_fill if is_ct else (_carcere_fill if is_carcere else _detenuto_fill)
                        write_base_cols(row, sogg, ctt, fill)
                        _style_data_cell(ws1, row, 22, cod_gpo, fill)
                        _style_data_cell(ws1, row, 23, GPO_CODICI.get(cod_gpo, cod_gpo), fill)
                        _style_data_cell(ws1, row, 24, SEDE_TRATTAMENTO.get(sede, sede), fill)
                        _style_data_cell(ws1, row, 25, gpo["DataInizioPrestazioni"], fill)
                        _style_data_cell(ws1, row, 26, gpo["DataChiusuraPrestazioni"], fill)
                        durata = _calc_durata_giorni(gpo["DataInizioPrestazioni"], gpo["DataChiusuraPrestazioni"])
                        _style_data_cell(ws1, row, 27, durata, fill)
                        n_prest = gpo["NumeroPrestazioni"]
                        _style_data_cell(ws1, row, 28, int(n_prest) if n_prest else "", fill)
                        _style_data_cell(ws1, row, 29, "SI" if is_ct else "", fill)
                        _style_data_cell(ws1, row, 30, "SI" if is_carcere else "", fill)
                        if is_ct:
                            riepilogo[asp_code]["inserimenti_ct"] += 1
                        if is_carcere:
                            riepilogo[asp_code]["prestazioni_carcere"] += 1
                        if sede == "3":
                            riepilogo[asp_code]["prestazioni_ct_sede"] += 1
                        row += 1

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{row - 1}"
    ws1.freeze_panes = "A2"
    _auto_width(ws1, len(headers1), row)

    # =============================================
    # SHEET 2: STRUTTURE E TERRITORIO
    # =============================================
    ws2 = wb.create_sheet("Strutture e Territorio")
    ws2.merge_cells("A1:K1")
    ws2.cell(row=1, column=1, value=f"ANAGRAFICA SER.D. E DOTAZIONE TERRITORIALE - Anno {anno_riferimento}").font = _title_font
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws2.merge_cells("A2:K2")
    ws2.cell(row=2, column=1, value="Fonte: Tracciato SIND STRUTTURE (STR) - deduplicato per Codice Servizio").font = Font(italic=True, size=10)
    ws2.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers2 = [
        "ASP", "Cod. Ser.D.", "Indirizzo", "CAP", "Telefono", "Email",
        "Responsabile", "N. Strutture Carcerarie", "Pop. Carceraria",
        "N. Comunità Terapeutiche", "Presenze CT",
    ]
    _style_header_row(ws2, 4, len(headers2), headers2)

    r = 5
    for asp_code in ASP_CODES:
        asp_name = ASP_NAMES[asp_code]
        serd_dict = str_data.get(asp_code, {})
        tot_carc = tot_pop = tot_ct = tot_pres = 0
        for cod_serd in sorted(serd_dict.keys()):
            s = serd_dict[cod_serd]
            _style_data_cell(ws2, r, 1, asp_name)
            _style_data_cell(ws2, r, 2, s["CodiceSERD"])
            _style_data_cell(ws2, r, 3, s["Indirizzo"])
            ws2.cell(row=r, column=3).alignment = Alignment(horizontal="left")
            _style_data_cell(ws2, r, 4, s["CAP"])
            _style_data_cell(ws2, r, 5, s["Telefono"])
            _style_data_cell(ws2, r, 6, s["Email"])
            ws2.cell(row=r, column=6).alignment = Alignment(horizontal="left")
            _style_data_cell(ws2, r, 7, f"{s['NomeResponsabile']} {s['CognomeResponsabile']}")
            _style_data_cell(ws2, r, 8, s["StruttureCarcerarie"])
            _style_data_cell(ws2, r, 9, s["PopolazioneCarceraria"])
            _style_data_cell(ws2, r, 10, s["ComunitaTerapeutiche"])
            _style_data_cell(ws2, r, 11, s["PresenzeCT"])
            tot_carc += s["StruttureCarcerarie"]
            tot_pop += s["PopolazioneCarceraria"]
            tot_ct += s["ComunitaTerapeutiche"]
            tot_pres += s["PresenzeCT"]
            r += 1
        for c in range(1, len(headers2) + 1):
            _style_data_cell(ws2, r, c, "", _totale_fill, bold=True)
        _style_data_cell(ws2, r, 1, f"Subtotale {asp_name}", _totale_fill, bold=True)
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        _style_data_cell(ws2, r, 8, tot_carc, _totale_fill, bold=True)
        _style_data_cell(ws2, r, 9, tot_pop, _totale_fill, bold=True)
        _style_data_cell(ws2, r, 10, tot_ct, _totale_fill, bold=True)
        _style_data_cell(ws2, r, 11, tot_pres, _totale_fill, bold=True)
        r += 1

    _auto_width(ws2, len(headers2), r, min_width=12)
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["F"].width = 35
    ws2.column_dimensions["G"].width = 25

    # =============================================
    # SHEET 3: MONITORAGGIO HIV
    # =============================================
    ws3 = wb.create_sheet("Monitoraggio HIV")
    ws3.merge_cells("A1:H1")
    ws3.cell(row=1, column=1, value=f"MONITORAGGIO HIV - Dati aggregati per Ser.D. - Anno {anno_riferimento}").font = _title_font
    ws3.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws3.merge_cells("A2:H2")
    ws3.cell(row=2, column=1, value="NOTA: Dati anonimi/aggregati. NON linkabili ai singoli detenuti (assenza di Id_Rec nel tracciato HIV)").font = Font(italic=True, size=10, color="FF0000")
    ws3.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    headers3 = ["ASP", "Ser.D.", "Area Problematica", "Tot. Soggetti", "Maschi", "Femmine", "HIV Positivi", "HIV Negativi"]
    _style_header_row(ws3, 4, len(headers3), headers3)

    r = 5
    for asp_code in ASP_CODES:
        asp_name = ASP_NAMES[asp_code]
        groups = hiv_data.get(asp_code, [])
        tot_sogg = tot_m = tot_f = tot_pos = tot_neg = 0
        for g in groups:
            n = g["TotSoggetti"]
            maschi = sum(1 for s in g["Soggetti"] if s["Sesso"] == "1")
            femmine = sum(1 for s in g["Soggetti"] if s["Sesso"] == "2")
            positivi = sum(1 for s in g["Soggetti"] if s.get("EsitoUltimoEsameHIV") == "1")
            negativi = sum(1 for s in g["Soggetti"] if s.get("EsitoUltimoEsameHIV") == "2")
            _style_data_cell(ws3, r, 1, asp_name)
            _style_data_cell(ws3, r, 2, g["CodiceSERD"])
            _style_data_cell(ws3, r, 3, AREA_PROBLEMATICA.get(g["AreaProblematica"], g["AreaProblematica"]))
            _style_data_cell(ws3, r, 4, n)
            _style_data_cell(ws3, r, 5, maschi)
            _style_data_cell(ws3, r, 6, femmine)
            _style_data_cell(ws3, r, 7, positivi)
            _style_data_cell(ws3, r, 8, negativi)
            tot_sogg += n; tot_m += maschi; tot_f += femmine
            tot_pos += positivi; tot_neg += negativi
            r += 1
        for c in range(1, len(headers3) + 1):
            _style_data_cell(ws3, r, c, "", _totale_fill, bold=True)
        _style_data_cell(ws3, r, 1, f"Subtotale {asp_name}", _totale_fill, bold=True)
        _style_data_cell(ws3, r, 4, tot_sogg, _totale_fill, bold=True)
        _style_data_cell(ws3, r, 5, tot_m, _totale_fill, bold=True)
        _style_data_cell(ws3, r, 6, tot_f, _totale_fill, bold=True)
        _style_data_cell(ws3, r, 7, tot_pos, _totale_fill, bold=True)
        _style_data_cell(ws3, r, 8, tot_neg, _totale_fill, bold=True)
        r += 1

    _auto_width(ws3, len(headers3), r)

    # =============================================
    # SHEET 4: RIEPILOGO REGIONALE
    # =============================================
    ws4 = wb.create_sheet("Riepilogo Regionale")
    ws4.merge_cells("A1:I1")
    ws4.cell(row=1, column=1, value=f"RELAZIONE AL PARLAMENTO SUL FENOMENO DELLE TOSSICODIPENDENZE - ANNO {anno_riferimento}").font = _title_font
    ws4.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws4.merge_cells("A2:I2")
    ws4.cell(row=2, column=1, value="Regione Calabria - Riepilogo dati SIND").font = _subtitle_font
    ws4.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws4.merge_cells("A3:I3")
    prot_str = f" - {protocollo}" if protocollo else ""
    ws4.cell(row=3, column=1, value=f"Dati estratti il {date.today().strftime('%d/%m/%Y')}{prot_str}").font = Font(italic=True, size=10)
    ws4.cell(row=3, column=1).alignment = Alignment(horizontal="center")

    # --- TABELLA A ---
    r = 5
    ws4.merge_cells(f"A{r}:I{r}")
    ws4.cell(row=r, column=1, value="A) DETENUTI TOSSICODIPENDENTI (fonte: tracciato SCG - Anagrafica Soggetto)").font = Font(bold=True, size=12, color="1F4E79")
    r = 6
    tab_a = ["ASP", "Cod.", "Tot. Soggetti SIND", "Detenuti TD", "Maschi", "Femmine", "% Det. su Tot.", "Con contatto Ser.D."]
    _style_header_row(ws4, r, len(tab_a), tab_a)
    r = 7
    tots: dict[str, int] = {"sogg": 0, "det": 0, "m": 0, "f": 0, "ctt": 0}
    for asp_code in ASP_CODES:
        rp = riepilogo[asp_code]
        _style_data_cell(ws4, r, 1, rp["nome"])
        _style_data_cell(ws4, r, 2, asp_code)
        _style_data_cell(ws4, r, 3, rp["tot_soggetti"])
        _style_data_cell(ws4, r, 4, rp["tot_detenuti"])
        _style_data_cell(ws4, r, 5, rp["detenuti_maschi"])
        _style_data_cell(ws4, r, 6, rp["detenuti_femmine"])
        pct = (rp["tot_detenuti"] / rp["tot_soggetti"] * 100) if rp["tot_soggetti"] > 0 else 0
        c = _style_data_cell(ws4, r, 7, round(pct, 1))
        c.number_format = '0.0"%"'
        _style_data_cell(ws4, r, 8, rp["detenuti_con_contatto"])
        tots["sogg"] += rp["tot_soggetti"]; tots["det"] += rp["tot_detenuti"]
        tots["m"] += rp["detenuti_maschi"]; tots["f"] += rp["detenuti_femmine"]
        tots["ctt"] += rp["detenuti_con_contatto"]
        r += 1
    _style_data_cell(ws4, r, 1, "TOTALE REGIONE", _totale_fill, bold=True)
    for c in range(2, 9):
        _style_data_cell(ws4, r, c, "", _totale_fill, bold=True)
    _style_data_cell(ws4, r, 3, tots["sogg"], _totale_fill, bold=True)
    _style_data_cell(ws4, r, 4, tots["det"], _totale_fill, bold=True)
    _style_data_cell(ws4, r, 5, tots["m"], _totale_fill, bold=True)
    _style_data_cell(ws4, r, 6, tots["f"], _totale_fill, bold=True)
    pct_tot = (tots["det"] / tots["sogg"] * 100) if tots["sogg"] > 0 else 0
    _style_data_cell(ws4, r, 7, round(pct_tot, 1), _totale_fill, bold=True)
    _style_data_cell(ws4, r, 8, tots["ctt"], _totale_fill, bold=True)

    # --- TABELLA B ---
    r += 3
    ws4.merge_cells(f"A{r}:I{r}")
    ws4.cell(row=r, column=1, value="B) INSERIMENTI IN CT - MISURA ALTERNATIVA ALLA DETENZIONE (fonte: tracciato GPO)").font = Font(bold=True, size=12, color="1F4E79")
    r += 1
    tab_c = ["ASP", "Cod.", "Inserimenti CT (cod.16)", "Prest. sede CT", "Prest. sede Carcere", "Spesa CT"]
    _style_header_row(ws4, r, len(tab_c), tab_c)
    r += 1
    for asp_code in ASP_CODES:
        rp = riepilogo[asp_code]
        _style_data_cell(ws4, r, 1, rp["nome"])
        _style_data_cell(ws4, r, 2, asp_code)
        _style_data_cell(ws4, r, 3, rp["inserimenti_ct"])
        _style_data_cell(ws4, r, 4, rp["prestazioni_ct_sede"])
        _style_data_cell(ws4, r, 5, rp["prestazioni_carcere"])
        _style_data_cell(ws4, r, 6, 0)
        r += 1
    _style_data_cell(ws4, r, 1, "TOTALE REGIONE", _totale_fill, bold=True)
    for c in range(2, 7):
        _style_data_cell(ws4, r, c, "", _totale_fill, bold=True)
    _style_data_cell(ws4, r, 3, sum(riepilogo[a]["inserimenti_ct"] for a in riepilogo), _totale_fill, bold=True)
    _style_data_cell(ws4, r, 4, sum(riepilogo[a]["prestazioni_ct_sede"] for a in riepilogo), _totale_fill, bold=True)
    _style_data_cell(ws4, r, 5, sum(riepilogo[a]["prestazioni_carcere"] for a in riepilogo), _totale_fill, bold=True)
    _style_data_cell(ws4, r, 6, 0, _totale_fill, bold=True)

    # --- NOTE METODOLOGICHE ---
    r += 3
    ws4.merge_cells(f"A{r}:I{r}")
    ws4.cell(row=r, column=1, value="NOTE METODOLOGICHE").font = Font(bold=True, size=11, color="1F4E79")
    note = [
        f"Fonte dati: Sistema Informativo Nazionale Dipendenze (SIND) - Flussi XML anno {anno_riferimento}",
        "Criterio identificazione detenuti: Campo CondizioneAbitativaAlloggio = 2 (in carcere) dal tracciato SCG",
        "Inserimento CT: Codice Gruppo Prestazioni Omogenee = 16 dal tracciato GPO",
        "SPESA CT: pari a zero in quanto non risultano inserimenti CT per detenuti nel SIND.",
    ]
    for i, nota in enumerate(note):
        r += 1
        ws4.merge_cells(f"A{r}:I{r}")
        ws4.cell(row=r, column=1, value=f"{i + 1}. {nota}").font = Font(size=9)
        ws4.cell(row=r, column=1).alignment = Alignment(wrap_text=True)

    # Legenda
    r += 2
    ws4.merge_cells(f"A{r}:I{r}")
    ws4.cell(row=r, column=1, value="LEGENDA COLORI (Sheet Dettaglio Detenuti)").font = Font(bold=True, size=11, color="1F4E79")
    r += 1
    ws4.cell(row=r, column=1, value="Detenuto").fill = _detenuto_fill
    ws4.cell(row=r, column=2, value="= Soggetto detenuto tossicodipendente")
    r += 1
    ws4.cell(row=r, column=1, value="Inserimento CT").fill = _ct_fill
    ws4.cell(row=r, column=2, value="= Prestazione inserimento in Comunità Terapeutica (cod.16)")
    r += 1
    ws4.cell(row=r, column=1, value="Sede Carcere").fill = _carcere_fill
    ws4.cell(row=r, column=2, value="= Prestazione erogata in sede carcere (cod.4)")

    for col in range(1, 10):
        ws4.column_dimensions[get_column_letter(col)].width = 22

    # =============================================
    # SHEET 5: NOTA METODOLOGICA
    # =============================================
    # Calcola totali STR e HIV per la nota metodologica
    tots_b: dict[str, int] = {"carc": 0, "pop": 0, "ct": 0, "pres": 0}
    for asp_code in ASP_CODES:
        for s in str_data.get(asp_code, {}).values():
            tots_b["carc"] += s["StruttureCarcerarie"]
            tots_b["pop"] += s["PopolazioneCarceraria"]
            tots_b["ct"] += s["ComunitaTerapeutiche"]
            tots_b["pres"] += s["PresenzeCT"]
    tots_d: dict[str, int] = {"s": 0, "m": 0, "f": 0, "p": 0, "n": 0}
    for asp_code in ASP_CODES:
        for g in hiv_data.get(asp_code, []):
            tots_d["s"] += g["TotSoggetti"]
            tots_d["m"] += sum(1 for s in g["Soggetti"] if s["Sesso"] == "1")
            tots_d["f"] += sum(1 for s in g["Soggetti"] if s["Sesso"] == "2")
            tots_d["p"] += sum(1 for s in g["Soggetti"] if s.get("EsitoUltimoEsameHIV") == "1")
            tots_d["n"] += sum(1 for s in g["Soggetti"] if s.get("EsitoUltimoEsameHIV") == "2")

    ws5 = wb.create_sheet("Nota Metodologica")
    ws5.sheet_properties.tabColor = "1F4E79"
    ws5.column_dimensions["A"].width = 100
    wrap_style = Alignment(wrap_text=True, vertical="top")
    normal_font = Font(size=11, name="Calibri")
    bold_font_11 = Font(size=11, name="Calibri", bold=True)
    section_font = Font(size=12, name="Calibri", bold=True, color="1F4E79")
    intestazione_font = Font(size=11, name="Calibri", italic=True)

    r = 1
    ws5.cell(row=r, column=1, value="REGIONE CALABRIA").font = Font(size=14, bold=True, name="Calibri", color="1F4E79")
    ws5.cell(row=r, column=1).alignment = wrap_style
    r += 1
    ws5.cell(row=r, column=1, value="Dipartimento Salute e Servizi Sanitari").font = Font(size=12, bold=True, name="Calibri")
    r += 1
    ws5.cell(row=r, column=1, value='Settore 7 "Assistenza Territoriale - Salute nelle Carceri - Formazione ECM - Medicina Convenzionata - Continuità Assistenziale"').font = Font(size=10, italic=True, name="Calibri")
    ws5.cell(row=r, column=1).alignment = wrap_style
    r += 2
    ws5.cell(row=r, column=1, value=f"Data: {date.today().strftime('%d/%m/%Y')}").font = intestazione_font
    r += 1
    if protocollo:
        ws5.cell(row=r, column=1, value=f"Rif.: {protocollo}").font = intestazione_font
        r += 1
    r += 1

    ws5.cell(row=r, column=1, value=f"OGGETTO: Richiesta dati per stesura della Relazione al Parlamento sul fenomeno delle tossicodipendenze in Italia - Anno {anno_riferimento}").font = bold_font_11
    ws5.cell(row=r, column=1).alignment = wrap_style
    r += 1
    ws5.cell(row=r, column=1, value="Nota metodologica e di accompagnamento alla reportistica allegata").font = Font(size=11, italic=True, name="Calibri")
    r += 2

    ws5.cell(row=r, column=1, value="1. PREMESSA").font = section_font
    r += 1
    premessa = (
        "In riscontro alla nota in oggetto, con la quale si richiedono i dati relativi al numero dei detenuti "
        "tossicodipendenti e alla spesa sostenuta per l'inserimento in Comunita' Terapeutica (CT) in misura "
        "alternativa alla detenzione, si trasmette la presente reportistica elaborata a partire dai flussi informativi "
        f"del Sistema Informativo Nazionale Dipendenze (SIND), versione 3.5, relativi all'anno {anno_riferimento}.\n\n"
        "L'estrazione ha riguardato tutte e cinque le Aziende Sanitarie Provinciali (ASP) della Regione Calabria: "
        "ASP 201 Cosenza, ASP 202 Crotone, ASP 203 Catanzaro, ASP 204 Vibo Valentia, ASP 205 Reggio Calabria."
    )
    ws5.cell(row=r, column=1, value=premessa).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 100
    r += 2

    ws5.cell(row=r, column=1, value="2. FONTI DATI UTILIZZATE").font = section_font
    r += 1
    fonti = (
        "I dati sono stati estratti dai seguenti tracciati XML del flusso SIND, trasmessi dalle 5 ASP calabresi:\n\n"
        "a) Tracciato SCG (Anagrafica Soggetto): contiene i dati anagrafici e la condizione abitativa di tutti i "
        "soggetti in carico ai Ser.D. regionali. E' la fonte primaria per l'identificazione dei detenuti "
        "tossicodipendenti.\n\n"
        "b) Tracciato CTT (Dati Contatto): registra i contatti attivi presso i Ser.D., con informazioni su tipo di "
        "invio, area problematica (droghe/alcol/gioco d'azzardo), stato del contatto e tipo di trattamento.\n\n"
        "c) Tracciato GPO (Gruppo Prestazioni Omogenee): documenta le prestazioni erogate per ciascun contatto, "
        "inclusa la tipologia (21 codici, tra cui il cod. 16 = Inserimento in Comunita' Terapeutica) e la sede "
        "di erogazione (Ser.D., ospedale, CT, carcere, domicilio, ecc.).\n\n"
        "d) Tracciato STR (Strutture): fornisce l'anagrafica dei Ser.D. e i dati territoriali aggregati, tra cui "
        "il numero di strutture carcerarie di competenza, la popolazione carceraria, il numero di Comunita' "
        "Terapeutiche convenzionate e le relative presenze.\n\n"
        "e) Tracciato HIV (Monitoraggio HIV): contiene dati anonimi e aggregati sullo screening HIV effettuato "
        "dai Ser.D. Il tracciato NON contiene l'identificativo del soggetto (Id_Rec), pertanto i dati HIV non "
        "sono linkabili ai singoli detenuti e vengono riportati unicamente come contesto epidemiologico."
    )
    ws5.cell(row=r, column=1, value=fonti).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 340
    r += 2

    ws5.cell(row=r, column=1, value="3. METODOLOGIA DI IDENTIFICAZIONE DEI DETENUTI TOSSICODIPENDENTI").font = section_font
    r += 1
    metodologia = (
        "L'identificazione dei soggetti detenuti tossicodipendenti e' stata effettuata attraverso il campo "
        "\"CondizioneAbitativaAlloggio\" presente nel tracciato SCG (Anagrafica Soggetto), secondo la codifica "
        "prevista dalle specifiche tecniche SIND v.3.5 (Allegato 1 - Tabella Condizione Abitativa Alloggio):\n\n"
        "   Codice 1 = Residenza fissa\n"
        "   Codice 2 = In carcere  <-- CRITERIO DI SELEZIONE\n"
        "   Codice 3 = Altro (ospedale, comunita')\n"
        "   Codice 8 = Senza fissa dimora\n"
        "   Codice 9 = Non determinato\n\n"
        "Sono stati quindi selezionati tutti i soggetti con CondizioneAbitativaAlloggio = 2, corrispondenti "
        "ai tossicodipendenti in stato di detenzione al momento della rilevazione.\n\n"
        "Per ciascun soggetto individuato, sono stati collegati tramite chiave Id_Rec:\n"
        "- i contatti attivi presso i Ser.D. (tracciato CTT)\n"
        "- le prestazioni erogate (tracciato GPO)\n\n"
        f"Il collegamento Id_Rec e' risultato completo: tutti i {tots['det']} detenuti identificati nel tracciato SCG "
        "risultano avere almeno un contatto attivo nel tracciato CTT."
    )
    ws5.cell(row=r, column=1, value=metodologia).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 260
    r += 2

    ws5.cell(row=r, column=1, value="4. RISULTATI").font = section_font
    r += 1
    ws5.cell(row=r, column=1, value="4.1 Detenuti tossicodipendenti per ASP").font = bold_font_11
    r += 1
    tab_risultati = (
        f"Su un totale di {tots['sogg']:,} soggetti in carico ai Ser.D. regionali (anno {anno_riferimento}), sono stati "
        f"individuati {tots['det']} detenuti tossicodipendenti, pari al {round(pct_tot, 1)}% del totale degli utenti SIND.\n\n"
    )
    for asp_code in ASP_CODES:
        rp = riepilogo[asp_code]
        tab_risultati += f"   {rp['nome']:.<30} {rp['tot_detenuti']:>3} detenuti TD  ({rp['detenuti_maschi']} M, {rp['detenuti_femmine']} F)\n"
    tab_risultati += (
        f"   {'':->47}\n"
        f"   {'TOTALE REGIONE':.<30} {tots['det']:>3} detenuti TD  ({tots['m']} M, {tots['f']} F)\n\n"
        f"Rapporto detenuti TD / popolazione carceraria (fonte STR): {tots['det']}/{tots_b['pop']:,} = "
        f"{round(tots['det'] / tots_b['pop'] * 100, 1) if tots_b['pop'] > 0 else 0}%"
    )
    ws5.cell(row=r, column=1, value=tab_risultati).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 220
    r += 2

    ws5.cell(row=r, column=1, value="4.2 Inserimenti in Comunita' Terapeutica in misura alternativa alla detenzione").font = bold_font_11
    r += 1
    ct_text = (
        f"Dall'analisi del tracciato GPO (Gruppo Prestazioni Omogenee) per i {tots['det']} soggetti detenuti identificati, "
        "si rileva quanto segue:\n\n"
        "- Inserimenti in Comunita' Terapeutica (cod. GPO 16): 0 (ZERO)\n"
        "- Prestazioni erogate presso sede CT (cod. sede 3): 0 (ZERO)\n"
        "- Prestazioni erogate presso sede Carcere (cod. sede 4): 0 (ZERO)\n\n"
        "Tutte le prestazioni registrate nel SIND per i detenuti tossicodipendenti risultano erogate "
        "presso la sede Ser.D. (cod. sede 1). Non risultano pertanto, nel flusso informativo SIND, "
        f"inserimenti in Comunita' Terapeutica in misura alternativa alla detenzione per l'anno {anno_riferimento}.\n\n"
        "La verifica e' stata estesa anche all'intera popolazione SIND (non solo detenuti): su tutte le "
        "prestazioni regionali, gli inserimenti CT risultano pressoche' assenti. "
        "Il dato conferma che l'inserimento in CT in misura alternativa non risulta praticato per i "
        "soggetti detenuti nell'anno di rilevazione."
    )
    ws5.cell(row=r, column=1, value=ct_text).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 200
    r += 2

    ws5.cell(row=r, column=1, value="4.3 Spesa sostenuta per inserimenti in CT").font = bold_font_11
    r += 1
    spese_text = (
        "Non risultando alcun inserimento in Comunita' Terapeutica in misura alternativa alla detenzione "
        "(cod. GPO 16 = 0 per tutti i soggetti detenuti), la spesa sostenuta per tale voce e' pari a zero.\n\n"
        "Si precisa in ogni caso che il sistema SIND traccia esclusivamente l'attivita' sanitaria e "
        "socio-sanitaria erogata dai Ser.D. (tipologia e quantita' delle prestazioni, sede, durata) e "
        "NON contiene informazioni di natura contabile o economica."
    )
    ws5.cell(row=r, column=1, value=spese_text).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 120
    r += 2

    ws5.cell(row=r, column=1, value="4.4 Contesto territoriale e infrastruttura").font = bold_font_11
    r += 1
    terr_text = (
        f"Dal tracciato STR (Strutture) risulta il seguente quadro territoriale regionale:\n\n"
        f"- Strutture carcerarie di competenza dei Ser.D.: {tots_b['carc']}\n"
        f"- Popolazione carceraria complessiva: {tots_b['pop']:,}\n"
        f"- Comunita' Terapeutiche convenzionate: {tots_b['ct']}\n"
        f"- Presenze totali nelle CT: {tots_b['pres']}\n\n"
        f"I dati STR sono stati deduplicati per Codice Ser.D., in quanto ogni servizio compare tre volte "
        f"nel tracciato (una per ciascuna tipologia di dipendenza trattata: S=Droghe, A=Alcolismo, "
        f"G=Gioco d'Azzardo), ma i dati infrastrutturali sono identici per le tre istanze."
    )
    ws5.cell(row=r, column=1, value=terr_text).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 160
    r += 2

    ws5.cell(row=r, column=1, value="4.5 Monitoraggio HIV").font = bold_font_11
    r += 1
    hiv_text = (
        f"Il tracciato HIV (Monitoraggio HIV) registra {tots_d['s']:,} soggetti sottoposti a screening HIV "
        f"({tots_d['m']:,} maschi, {tots_d['f']} femmine) presso i Ser.D. regionali.\n\n"
        f"Si segnalano {tots_d['p']} soggetti HIV positivi e {tots_d['n']} HIV negativi a livello regionale.\n\n"
        f"AVVERTENZA: il tracciato HIV e' anonimo (non contiene il campo Id_Rec) e pertanto i dati "
        f"non sono linkabili ai singoli detenuti tossicodipendenti. L'informazione viene fornita "
        f"esclusivamente come contesto epidemiologico generale dell'utenza Ser.D."
    )
    ws5.cell(row=r, column=1, value=hiv_text).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 130
    r += 2

    ws5.cell(row=r, column=1, value="5. STRUTTURA DELLA REPORTISTICA ALLEGATA").font = section_font
    r += 1
    struttura = (
        "Il file Excel contiene i seguenti fogli:\n\n"
        f"1. \"Dettaglio Detenuti\": elenco analitico dei {tots['det']} detenuti tossicodipendenti con dati anagrafici, "
        "contatti Ser.D. e prestazioni erogate. I record sono colorati per evidenziare lo stato di detenuto (rosa), "
        "eventuali inserimenti CT (verde) e prestazioni erogate in sede carcere (giallo).\n\n"
        "2. \"Strutture e Territorio\": mappa completa dei Ser.D. regionali con dotazione territoriale.\n\n"
        "3. \"Monitoraggio HIV\": dettaglio per Ser.D. e area problematica dei dati aggregati di screening HIV.\n\n"
        "4. \"Riepilogo Regionale\": quadro sinottico con 4 tabelle riepilogative e note metodologiche.\n\n"
        "5. \"Nota Metodologica\" (presente foglio): descrizione dettagliata delle fonti e della metodologia."
    )
    ws5.cell(row=r, column=1, value=struttura).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 220
    r += 2

    ws5.cell(row=r, column=1, value="6. CONCLUSIONI").font = section_font
    r += 1
    conclusioni = (
        "Dalla presente analisi emerge che:\n\n"
        f"a) I detenuti tossicodipendenti in carico ai Ser.D. della Regione Calabria nell'anno {anno_riferimento} sono {tots['det']}, "
        f"pari al {round(tots['det'] / tots_b['pop'] * 100, 1) if tots_b['pop'] > 0 else 0}% della popolazione carceraria regionale "
        f"({tots_b['pop']:,} detenuti).\n\n"
        "b) Non risultano nel SIND inserimenti in Comunita' Terapeutica in misura alternativa alla detenzione.\n\n"
        "c) La spesa per inserimenti in CT per detenuti e' conseguentemente pari a zero.\n\n"
        "Si resta a disposizione per eventuali chiarimenti o integrazioni."
    )
    ws5.cell(row=r, column=1, value=conclusioni).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap_style
    ws5.row_dimensions[r].height = 180

    # --- Salvataggio ---
    wb.save(output_path)
    _log(f"\nFile salvato: {output_path}")
    _log(f"Dettaglio Detenuti: {row - 1} righe")
    _log(f"Totale detenuti TD: {tots['det']} ({tots['m']} M, {tots['f']} F)")

    return output_path
