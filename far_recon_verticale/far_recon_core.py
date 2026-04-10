#!/usr/bin/env python3
"""Logica di conteggio FAR: acquisiti NSIS, nuovi FLUSSI, scarti SISR, proiezione.

Regole di conteggio (DM 17/12/2008, Specifiche FAR v6.3 § 4.3):
- T1: 1 record = 1 wrapper <FlsResSemires_1>.
      Chiave NSIS: (CodASL, CodStruttura, ID_REC, Data, TipoPrestazione).
- T2: 1 record = 1 evento (Tariffa|PrestazioniSR|Valutazione|Sospensione|Dimissione).
      Chiave NSIS per ogni tipo di evento (dedup):
        Tariffa       → (CodASL, CodStruttura, ID_REC, "Tariffa", Data)
        PrestazioniSR → (CodASL, CodStruttura, ID_REC, "PrestazioniSR", TempoPieno, TempoParziale)
        Valutazione   → (CodASL, CodStruttura, ID_REC, "Valutazione", Tipo, Data)
        Sospensione   → (CodASL, CodStruttura, ID_REC, "Sospensione", Data, DataFine)
        Dimissione    → (CodASL, CodStruttura, ID_REC, "Dimissione", Data)
- Scarti SISR: conteggio righe FLAG_OK_KO per Azienda × Trimestre × Tracciato
  (NO dedup globale per id_rec — logica Nicoletta).
"""

from __future__ import annotations

import os
import re
import shutil
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from typing import Any, Callable

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Costanti
# ---------------------------------------------------------------------------

AZIENDE = {
    "201": "201 ASP CS",
    "202": "202 ASP KR",
    "203": "203 ASP CZ",
    "204": "204 ASP VV",
    "205": "205 ASP RC",
}

T2_EVENTS = ("Tariffa", "PrestazioniSR", "Valutazione", "Sospensione", "Dimissione")

TAG1 = b"<FlsResSemires_1"
T2_WRAPPER_RE = re.compile(rb"<FlsResSemires_2[ >].*?</FlsResSemires_2>", re.DOTALL)
ID_REC_RE = re.compile(rb"<ID_REC>([^<]+)</ID_REC>")
NS_RE = re.compile(r"^\{[^}]+\}")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

LogFn = Callable[[str], None]


def _nolog(_msg: str) -> None:
    pass


def _local(tag: str) -> str:
    return NS_RE.sub("", tag)


def detect_azienda(path: str, fname: str) -> str:
    m = re.search(r"FAR(\d{3})", fname)
    if m and m.group(1) in AZIENDE:
        return AZIENDE[m.group(1)]
    m = re.search(r"FAR_[12]_(\d{3})_", fname)
    if m and m.group(1) in AZIENDE:
        return AZIENDE[m.group(1)]
    for seg in re.split(r"[\\/]", path):
        m2 = re.match(r"^(20[1-5])[ _]", seg)
        if m2:
            return AZIENDE[m2.group(1)]
    return "SCONOSCIUTO"


def detect_tracciato(fname: str) -> str | None:
    stem = os.path.splitext(fname)[0]
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
    m = re.search(r"FAR_([12])_", stem)
    if m:
        return m.group(1)
    base = stem.split("_")[0]
    if base and base[-1] in ("1", "2"):
        return base[-1]
    return None


def _text(el: ET.Element | None) -> str:
    return (el.text or "").strip() if el is not None else ""


def _find(parent: ET.Element, local_name: str) -> ET.Element | None:
    for child in parent:
        if _local(child.tag) == local_name:
            return child
    return None


# ---------------------------------------------------------------------------
# Preparazione: estrazione ZIP in cartella di lavoro temporanea
# ---------------------------------------------------------------------------

def prepare_workdir(
    source_dir: str,
    exclude_dirs: tuple[str, ...] = (),
    log: LogFn = _nolog,
) -> tuple[str, list[str]]:
    """Estrae ZIP e copia XML in una cartella di lavoro temporanea.

    Mantiene la struttura delle sottocartelle relativa a source_dir.
    Se due file (da ZIP diversi o da ZIP + plain) hanno lo stesso nome
    nella stessa directory, il secondo viene rinominato con suffisso _dupN
    e viene segnalata la collisione.

    Returns:
        (workdir_path, list_of_collision_warnings)
    """
    workdir = tempfile.mkdtemp(prefix="far_recon_")
    collisions: list[str] = []
    # Track which names are already placed in each output subdir
    placed: dict[str, set[str]] = defaultdict(set)  # rel_dir -> set of lowercase filenames

    for dirpath, dirnames, filenames in os.walk(source_dir):
        rel = os.path.relpath(dirpath, source_dir)
        skip = False
        for ex in exclude_dirs:
            if rel == ex or rel.startswith(ex + os.sep):
                skip = True
                break
        if skip:
            dirnames[:] = []
            continue

        out_subdir = os.path.join(workdir, rel) if rel != "." else workdir
        os.makedirs(out_subdir, exist_ok=True)

        for fn in filenames:
            if fn.startswith(".") or fn.endswith(".7z"):
                continue
            full = os.path.join(dirpath, fn)
            ext = os.path.splitext(fn)[1].lower()

            if ext == ".xml":
                dest_name = _safe_place(fn, out_subdir, rel, placed, collisions)
                shutil.copy2(full, os.path.join(out_subdir, dest_name))

            elif ext == ".zip":
                try:
                    with zipfile.ZipFile(full) as z:
                        for member in z.namelist():
                            m_ext = os.path.splitext(member)[1].lower()
                            if m_ext not in (".xml", ".xls"):
                                continue
                            m_basename = os.path.basename(member)
                            if not m_basename:
                                continue
                            dest_name = _safe_place(
                                m_basename, out_subdir, rel, placed, collisions,
                                source_zip=fn,
                            )
                            data = z.read(member)
                            with open(os.path.join(out_subdir, dest_name), "wb") as fout:
                                fout.write(data)
                except Exception as e:
                    log(f"  WARN: impossibile aprire {fn}: {e}")

    if collisions:
        log(f"  ATTENZIONE: {len(collisions)} collisione/i di nomi file rilevate:")
        for c in collisions:
            log(f"    {c}")
    else:
        log("  Nessuna collisione di nomi file.")

    return workdir, collisions


def _safe_place(
    filename: str,
    out_dir: str,
    rel_dir: str,
    placed: dict[str, set[str]],
    collisions: list[str],
    source_zip: str | None = None,
) -> str:
    """Restituisce il nome da usare in out_dir, rinominando se collisione."""
    key = filename.lower()
    if key not in placed[rel_dir]:
        placed[rel_dir].add(key)
        return filename

    # Collisione: trova un nome libero
    stem, ext = os.path.splitext(filename)
    n = 1
    while True:
        candidate = f"{stem}_dup{n}{ext}"
        if candidate.lower() not in placed[rel_dir]:
            break
        n += 1
    placed[rel_dir].add(candidate.lower())

    origin = f" (da {source_zip})" if source_zip else ""
    collisions.append(
        f"{filename}{origin} in {rel_dir}/ → rinominato {candidate}"
    )
    return candidate


def cleanup_workdir(workdir: str) -> None:
    """Rimuove la cartella di lavoro temporanea."""
    try:
        shutil.rmtree(workdir)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Conteggio grezzo (byte-level) — usato per vista rapida
# ---------------------------------------------------------------------------

def _new_row() -> dict[str, Any]:
    return {"t1": 0, "files_t1": 0, "files_t2": 0,
            "t2_idrec": set(), "t2_wrap": 0,
            **{n: 0 for n in T2_EVENTS}}


def _count_bytes(data: bytes) -> tuple[int, dict[str, int], set[bytes], int]:
    c1 = data.count(TAG1)
    ev: dict[str, int] = {}
    for name in T2_EVENTS:
        ev[name] = data.count(b"<" + name.encode() + b">") + data.count(b"<" + name.encode() + b" ")
    t2_ids: set[bytes] = set()
    wraps = T2_WRAPPER_RE.findall(data)
    for w in wraps:
        m = ID_REC_RE.search(w)
        if m:
            t2_ids.add(m.group(1))
    return c1, ev, t2_ids, len(wraps)


def _process_file_bytes(full_path: str) -> tuple[int, dict[str, int], set[bytes], int]:
    ext = os.path.splitext(full_path)[1].lower()
    try:
        if ext == ".zip":
            with zipfile.ZipFile(full_path) as z:
                c1 = 0
                ev_tot = {n: 0 for n in T2_EVENTS}
                ids: set[bytes] = set()
                wraps = 0
                for n in z.namelist():
                    if n.lower().endswith(".xml"):
                        with z.open(n) as f:
                            a, ev, t2ids, w = _count_bytes(f.read())
                            c1 += a
                            ids |= t2ids
                            wraps += w
                            for k in T2_EVENTS:
                                ev_tot[k] += ev[k]
                return c1, ev_tot, ids, wraps
        elif ext == ".xml":
            with open(full_path, "rb") as f:
                return _count_bytes(f.read())
    except Exception:
        pass
    return 0, {n: 0 for n in T2_EVENTS}, set(), 0


def scan_bytes(root: str, exclude_dirs: tuple[str, ...] = (), log: LogFn = _nolog) -> dict[str, dict]:
    """Scansione rapida byte-level (tag count). Per sheet acquisiti/nuovi."""
    agg: dict[str, dict] = defaultdict(_new_row)
    for dirpath, dirnames, filenames in os.walk(root):
        rel = os.path.relpath(dirpath, root)
        skip = False
        for ex in exclude_dirs:
            if rel == ex or rel.startswith(ex + os.sep):
                skip = True
                break
        if skip:
            dirnames[:] = []
            continue
        for fn in filenames:
            if fn.startswith(".") or fn.endswith(".7z"):
                continue
            if not (fn.lower().endswith(".zip") or fn.lower().endswith(".xml")):
                continue
            tr = detect_tracciato(fn)
            if tr not in ("1", "2"):
                continue
            full = os.path.join(dirpath, fn)
            az = detect_azienda(full, fn)
            c1, ev, ids, wraps = _process_file_bytes(full)
            row = agg[az]
            if tr == "1":
                row["t1"] += c1
                row["files_t1"] += 1
            else:
                for k in T2_EVENTS:
                    row[k] += ev[k]
                row["t2_idrec"] |= ids
                row["t2_wrap"] += wraps
                row["files_t2"] += 1
            log(f"  {fn} → {az} T{tr}")
    return dict(agg)


# ---------------------------------------------------------------------------
# Conteggio con chiave NSIS § 4.3 (dedup eventi) — per proiezione
# ---------------------------------------------------------------------------

def _find_deep(parent: ET.Element, local_name: str) -> ET.Element | None:
    """Cerca ricorsivamente un elemento per nome locale (senza namespace)."""
    for child in parent.iter():
        if _local(child.tag) == local_name:
            return child
    return None


def _get_chiave_fields(wrapper: ET.Element) -> tuple[str, str, str]:
    """Estrae (CodiceASL, CodiceStruttura, ID_REC) da Chiave > Erogatore."""
    chiave = _find(wrapper, "Chiave")
    if chiave is None:
        return ("", "", "")
    erog = _find(chiave, "Erogatore")
    cod_asl = _text(_find(erog, "CodiceASL")) if erog is not None else ""
    cod_str = _text(_find(erog, "CodiceStruttura")) if erog is not None else ""
    id_rec = _text(_find(chiave, "ID_REC"))
    return (cod_asl, cod_str, id_rec)


def _extract_event_keys(wrapper: ET.Element, cod_asl: str, cod_struttura: str, id_rec: str) -> list[tuple]:
    """Estrae le chiavi-evento da un wrapper FlsResSemires_2.
    Struttura XML reale: eventi (Tariffa, Valutazione, ...) sono figli diretti del wrapper.
    """
    keys: list[tuple] = []
    for child in wrapper:
        name = _local(child.tag)
        if name == "Tariffa":
            keys.append(("Tariffa", cod_asl, cod_struttura, id_rec,
                         _text(_find(child, "DataInizioTariffa"))))
        elif name == "PrestazioniSR":
            keys.append(("PrestazioniSR", cod_asl, cod_struttura, id_rec,
                         _text(_find(child, "TempoPieno")), _text(_find(child, "TempoParziale"))))
        elif name == "Valutazione":
            keys.append(("Valutazione", cod_asl, cod_struttura, id_rec,
                         _text(_find(child, "Tipo")), _text(_find(child, "Data"))))
        elif name == "Sospensione":
            keys.append(("Sospensione", cod_asl, cod_struttura, id_rec,
                         _text(_find(child, "Data")), _text(_find(child, "DataFine"))))
        elif name == "Dimissione":
            keys.append(("Dimissione", cod_asl, cod_struttura, id_rec,
                         _text(_find(child, "Data"))))
    return keys


def _parse_xml_keys(data: bytes) -> tuple[set[tuple], set[tuple]]:
    """Parsa un XML FAR e restituisce (chiavi_T1, chiavi_T2).
    Struttura XML: FlsResSemires_X > Chiave > Erogatore > CodiceASL/CodiceStruttura
                   FlsResSemires_X > Chiave > ID_REC/Data/tipoPrestazione
    """
    t1_keys: set[tuple] = set()
    t2_keys: set[tuple] = set()
    try:
        root_el = ET.fromstring(data)
    except ET.ParseError:
        return t1_keys, t2_keys

    for el in root_el.iter():
        lname = _local(el.tag)
        if lname == "FlsResSemires_1":
            cod_asl, cod_str, id_rec = _get_chiave_fields(el)
            chiave = _find(el, "Chiave")
            data_val = _text(_find(chiave, "Data")) if chiave is not None else ""
            tipo_pr = _text(_find(chiave, "tipoPrestazione")) if chiave is not None else ""
            t1_keys.add((cod_asl, cod_str, id_rec, data_val, tipo_pr))
        elif lname == "FlsResSemires_2":
            cod_asl, cod_str, id_rec = _get_chiave_fields(el)
            t2_keys.update(_extract_event_keys(el, cod_asl, cod_str, id_rec))

    return t1_keys, t2_keys


def scan_nsis_keys(root: str, exclude_dirs: tuple[str, ...] = (), log: LogFn = _nolog) -> dict[str, dict[str, set[tuple]]]:
    """Scansione con parsing XML completo per chiavi NSIS § 4.3.
    Ritorna {azienda: {"t1": set_chiavi, "t2": set_chiavi}}.
    """
    agg: dict[str, dict[str, set[tuple]]] = defaultdict(lambda: {"t1": set(), "t2": set()})
    for dirpath, dirnames, filenames in os.walk(root):
        rel = os.path.relpath(dirpath, root)
        skip = False
        for ex in exclude_dirs:
            if rel == ex or rel.startswith(ex + os.sep):
                skip = True
                break
        if skip:
            dirnames[:] = []
            continue
        for fn in filenames:
            if fn.startswith(".") or fn.endswith(".7z"):
                continue
            ext = os.path.splitext(fn)[1].lower()
            if ext not in (".zip", ".xml"):
                continue
            full = os.path.join(dirpath, fn)
            az = detect_azienda(full, fn)
            log(f"  [NSIS keys] {fn} → {az}")
            try:
                if ext == ".zip":
                    with zipfile.ZipFile(full) as z:
                        for n in z.namelist():
                            if n.lower().endswith(".xml"):
                                t1k, t2k = _parse_xml_keys(z.read(n))
                                agg[az]["t1"] |= t1k
                                agg[az]["t2"] |= t2k
                else:
                    with open(full, "rb") as f:
                        t1k, t2k = _parse_xml_keys(f.read())
                        agg[az]["t1"] |= t1k
                        agg[az]["t2"] |= t2k
            except Exception as e:
                log(f"  ERR {fn}: {e}")
    return dict(agg)


# ---------------------------------------------------------------------------
# Scarti SISR — per trimestre (logica Nicoletta)
# ---------------------------------------------------------------------------

def _read_xls_workbooks(full_path: str, log: LogFn = _nolog):
    """Genera coppie (nome_file, xlrd.Book) da un .zip (contenente .xls) o da un .xls plain."""
    ext = os.path.splitext(full_path)[1].lower()
    if ext == ".zip":
        try:
            with zipfile.ZipFile(full_path) as z:
                for n in z.namelist():
                    if not n.lower().endswith(".xls"):
                        continue
                    try:
                        yield n, xlrd.open_workbook(file_contents=z.read(n))
                    except Exception as e:
                        log(f"  ERR xls {full_path}/{n}: {e}")
        except Exception as e:
            log(f"  ERR {full_path}: {e}")
    elif ext == ".xls":
        try:
            yield os.path.basename(full_path), xlrd.open_workbook(full_path)
        except Exception as e:
            log(f"  ERR xls {full_path}: {e}")


def scan_scarti_per_trimestre(root: str, log: LogFn = _nolog) -> dict[tuple[str, str, str], dict[str, int]]:
    """Conta righe FLAG_OK_KO per (azienda, trimestre, tracciato) → {KO: n, OK: n}."""
    out: dict[tuple[str, str, str], dict[str, int]] = {}
    tri_re = re.compile(r"_T([1-4])")
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            fn_low = fn.lower()
            if not (fn_low.endswith(".zip") or fn_low.endswith(".xls")):
                continue
            tr = detect_tracciato(fn)
            if tr not in ("1", "2"):
                continue
            mt = tri_re.search(fn)
            if not mt:
                continue
            trim = "T" + mt.group(1)
            full = os.path.join(dirpath, fn)
            az = detect_azienda(full, fn)
            log(f"  [scarti] {fn} → {az} {trim} T{tr}")
            for _xls_name, wb in _read_xls_workbooks(full, log):
                for sn in wb.sheet_names():
                    s = wb.sheet_by_name(sn)
                    if s.nrows < 2:
                        continue
                    header = [str(c).strip() for c in s.row_values(0)]
                    try:
                        i_flag = header.index("FLAG_OK_KO")
                    except ValueError:
                        continue
                    for r in range(1, s.nrows):
                        flag = str(s.cell_value(r, i_flag)).strip()
                        key = (az, trim, tr)
                        d = out.setdefault(key, {"KO": 0, "OK": 0})
                        if flag in ("KO", "OK"):
                            d[flag] += 1
    return out


# ---------------------------------------------------------------------------
# Scarti dedup globale per id_rec (usata nel delta_validi)
# ---------------------------------------------------------------------------

def scan_scarti_dedup(root: str, log: LogFn = _nolog) -> dict[str, dict]:
    """Legge scarti .xls con dedup globale per id_rec. Per sheet delta_validi."""
    agg: dict[str, dict] = defaultdict(_new_row)
    id_re = re.compile(r"<ID_REC>([^<]+)</ID_REC>")
    seen: dict[tuple[str, str], set[str]] = defaultdict(set)
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            fn_low = fn.lower()
            if not (fn_low.endswith(".zip") or fn_low.endswith(".xls")):
                continue
            tr = detect_tracciato(fn)
            if tr not in ("1", "2"):
                continue
            full = os.path.join(dirpath, fn)
            az = detect_azienda(full, fn)
            for _xls_name, wb in _read_xls_workbooks(full, log):
                for sn in wb.sheet_names():
                    s = wb.sheet_by_name(sn)
                    if s.nrows < 2:
                        continue
                    header = [str(c).strip() for c in s.row_values(0)]
                    try:
                        i_flag = header.index("FLAG_OK_KO")
                        i_xml = header.index("XMLDOC")
                    except ValueError:
                        continue
                    for r in range(1, s.nrows):
                        if str(s.cell_value(r, i_flag)).strip() != "KO":
                            continue
                        xml_str = str(s.cell_value(r, i_xml))
                        m = id_re.search(xml_str)
                        key = (az, tr)
                        if m:
                            rid = m.group(1)
                            if rid in seen[key]:
                                continue
                            seen[key].add(rid)
                        row = agg[az]
                        if tr == "1":
                            row["t1"] += 1
                        else:
                            row["t2_wrap"] += 1
                            if m:
                                row["t2_idrec"].add(m.group(1).encode())
                            xb = xml_str.encode("utf-8", "ignore")
                            for k in T2_EVENTS:
                                row[k] += (xb.count(b"<" + k.encode() + b">")
                                           + xb.count(b"<" + k.encode() + b" "))
    return dict(agg)


# ---------------------------------------------------------------------------
# Proiezione post-upload
# ---------------------------------------------------------------------------

def compute_projection(
    nsis_keys: dict[str, dict[str, set[tuple]]],
    nuovi_keys: dict[str, dict[str, set[tuple]]],
) -> list[dict[str, Any]]:
    """Per ogni azienda calcola NSIS, Nuovi, overlap, veri nuovi, post-upload per T1 e T2."""
    all_az = sorted(set(nsis_keys) | set(nuovi_keys))
    rows: list[dict[str, Any]] = []
    for az in all_az:
        nk = nsis_keys.get(az, {"t1": set(), "t2": set()})
        nuk = nuovi_keys.get(az, {"t1": set(), "t2": set()})
        overlap_t1 = nk["t1"] & nuk["t1"]
        overlap_t2 = nk["t2"] & nuk["t2"]
        veri_t1 = len(nuk["t1"]) - len(overlap_t1)
        veri_t2 = len(nuk["t2"]) - len(overlap_t2)
        rows.append({
            "azienda": az,
            "nsis_t1": len(nk["t1"]), "nuovi_t1": len(nuk["t1"]),
            "overlap_t1": len(overlap_t1), "veri_nuovi_t1": veri_t1,
            "post_t1": len(nk["t1"]) + veri_t1,
            "nsis_t2": len(nk["t2"]), "nuovi_t2": len(nuk["t2"]),
            "overlap_t2": len(overlap_t2), "veri_nuovi_t2": veri_t2,
            "post_t2": len(nk["t2"]) + veri_t2,
        })
    return rows


# ---------------------------------------------------------------------------
# Scrittura Excel
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_TOTAL_FONT = Font(bold=True)
_ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_total(ws) -> None:
    for cell in ws[ws.max_row]:
        cell.font = _TOTAL_FONT


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def write_count_sheet(ws, title: str, agg: dict[str, dict], scarti: dict[str, dict] | None = None) -> None:
    ws.title = title
    headers = ["Azienda", "Record T1", "Tariffa", "PrestazioniSR",
               "Valutazione", "Sospensione", "Dimissione",
               "Tot Record T2 (eventi)",
               "T2 ID_REC distinti", "T2 ID_REC con ripetizioni"]
    if scarti is not None:
        headers.append("T2 ID_REC con rip. al netto scarti")
    headers += ["File T1", "File T2"]
    ws.append(headers)
    _style_header(ws)
    tot = {h: 0 for h in headers[1:]}
    for az in sorted(agg.keys()):
        r = agg[az]
        t2_tot = sum(r[k] for k in T2_EVENTS)
        t2_ids = len(r["t2_idrec"])
        row = [az, r["t1"], r["Tariffa"], r["PrestazioniSR"], r["Valutazione"],
               r["Sospensione"], r["Dimissione"], t2_tot, t2_ids, r["t2_wrap"]]
        if scarti is not None:
            s = scarti.get(az, _new_row())
            row.append(r["t2_wrap"] - s["t2_wrap"])
        row += [r["files_t1"], r["files_t2"]]
        ws.append(row)
        for h, v in zip(headers[1:], row[1:]):
            tot[h] += v
    ws.append(["TOTALE"] + [tot[h] for h in headers[1:]])
    _style_total(ws)
    _auto_width(ws)


def write_scarti_trim_sheet(ws, dati: dict[tuple[str, str, str], dict[str, int]]) -> None:
    ws.title = "scarti_per_trimestre"
    ws.append(["Azienda", "Trimestre", "Tracciato",
               "Scartati SISR (KO)", "Validi SISR (OK)", "Totale SISR"])
    _style_header(ws)
    tot = [0, 0, 0]
    for key in sorted(dati.keys()):
        az, trim, tr = key
        d = dati[key]
        totale = d["KO"] + d["OK"]
        ws.append([az, trim, tr, d["KO"], d["OK"], totale])
        tot[0] += d["KO"]
        tot[1] += d["OK"]
        tot[2] += totale
    ws.append(["TOTALE", "", "", *tot])
    _style_total(ws)
    _auto_width(ws)


def write_delta_sheet(ws, nuovi: dict[str, dict], scarti: dict[str, dict]) -> None:
    ws.title = "delta_validi"
    ws.append(["Azienda",
               "T1 inviati", "T1 scarti", "T1 validi",
               "T2 eventi inviati", "T2 eventi scarti", "T2 eventi validi",
               "T2 ID_REC distinti inv.", "T2 ID_REC distinti scart.", "T2 ID_REC distinti validi",
               "T2 ID_REC con rip. inv.", "T2 ID_REC con rip. scart.", "T2 ID_REC con rip. validi"])
    _style_header(ws)
    aziende = sorted(set(nuovi) | set(scarti))
    tot = [0] * 12
    for az in aziende:
        n = nuovi.get(az, _new_row())
        s = scarti.get(az, _new_row())
        n_t2 = sum(n[k] for k in T2_EVENTS)
        s_t2 = sum(s[k] for k in T2_EVENTS)
        n_id = len(n["t2_idrec"])
        s_id = len(s["t2_idrec"])
        n_wr = n["t2_wrap"]
        s_wr = s["t2_wrap"]
        row = [az, n["t1"], s["t1"], n["t1"] - s["t1"],
               n_t2, s_t2, n_t2 - s_t2,
               n_id, s_id, n_id - s_id,
               n_wr, s_wr, n_wr - s_wr]
        ws.append(row)
        for i, v in enumerate(row[1:]):
            tot[i] += v
    ws.append(["TOTALE"] + tot)
    _style_total(ws)
    _auto_width(ws)


def write_projection_sheet(ws, proj: list[dict[str, Any]]) -> None:
    ws.title = "proiezione_post_upload"
    ws.append([
        "Azienda",
        "NSIS T1", "Nuovi T1", "Già in NSIS T1", "Veri nuovi T1", "Post-upload T1",
        "NSIS T2", "Nuovi T2", "Già in NSIS T2", "Veri nuovi T2", "Post-upload T2",
    ])
    _style_header(ws)
    tot = [0] * 10
    for r in proj:
        row = [r["azienda"],
               r["nsis_t1"], r["nuovi_t1"], r["overlap_t1"], r["veri_nuovi_t1"], r["post_t1"],
               r["nsis_t2"], r["nuovi_t2"], r["overlap_t2"], r["veri_nuovi_t2"], r["post_t2"]]
        ws.append(row)
        for i, v in enumerate(row[1:]):
            tot[i] += v
    ws.append(["TOTALE"] + tot)
    _style_total(ws)
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Flusso completo
# ---------------------------------------------------------------------------

def generate_report(
    *,
    nsis_dir: str | None = None,
    nuovi_dir: str | None = None,
    scarti_dir: str | None = None,
    output_path: str,
    log: LogFn = _nolog,
) -> str:
    """Genera il report Excel completo.

    Fase 1: Preparazione — estrae ZIP in cartella temporanea, rileva collisioni nomi.
    Fase 2: Scansione e conteggio sui file preparati.
    Fase 3: Scrittura Excel e cleanup.

    Args:
        nsis_dir: cartella acquisiti_NSIS (opzionale).
        nuovi_dir: cartella nuovi_FLUSSI — gli scarti vengono cercati in <nuovi_dir>/scarti/
                   se scarti_dir non è specificato.
        scarti_dir: cartella scarti SISR (opzionale, default <nuovi_dir>/scarti/).
        output_path: percorso file .xlsx output.
        log: callback per messaggi di log.
    Returns:
        Percorso del file generato.
    """
    # Scarti dir default
    if scarti_dir is None and nuovi_dir:
        candidate = os.path.join(nuovi_dir, "scarti")
        if os.path.isdir(candidate):
            scarti_dir = candidate

    # --- Fase 1: preparazione workdir ---
    workdirs: list[str] = []
    all_collisions: list[str] = []
    nsis_work: str | None = None
    nuovi_work: str | None = None
    scarti_work: str | None = None

    try:
        if nsis_dir and os.path.isdir(nsis_dir):
            log("Preparazione acquisiti_NSIS (estrazione ZIP)...")
            nsis_work, cols = prepare_workdir(nsis_dir, log=log)
            workdirs.append(nsis_work)
            all_collisions.extend(cols)

        if nuovi_dir and os.path.isdir(nuovi_dir):
            log("Preparazione nuovi_FLUSSI (estrazione ZIP)...")
            nuovi_work, cols = prepare_workdir(nuovi_dir, exclude_dirs=("scarti",), log=log)
            workdirs.append(nuovi_work)
            all_collisions.extend(cols)

        if scarti_dir and os.path.isdir(scarti_dir):
            log("Preparazione scarti SISR (estrazione ZIP)...")
            scarti_work, cols = prepare_workdir(scarti_dir, log=log)
            workdirs.append(scarti_work)
            all_collisions.extend(cols)

        # --- Fase 2: scansione e conteggio ---
        wb = Workbook()
        sheet_idx = 0

        # Sheet collisioni (se presenti)
        if all_collisions:
            ws_col = wb.active if sheet_idx == 0 else wb.create_sheet()
            ws_col.title = "collisioni_nomi"
            ws_col.append(["File originale", "Cartella", "Rinominato in"])
            _style_header(ws_col)
            for c in all_collisions:
                # parse "filename (da zip) in dir/ → rinominato candidate"
                parts = c.split(" → rinominato ")
                if len(parts) == 2:
                    ws_col.append([parts[0], "", parts[1]])
                else:
                    ws_col.append([c, "", ""])
            _auto_width(ws_col)
            sheet_idx += 1

        # Sheet acquisiti_NSIS
        acq: dict[str, dict] = {}
        if nsis_work:
            log("Scansione acquisiti_NSIS (byte-level)...")
            acq = scan_bytes(nsis_work, log=log)
            ws = wb.active if sheet_idx == 0 else wb.create_sheet()
            write_count_sheet(ws, "acquisiti_NSIS", acq)
            sheet_idx += 1

        # Sheet nuovi_FLUSSI
        nuovi: dict[str, dict] = {}
        scarti_dedup: dict[str, dict] = {}
        if nuovi_work:
            log("Scansione nuovi_FLUSSI (byte-level)...")
            nuovi = scan_bytes(nuovi_work, log=log)
            # scarti dedup
            if scarti_work:
                log("Scansione scarti SISR (dedup globale)...")
                scarti_dedup = scan_scarti_dedup(scarti_work, log=log)
            ws = wb.active if sheet_idx == 0 else wb.create_sheet()
            write_count_sheet(ws, "nuovi_FLUSSI", nuovi, scarti=scarti_dedup)
            sheet_idx += 1

            # Sheet scarti
            if scarti_dedup:
                ws = wb.create_sheet()
                write_count_sheet(ws, "scarti_nuovi_FLUSSI", scarti_dedup)
                sheet_idx += 1

            # Sheet delta_validi
            if scarti_dedup:
                ws = wb.create_sheet()
                write_delta_sheet(ws, nuovi, scarti_dedup)
                sheet_idx += 1

            # Sheet scarti_per_trimestre
            if scarti_work:
                log("Scansione scarti per trimestre (logica Nicoletta)...")
                trim = scan_scarti_per_trimestre(scarti_work, log=log)
                if trim:
                    ws = wb.create_sheet()
                    write_scarti_trim_sheet(ws, trim)
                    sheet_idx += 1

        # Sheet proiezione (se entrambe le sorgenti presenti)
        if nsis_work and nuovi_work:
            log("Parsing XML per chiavi NSIS § 4.3 (acquisiti)...")
            nsis_keys = scan_nsis_keys(nsis_work, log=log)
            log("Parsing XML per chiavi NSIS § 4.3 (nuovi)...")
            nuovi_keys = scan_nsis_keys(nuovi_work, log=log)
            proj = compute_projection(nsis_keys, nuovi_keys)
            ws = wb.active if sheet_idx == 0 else wb.create_sheet()
            write_projection_sheet(ws, proj)
            sheet_idx += 1

        # Fallback
        if sheet_idx == 0:
            ws = wb.active
            ws.title = "info"
            ws.append(["Nessuna cartella valida selezionata."])

        wb.save(output_path)
        log(f"Report salvato: {output_path}")
        return output_path

    finally:
        # --- Fase 3: cleanup workdir ---
        for wd in workdirs:
            cleanup_workdir(wd)
