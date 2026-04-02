from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Callable

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


FLOW_NAMES = ("attiva", "passiva")
ORDERED_EROGATRICI = ["201", "202", "203", "204", "205", "912", "914", "915", "916"]
ORDER_MAP = {code: index for index, code in enumerate(ORDERED_EROGATRICI)}
EROGATRICE_LABELS = {
    "201": "180201 - ASP COSENZA",
    "202": "180202 - ASP CROTONE",
    "203": "180203 - ASP CATANZARO",
    "204": "180204 - ASP VIBO VALENTIA",
    "205": "180205 - ASP REGGIO CALABRIA",
    "912": "912 - AO ANNUNZIATA - COSENZA",
    "914": "914 - AOU RENATO DULBECCO - CATANZARO",
    "915": "915 - AO BIANCHI MELACRINO MORELLI GOM - REGGIO CALABRIA",
    "916": "916 - INRCA",
}
FILE_STEM_RE = re.compile(r"^(?P<code>\d+)\s+(?P<kind>importi|prestazioni)$", re.IGNORECASE)
HEADER_FARMACEUTICA = "FARMACEUTICA"
HEADER_SOMM_DIRETTA = "SOMM. DIRETTA DI FARMACI"
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ALT_FILL = PatternFill("solid", fgColor="F7FBFF")
SUBTOTAL_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="B8CCE4")
AMOUNT_QUANTIZER = Decimal("0.01")
MAX_VALIDATION_ERRORS = 20


@dataclass(frozen=True)
class ColumnDefinition:
    key: str
    label: str
    aliases: tuple[str, ...]


COLUMN_DEFINITIONS = (
    ColumnDefinition("ricoveri", "RICOV. ORD. E DAY HOSPITAL", ("RICOV. ORD. E DAY HOSPITAL", "RICOVERI ORD. E DAY  HOSPITAL")),
    ColumnDefinition("medicina_generale", "MEDICINA GENERALE", ("MEDICINA GENERALE",)),
    ColumnDefinition("specialistica_ambulatoriale", "SPECIALISTICA AMBULATORIALE", ("SPECIALISTICAAMBULATORIALE", "SPECIALISTICA AMBULATORIALE")),
    ColumnDefinition("farmaceutica", "FARMACEUTICA", ("FARMACEUTICA",)),
    ColumnDefinition("cure_termali", "CURE TERMALI", ("CURE TERMALI",)),
    ColumnDefinition("somm_diretta_farmaci", "SOMM. DIRETTA DI FARMACI", ("SOMM. DIRETTA DI FARMACI", "SOMM. DIRETTA DI  FARMACI")),
    ColumnDefinition("trasporti", "TRASPORTI", ("TRASPORTI CON AMBULANZA", "TRASPORTI")),
    ColumnDefinition("riabilitazione", "RIABILITAZIONE", ("RIABILITAZIONE", "RIABILITAZIONI")),
)
COLUMN_DEFINITION_BY_KEY = {definition.key: definition for definition in COLUMN_DEFINITIONS}
DEFAULT_SELECTED_COLUMN_KEYS = ("farmaceutica", "somm_diretta_farmaci")


@dataclass(frozen=True)
class FlowRecord:
    azienda_erogatrice: str
    controparte: str
    importi_by_column: dict[str, Decimal]
    prestazioni_by_column: dict[str, int]


@dataclass(frozen=True)
class FlowBuildResult:
    sheet_name: str
    counterparty_header: str
    erogatrici: list[str]
    selected_columns: list[ColumnDefinition]
    records: list[FlowRecord]


@dataclass(frozen=True)
class AmountValidationSummary:
    sheet_name: str
    checked_rows: int
    subtotal_rows: int
    column_keys: tuple[str, ...]
    totals_by_column: dict[str, Decimal]


def normalize_header(value: str) -> str:
    return " ".join(value.strip().upper().split())


def parse_decimal_italian(value: str) -> Decimal:
    text = value.strip().replace('"', "")
    if not text:
        return Decimal("0")
    normalized = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation as exc:
        raise ValueError(f"Valore numerico non riconosciuto: {value!r}") from exc


def decimal_to_excel_number(value: Decimal) -> float:
    return float(value.quantize(AMOUNT_QUANTIZER))


def quantize_amount(value: Decimal) -> Decimal:
    return value.quantize(AMOUNT_QUANTIZER)


def get_selected_columns(selected_column_keys: list[str] | tuple[str, ...] | None) -> list[ColumnDefinition]:
    keys = list(selected_column_keys or DEFAULT_SELECTED_COLUMN_KEYS)
    if not keys:
        raise ValueError("Seleziona almeno una tipologia da riportare nel report.")

    selected: list[ColumnDefinition] = []
    seen: set[str] = set()
    for key in keys:
        definition = COLUMN_DEFINITION_BY_KEY.get(key)
        if definition is None:
            raise ValueError(f"Tipologia non riconosciuta: {key}")
        if key in seen:
            continue
        selected.append(definition)
        seen.add(key)
    return selected


def parse_counterparty_sort_key(label: str) -> tuple[int, int | str]:
    match = re.match(r"\s*(\d+)", label)
    if match:
        return (0, int(match.group(1)))
    return (1, label.upper())


def parse_erogatrice_sort_key(code: str) -> tuple[int, int | str]:
    if code in ORDER_MAP:
        return (ORDER_MAP[code], int(code))
    if code.isdigit():
        return (len(ORDER_MAP), int(code))
    return (len(ORDER_MAP) + 1, code.upper())


def translate_erogatrice(code: str) -> str:
    return EROGATRICE_LABELS.get(code, code)


def display_counterparty_header(header: str) -> str:
    if normalize_header(header) == "ASL":
        return "AZIENDA SANITARIA DEBITRICE"
    return header


def next_output_path(base_dir: Path, stem: str = "report_mobilita_farmaci", ext: str = ".xlsx") -> Path:
    candidate = base_dir / f"{stem}{ext}"
    if not candidate.exists():
        return candidate

    index = 1
    while True:
        candidate = base_dir / f"{stem}_{index}{ext}"
        if not candidate.exists():
            return candidate
        index += 1


def ensure_available_output_path(output_path: Path) -> Path:
    if not output_path.exists():
        return output_path

    stem = output_path.stem
    suffix = output_path.suffix or ".xlsx"
    index = 1
    while True:
        candidate = output_path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def read_csv_file(path: Path) -> tuple[list[str], list[list[str]]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle, delimiter=";")
                header = next(reader)
                rows = [row for row in reader if any(cell.strip() for cell in row)]
                return header, rows
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    raise ValueError(f"File CSV vuoto: {path}")


def get_column_index(header: list[str], definition: ColumnDefinition, csv_path: Path) -> int:
    normalized_header_map = {
        normalize_header(column_name): index for index, column_name in enumerate(header)
    }
    for alias in definition.aliases:
        index = normalized_header_map.get(normalize_header(alias))
        if index is not None:
            return index
    raise ValueError(
        f"Colonna {definition.label!r} non trovata in {csv_path.name}. "
        f"Header disponibili: {', '.join(header)}"
    )


def extract_metrics(
    path: Path,
    selected_columns: list[ColumnDefinition],
) -> tuple[str, dict[str, dict[str, Decimal]]]:
    header, rows = read_csv_file(path)
    if not header:
        raise ValueError(f"Header mancante in {path}")

    first_header = header[0].strip()
    column_indexes = {
        definition.key: get_column_index(header, definition, path)
        for definition in selected_columns
    }

    data: dict[str, dict[str, Decimal]] = {}
    for row_number, row in enumerate(rows, start=2):
        padded_row = row + [""] * (len(header) - len(row))
        counterparty = padded_row[0].strip()
        if not counterparty:
            continue
        if normalize_header(counterparty).startswith("TOTALE"):
            continue
        if counterparty in data:
            raise ValueError(f"Valore duplicato nella prima colonna di {path.name} alla riga {row_number}: {counterparty}")
        data[counterparty] = {
            definition.key: parse_decimal_italian(padded_row[column_indexes[definition.key]])
            for definition in selected_columns
        }
    return first_header, data


def collect_flow_pairs(flow_dir: Path) -> tuple[dict[str, dict[str, Path]], list[str]]:
    pairs: dict[str, dict[str, Path]] = {}
    invalid_files: list[str] = []

    for path in sorted(flow_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() != ".csv":
            continue
        match = FILE_STEM_RE.match(path.stem.strip())
        if match is None:
            invalid_files.append(path.name)
            continue
        code = match.group("code")
        kind = match.group("kind").lower()
        pairs.setdefault(code, {})
        if kind in pairs[code]:
            raise ValueError(f"File duplicato per {flow_dir.name} {code} {kind}: {path.name}")
        pairs[code][kind] = path

    return pairs, invalid_files


def validate_flow_directory(flow_name: str, flow_dir: str | Path) -> list[str]:
    directory = Path(flow_dir).expanduser()
    errors: list[str] = []

    if not directory.exists() or not directory.is_dir():
        return [f"La cartella {flow_name!r} non esiste: {directory}"]

    try:
        pairs, invalid_files = collect_flow_pairs(directory)
    except ValueError as exc:
        return [str(exc)]

    if invalid_files:
        errors.append(
            f"{flow_name}: nomi file non riconosciuti: {', '.join(sorted(invalid_files))}"
        )
    if not pairs:
        errors.append(f"{flow_name}: nessuna coppia CSV trovata.")
        return errors

    for code, file_map in sorted(pairs.items(), key=lambda item: parse_erogatrice_sort_key(item[0])):
        missing_kinds = {"importi", "prestazioni"} - set(file_map)
        if missing_kinds:
            errors.append(
                f"{flow_name}: per l'azienda erogatrice {code} manca/no il/i file/i {', '.join(sorted(missing_kinds))}."
            )

    return errors


def validate_input_directories(attiva_dir: str | Path, passiva_dir: str | Path) -> list[str]:
    errors: list[str] = []
    errors.extend(validate_flow_directory("attiva", attiva_dir))
    errors.extend(validate_flow_directory("passiva", passiva_dir))
    return errors


def validate_input_folder(base_dir: str | Path) -> list[str]:
    root = Path(base_dir).expanduser()
    errors: list[str] = []

    if not root.exists() or not root.is_dir():
        return [f"La cartella selezionata non esiste: {root}"]

    errors.extend(validate_input_directories(root / "attiva", root / "passiva"))
    return errors


def build_flow_records(
    flow_name: str,
    flow_dir: Path,
    selected_column_keys: list[str] | tuple[str, ...] | None = None,
) -> FlowBuildResult:
    selected_columns = get_selected_columns(selected_column_keys)
    pairs, invalid_files = collect_flow_pairs(flow_dir)
    if invalid_files:
        raise ValueError(
            f"{flow_name}: nomi file non riconosciuti: {', '.join(sorted(invalid_files))}"
        )
    if not pairs:
        raise ValueError(f"{flow_name}: nessuna coppia CSV disponibile.")

    counterparty_header: str | None = None
    records: list[FlowRecord] = []
    erogatrici = sorted(pairs, key=parse_erogatrice_sort_key)

    for erogatrice in erogatrici:
        files = pairs[erogatrice]
        missing_kinds = {"importi", "prestazioni"} - set(files)
        if missing_kinds:
            raise ValueError(
                f"{flow_name}: per l'azienda erogatrice {erogatrice} manca/no il/i file/i {', '.join(sorted(missing_kinds))}."
            )

        importi_header, importi_data = extract_metrics(files["importi"], selected_columns)
        prestazioni_header, prestazioni_data = extract_metrics(files["prestazioni"], selected_columns)

        if importi_header != prestazioni_header:
            raise ValueError(
                f"{flow_name}: la prima colonna non coincide tra {files['importi'].name} e {files['prestazioni'].name}."
            )
        if counterparty_header is None:
            counterparty_header = importi_header
        elif counterparty_header != importi_header:
            raise ValueError(
                f"{flow_name}: header della prima colonna incoerente. Atteso {counterparty_header!r}, trovato {importi_header!r}."
            )

        missing_in_prestazioni = sorted(set(importi_data) - set(prestazioni_data))
        missing_in_importi = sorted(set(prestazioni_data) - set(importi_data))
        if missing_in_prestazioni or missing_in_importi:
            details: list[str] = []
            if missing_in_prestazioni:
                details.append(f"presenti solo in importi: {', '.join(missing_in_prestazioni)}")
            if missing_in_importi:
                details.append(f"presenti solo in prestazioni: {', '.join(missing_in_importi)}")
            raise ValueError(
                f"{flow_name}: righe non allineate per l'azienda erogatrice {erogatrice} ({'; '.join(details)})."
            )

        for controparte in sorted(importi_data, key=parse_counterparty_sort_key):
            importi_values = importi_data[controparte]
            prestazioni_values = prestazioni_data[controparte]
            records.append(
                FlowRecord(
                    azienda_erogatrice=translate_erogatrice(erogatrice),
                    controparte=controparte,
                    importi_by_column={
                        definition.key: importi_values[definition.key]
                        for definition in selected_columns
                    },
                    prestazioni_by_column={
                        definition.key: int(prestazioni_values[definition.key])
                        for definition in selected_columns
                    },
                )
            )

    if counterparty_header is None:
        raise ValueError(f"{flow_name}: impossibile determinare l'header della prima colonna.")

    return FlowBuildResult(
        sheet_name=flow_name,
        counterparty_header=counterparty_header,
        erogatrici=erogatrici,
        selected_columns=selected_columns,
        records=records,
    )


def style_sheet(worksheet, result: FlowBuildResult) -> None:
    headers = ["AZIENDA EROGATRICE", display_counterparty_header(result.counterparty_header)]
    for definition in result.selected_columns:
        headers.extend([f"{definition.label} - IMPORTI", f"{definition.label} - PRESTAZIONI"])
    worksheet.append(headers)

    subtotal_importi = {definition.key: Decimal("0") for definition in result.selected_columns}
    subtotal_prestazioni = {definition.key: 0 for definition in result.selected_columns}
    total_importi = {definition.key: Decimal("0") for definition in result.selected_columns}
    total_prestazioni = {definition.key: 0 for definition in result.selected_columns}
    current_erogatrice: str | None = None

    def append_subtotal_row(erogatrice: str) -> None:
        row = [erogatrice, "SUBTOTALE"]
        for definition in result.selected_columns:
            row.extend(
                [
                    decimal_to_excel_number(subtotal_importi[definition.key]),
                    subtotal_prestazioni[definition.key],
                ]
            )
        worksheet.append(row)

    for record in result.records:
        if current_erogatrice is None:
            current_erogatrice = record.azienda_erogatrice
        elif record.azienda_erogatrice != current_erogatrice:
            append_subtotal_row(current_erogatrice)
            subtotal_importi = {definition.key: Decimal("0") for definition in result.selected_columns}
            subtotal_prestazioni = {definition.key: 0 for definition in result.selected_columns}
            current_erogatrice = record.azienda_erogatrice

        row = [record.azienda_erogatrice, record.controparte]
        for definition in result.selected_columns:
            row.extend(
                [
                    decimal_to_excel_number(record.importi_by_column[definition.key]),
                    record.prestazioni_by_column[definition.key],
                ]
            )
            subtotal_importi[definition.key] += record.importi_by_column[definition.key]
            subtotal_prestazioni[definition.key] += record.prestazioni_by_column[definition.key]
            total_importi[definition.key] += record.importi_by_column[definition.key]
            total_prestazioni[definition.key] += record.prestazioni_by_column[definition.key]
        worksheet.append(row)

    if current_erogatrice is not None:
        append_subtotal_row(current_erogatrice)

    total_row = ["TOTALE GENERALE", ""]
    for definition in result.selected_columns:
        total_row.extend(
            [
                decimal_to_excel_number(total_importi[definition.key]),
                total_prestazioni[definition.key],
            ]
        )
    worksheet.append(total_row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, worksheet.max_row)}"
    worksheet.sheet_view.showGridLines = True

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        row_kind = "detail"
        if row[1].value == "SUBTOTALE":
            row_kind = "subtotal"
        elif row[0].value == "TOTALE GENERALE":
            row_kind = "total"

        is_alt = row_index % 2 == 0
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")
            if row_kind == "subtotal":
                cell.font = Font(bold=True)
                cell.fill = SUBTOTAL_FILL
            elif row_kind == "total":
                cell.font = Font(bold=True)
                cell.fill = TOTAL_FILL
            elif is_alt:
                cell.fill = ALT_FILL
        for column_index in range(3, len(headers) + 1, 2):
            worksheet.cell(row=row_index, column=column_index).number_format = "#,##0.00"
        for column_index in range(4, len(headers) + 1, 2):
            worksheet.cell(row=row_index, column=column_index).number_format = "#,##0"

    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 58
    for offset, definition in enumerate(result.selected_columns):
        import_col = get_column_letter(3 + offset * 2)
        prest_col = get_column_letter(4 + offset * 2)
        worksheet.column_dimensions[import_col].width = max(24, len(definition.label) + 6)
        worksheet.column_dimensions[prest_col].width = max(24, len(definition.label) + 6)


def build_expected_amount_rows(
    result: FlowBuildResult,
) -> tuple[list[tuple[object, ...]], AmountValidationSummary]:
    rows: list[tuple[object, ...]] = []
    subtotal_importi = {definition.key: Decimal("0") for definition in result.selected_columns}
    total_importi = {definition.key: Decimal("0") for definition in result.selected_columns}
    current_erogatrice: str | None = None
    subtotal_rows = 0

    for record in result.records:
        if current_erogatrice is None:
            current_erogatrice = record.azienda_erogatrice
        elif record.azienda_erogatrice != current_erogatrice:
            rows.append(
                (current_erogatrice, "SUBTOTALE")
                + tuple(quantize_amount(subtotal_importi[definition.key]) for definition in result.selected_columns)
            )
            subtotal_rows += 1
            subtotal_importi = {definition.key: Decimal("0") for definition in result.selected_columns}
            current_erogatrice = record.azienda_erogatrice

        rows.append(
            (record.azienda_erogatrice, record.controparte)
            + tuple(quantize_amount(record.importi_by_column[definition.key]) for definition in result.selected_columns)
        )
        for definition in result.selected_columns:
            subtotal_importi[definition.key] += record.importi_by_column[definition.key]
            total_importi[definition.key] += record.importi_by_column[definition.key]

    if current_erogatrice is not None:
        rows.append(
            (current_erogatrice, "SUBTOTALE")
            + tuple(quantize_amount(subtotal_importi[definition.key]) for definition in result.selected_columns)
        )
        subtotal_rows += 1

    rows.append(
        ("TOTALE GENERALE", "")
        + tuple(quantize_amount(total_importi[definition.key]) for definition in result.selected_columns)
    )

    return rows, AmountValidationSummary(
        sheet_name=result.sheet_name,
        checked_rows=len(rows),
        subtotal_rows=subtotal_rows,
        column_keys=tuple(definition.key for definition in result.selected_columns),
        totals_by_column={
            definition.key: quantize_amount(total_importi[definition.key])
            for definition in result.selected_columns
        },
    )


def validate_output_amounts(
    output_path: str | Path,
    results: dict[str, FlowBuildResult],
) -> list[AmountValidationSummary]:
    workbook = load_workbook(Path(output_path).expanduser(), data_only=True)
    errors: list[str] = []
    summaries: list[AmountValidationSummary] = []

    for flow_name in FLOW_NAMES:
        result = results[flow_name]
        expected_rows, summary = build_expected_amount_rows(result)
        summaries.append(summary)
        worksheet = workbook[flow_name]

        actual_rows: list[tuple[object, ...]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            amount_values = []
            for offset in range(len(result.selected_columns)):
                amount_values.append(quantize_amount(Decimal(str(row[2 + offset * 2]))))
            actual_rows.append(
                (str(row[0]), "" if row[1] is None else str(row[1])) + tuple(amount_values)
            )

        if len(actual_rows) != len(expected_rows):
            errors.append(
                f"{flow_name}: righe importi output={len(actual_rows)} diverse da attese={len(expected_rows)}."
            )

        for row_index, (expected_row, actual_row) in enumerate(zip(expected_rows, actual_rows), start=2):
            if expected_row != actual_row:
                errors.append(
                    f"{flow_name}: mismatch importi riga {row_index}: expected={expected_row} actual={actual_row}"
                )
                if len(errors) >= MAX_VALIDATION_ERRORS:
                    break
        if len(errors) >= MAX_VALIDATION_ERRORS:
            break

    workbook.close()

    if errors:
        raise ValueError(
            "Controllo importi input/output fallito.\n" + "\n".join(errors[:MAX_VALIDATION_ERRORS])
        )

    return summaries


def generate_report_from_directories(
    attiva_dir: str | Path,
    passiva_dir: str | Path,
    output_path: str | Path,
    selected_column_keys: list[str] | tuple[str, ...] | None = None,
    log: Callable[[str], None] | None = None,
) -> Path:
    attiva = Path(attiva_dir).expanduser()
    passiva = Path(passiva_dir).expanduser()
    output = Path(output_path).expanduser()
    selected_columns = get_selected_columns(selected_column_keys)

    errors = validate_input_directories(attiva, passiva)
    if errors:
        raise ValueError("\n".join(errors))

    if log:
        log(f"Cartella attiva: {attiva}")
        log(f"Cartella passiva: {passiva}")

    results = {
        "attiva": build_flow_records("attiva", attiva, [definition.key for definition in selected_columns]),
        "passiva": build_flow_records("passiva", passiva, [definition.key for definition in selected_columns]),
    }
    for flow_name in FLOW_NAMES:
        flow_result = results[flow_name]
        if log:
            log(
                f"{flow_name}: aziende erogatrici={len(flow_result.erogatrici)}, "
                f"righe report={len(flow_result.records)}, "
                f"tipologie={', '.join(definition.label for definition in flow_result.selected_columns)}."
            )

    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for flow_name in FLOW_NAMES:
        worksheet = workbook.create_sheet(flow_name)
        style_sheet(worksheet, results[flow_name])

    workbook.save(output)
    validation_summaries = validate_output_amounts(output, results)
    if log:
        log(f"Workbook creato: {output}")
        log("Controllo importi input/output: OK.")
        for summary in validation_summaries:
            totals_text = ", ".join(
                f"{COLUMN_DEFINITION_BY_KEY[key].label}={summary.totals_by_column[key]}"
                for key in summary.column_keys
            )
            log(
                f"{summary.sheet_name}: righe verificate={summary.checked_rows}, "
                f"subtotali={summary.subtotal_rows}, "
                f"totali importi -> {totals_text}."
            )
    return output


def generate_report(
    base_dir: str | Path,
    output_path: str | Path,
    selected_column_keys: list[str] | tuple[str, ...] | None = None,
    log: Callable[[str], None] | None = None,
) -> Path:
    root = Path(base_dir).expanduser()
    return generate_report_from_directories(
        root / "attiva",
        root / "passiva",
        output_path,
        selected_column_keys=selected_column_keys,
        log=log,
    )


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Genera il report Excel della mobilita infraregionale farmaci."
    )
    parser.add_argument("input_dir", help="Cartella che contiene le sottocartelle attiva/ e passiva/")
    parser.add_argument("output_file", help="Percorso del file Excel di output")
    return parser


def main() -> None:
    args = build_arg_parser().parse_args()
    output = generate_report(args.input_dir, args.output_file, selected_column_keys=DEFAULT_SELECTED_COLUMN_KEYS, log=print)
    print(output)


if __name__ == "__main__":
    main()
