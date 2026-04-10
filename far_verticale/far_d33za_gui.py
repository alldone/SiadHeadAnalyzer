#!/usr/bin/env python3

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
import json
import os
from pathlib import Path
import re
import subprocess
import sys
import threading
import traceback
from typing import Any
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


TRACK_TOKEN = "FAR"
TRACK1_ROOT = "Tracciato1"
TRACK2_ROOT = "Tracciato2"
TRACK1_RECORD = "FlsResSemires_1"
TRACK2_RECORD = "FlsResSemires_2"
ALLOWED_LEVELS = ("R1", "R2", "R3")
LEVEL_RANK = {"R1": 1, "R2": 2, "R3": 3}
XML_NS_RE = re.compile(r"^\{(?P<ns>[^}]+)\}(?P<name>.+)$")
ANAGRAFICA_FILE_RE = re.compile(r"Export_Assistiti_([A-Z]{2})_")
STATE_FILE = Path(__file__).with_name(".far_d33za_gui_state.json")

# Posizioni ricavate dal PDF del tracciato anagrafe.
ANAG_FIELDS = {
    "primary_code": (1, 16),
    "birth_date": (177, 186),
    "new_code": (208, 223),
    "old_code": (224, 239),
    "asl_iscrizione": (842, 847),
}
FALLBACK_ASL_BY_AREA = {
    "CS": "201",
    "KR": "202",
    "CZ": "203",
    "VV": "204",
    "RC": "205",
}


def local_name(tag: str) -> str:
    match = XML_NS_RE.match(tag)
    return match.group("name") if match else tag


def parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_ddmmyyyy_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%d/%m/%Y").date()
    except ValueError:
        return None


def normalize_code(value: str | None) -> str:
    return (value or "").strip().upper()


def normalize_digits(value: str | None) -> str:
    return "".join(ch for ch in (value or "") if ch.isdigit())


def normalize_asl_code(value: str | None, fallback: str = "") -> str:
    digits = normalize_digits(value)
    if len(digits) >= 3:
        return digits[-3:]
    return normalize_code(value) or fallback


def normalize_region_code(value: str | None) -> str:
    digits = normalize_digits(value)
    if len(digits) >= 6:
        return digits[:3]
    return ""


def age_on_reference_date(birth_date: date, reference_date: date) -> int:
    years = reference_date.year - birth_date.year
    if (reference_date.month, reference_date.day) < (birth_date.month, birth_date.day):
        years -= 1
    return years


def find_child(parent: ET.Element | None, name: str) -> ET.Element | None:
    if parent is None:
        return None
    for child in parent:
        if local_name(child.tag) == name:
            return child
    return None


def find_text(parent: ET.Element | None, path: list[str]) -> str | None:
    current = parent
    for name in path:
        current = find_child(current, name)
        if current is None:
            return None
    return current.text.strip() if current.text else None


def count_file_lines(path: Path) -> int:
    count = 0
    with path.open("r", encoding="utf-8", errors="ignore", newline=None) as handle:
        for _ in handle:
            count += 1
    if count > 0:
        return count
    return 1 if path.stat().st_size > 0 else 0


def file_size_mb(path: Path) -> float:
    return path.stat().st_size / (1024 * 1024)


def quarter_from_path(relative_path: str) -> str:
    for part in Path(relative_path).parts:
        if re.fullmatch(r"T[1-4]", part):
            return part
    return ""


def next_available_report_path(xml_dir: Path) -> Path:
    base_name = "report_far_d33za"
    candidate = xml_dir / f"{base_name}.xlsx"
    if not candidate.exists():
        return candidate
    index = 1
    while True:
        candidate = xml_dir / f"{base_name}_{index}.xlsx"
        if not candidate.exists():
            return candidate
        index += 1


def extract_last_16(value: str | None) -> str:
    normalized = normalize_code(value)
    return normalized[-16:] if len(normalized) >= 16 else normalized


def safe_fixed_slice(line: str, start: int, end: int) -> str:
    if len(line) < start:
        return ""
    return line[start - 1:end].strip()


@dataclass
class FarXmlFileInfo:
    path: Path
    relative_path: str
    track: int
    root_name: str
    line_count: int
    size_mb: float


@dataclass
class AnagrafeRecord:
    primary_code: str
    birth_date: date | None
    resident_asl: str
    resident_region: str
    source_file: str
    raw_asl_iscrizione: str


@dataclass
class FarRecord:
    track: int
    source_file: str
    quarter: str
    event_date: date
    erogatore_asl: str
    resident_region_xml: str
    resident_asl_xml: str
    id_rec: str
    cuni: str
    link_candidates: tuple[str, ...]
    tipo_prestazione: str
    anno_nascita_xml: int | None


def classify_far_xml_file(xml_path: Path) -> tuple[int, str]:
    root = ET.parse(xml_path).getroot()
    root_name = local_name(root.tag)
    if root_name == TRACK1_ROOT:
        return 1, root_name
    if root_name == TRACK2_ROOT:
        return 2, root_name
    raise ValueError(f"Root XML non riconosciuta per {xml_path}: {root_name}")


def scan_xml_files(base_dir: Path, filename_token: str = TRACK_TOKEN) -> list[FarXmlFileInfo]:
    pattern = f"{filename_token}*.xml" if filename_token else "*.xml"
    xml_files: list[FarXmlFileInfo] = []
    for path in sorted(base_dir.rglob(pattern)):
        if not path.is_file():
            continue
        track, root_name = classify_far_xml_file(path)
        xml_files.append(
            FarXmlFileInfo(
                path=path,
                relative_path=path.relative_to(base_dir).as_posix(),
                track=track,
                root_name=root_name,
                line_count=count_file_lines(path),
                size_mb=file_size_mb(path),
            )
        )
    return xml_files


def iter_track_records(root: ET.Element, expected_name: str):
    for child in root:
        if local_name(child.tag) == expected_name:
            yield child


def unique_candidates(*values: str) -> tuple[str, ...]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        normalized = normalize_code(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    return tuple(result)


def parse_track1_record(record: ET.Element, source_file: str) -> FarRecord | None:
    resident_region_xml = normalize_code(
        find_text(record, ["AssistitoAmmissione", "Assistito", "DatiAnagrafici", "Residenza", "Regione"])
    )
    resident_asl_xml = normalize_asl_code(
        find_text(record, ["AssistitoAmmissione", "Assistito", "DatiAnagrafici", "Residenza", "ASL"])
    )
    erogatore_asl = normalize_asl_code(find_text(record, ["Chiave", "Erogatore", "CodiceASL"]))
    id_rec = normalize_code(find_text(record, ["Chiave", "ID_REC"]))
    data_text = find_text(record, ["Chiave", "Data"])
    cuni = normalize_code(find_text(record, ["AssistitoAmmissione", "Assistito", "DatiAnagrafici", "CUNI"]))
    tipo_prestazione = normalize_code(find_text(record, ["Chiave", "tipoPrestazione"]))
    anno_nascita_text = find_text(record, ["AssistitoAmmissione", "Assistito", "DatiAnagrafici", "AnnoNascita"])

    event_date = parse_iso_date(data_text)
    if not erogatore_asl or not id_rec or event_date is None or not tipo_prestazione:
        return None

    anno_nascita = int(anno_nascita_text) if anno_nascita_text and anno_nascita_text.isdigit() else None
    return FarRecord(
        track=1,
        source_file=source_file,
        quarter=quarter_from_path(source_file),
        event_date=event_date,
        erogatore_asl=erogatore_asl,
        resident_region_xml=resident_region_xml,
        resident_asl_xml=resident_asl_xml,
        id_rec=id_rec,
        cuni=cuni,
        link_candidates=unique_candidates(cuni, extract_last_16(id_rec)),
        tipo_prestazione=tipo_prestazione,
        anno_nascita_xml=anno_nascita,
    )


def parse_track2_record(record: ET.Element, source_file: str) -> FarRecord | None:
    erogatore_asl = normalize_asl_code(find_text(record, ["Chiave", "Erogatore", "CodiceASL"]))
    id_rec = normalize_code(find_text(record, ["Chiave", "ID_REC"]))
    tipo_prestazione = normalize_code(find_text(record, ["Chiave", "tipoPrestazione"]))
    key_date = parse_iso_date(find_text(record, ["Chiave", "Data"]))
    valutazione_date = parse_iso_date(find_text(record, ["Valutazione", "Data"]))
    dimissione_date = parse_iso_date(find_text(record, ["Dimissione", "Data"]))
    event_date = max(
        [value for value in (valutazione_date, dimissione_date, key_date) if value is not None],
        default=None,
    )

    if not erogatore_asl or not id_rec or event_date is None or not tipo_prestazione:
        return None

    tail_code = extract_last_16(id_rec)
    return FarRecord(
        track=2,
        source_file=source_file,
        quarter=quarter_from_path(source_file),
        event_date=event_date,
        erogatore_asl=erogatore_asl,
        resident_region_xml="",
        resident_asl_xml="",
        id_rec=id_rec,
        cuni=tail_code,
        link_candidates=unique_candidates(tail_code),
        tipo_prestazione=tipo_prestazione,
        anno_nascita_xml=None,
    )


def parse_anagrafe_line(line: str, source_file: str, fallback_asl: str) -> tuple[AnagrafeRecord | None, tuple[str, ...]]:
    primary_code = normalize_code(safe_fixed_slice(line, *ANAG_FIELDS["primary_code"]))
    birth_date = parse_ddmmyyyy_date(safe_fixed_slice(line, *ANAG_FIELDS["birth_date"]))
    new_code = normalize_code(safe_fixed_slice(line, *ANAG_FIELDS["new_code"]))
    old_code = normalize_code(safe_fixed_slice(line, *ANAG_FIELDS["old_code"]))
    raw_asl = safe_fixed_slice(line, *ANAG_FIELDS["asl_iscrizione"])
    resident_asl = normalize_asl_code(raw_asl, fallback=fallback_asl)
    resident_region = normalize_region_code(raw_asl)

    if not primary_code:
        return None, ()

    record = AnagrafeRecord(
        primary_code=primary_code,
        birth_date=birth_date,
        resident_asl=resident_asl,
        resident_region=resident_region,
        source_file=source_file,
        raw_asl_iscrizione=raw_asl,
    )
    aliases = unique_candidates(primary_code, new_code, old_code)
    return record, aliases


def load_anagrafe_data(
    anagrafe_dir: Path,
    analysis_year: int,
) -> tuple[dict[str, AnagrafeRecord], list[dict[str, object]], dict[str, object]]:
    lookup: dict[str, AnagrafeRecord] = {}
    reference_date = date(analysis_year, 12, 31)
    population_over75: dict[str, set[str]] = defaultdict(set)
    rows_by_file: dict[str, int] = defaultdict(int)
    file_for_asl: dict[str, str] = {}
    stats: dict[str, object] = {
        "anagrafe_files": 0,
        "anagrafe_rows": 0,
        "anagrafe_alias_conflicts": 0,
    }

    txt_files = sorted(anagrafe_dir.glob("Export_Assistiti_*.txt"))
    if not txt_files:
        raise ValueError(f"Nessun file anagrafico trovato in {anagrafe_dir}")

    for txt_path in txt_files:
        match = ANAGRAFICA_FILE_RE.search(txt_path.name)
        area_code = match.group(1) if match else ""
        fallback_asl = FALLBACK_ASL_BY_AREA.get(area_code, "")
        stats["anagrafe_files"] = int(stats["anagrafe_files"]) + 1

        with txt_path.open("r", encoding="latin-1", errors="replace") as handle:
            for line in handle:
                line = line.rstrip("\r\n")
                if not line:
                    continue
                stats["anagrafe_rows"] = int(stats["anagrafe_rows"]) + 1
                record, aliases = parse_anagrafe_line(line, txt_path.name, fallback_asl)
                if record is None:
                    continue

                rows_by_file[txt_path.name] += 1
                if record.resident_asl:
                    file_for_asl.setdefault(record.resident_asl, txt_path.name)
                for alias in aliases:
                    existing = lookup.get(alias)
                    if existing is None:
                        lookup[alias] = record
                    elif existing.primary_code != record.primary_code:
                        stats["anagrafe_alias_conflicts"] = int(stats["anagrafe_alias_conflicts"]) + 1

                if record.birth_date is not None and record.resident_asl:
                    age = age_on_reference_date(record.birth_date, reference_date)
                    if age >= 75:
                        population_over75[record.resident_asl].add(record.primary_code)

    population_rows: list[dict[str, object]] = []
    for asl in sorted(population_over75):
        population_rows.append(
            {
                "ASL Residenza": asl,
                "File TXT": file_for_asl.get(asl, ""),
                "Popolazione residente >=75": len(population_over75[asl]),
            }
        )

    stats["anagrafe_lookup_keys"] = len(lookup)
    stats["population_over75_by_asl"] = {asl: len(values) for asl, values in population_over75.items()}
    stats["population_total_over75"] = sum(len(values) for values in population_over75.values())
    return lookup, population_rows, stats


def resolve_anagrafe_record(record: FarRecord, anagrafe_lookup: dict[str, AnagrafeRecord]) -> tuple[AnagrafeRecord | None, str]:
    for candidate in record.link_candidates:
        resolved = anagrafe_lookup.get(candidate)
        if resolved is not None:
            return resolved, candidate
    return None, ""


def choose_winner(records: list[dict[str, object]]) -> dict[str, object]:
    def winner_key(item: dict[str, object]) -> tuple[int, date, int, str]:
        return (
            LEVEL_RANK[str(item["Tipo prestazione"])],
            item["Data evento oggetto"],
            int(item["Tracciato"]),
            str(item["ID_REC"]),
        )

    return max(records, key=winner_key)


def indicator_value(numerator: int, denominator: int) -> float | None:
    if denominator <= 0:
        return None
    return round((numerator / denominator) * 1000, 2)


def build_summary_total_row(summary_rows: list[dict[str, object]]) -> dict[str, object] | None:
    if not summary_rows:
        return None

    total_denominator = sum(int(row["Popolazione residente >=75"]) for row in summary_rows)
    total_all = sum(int(row["Assistiti unici >=75"]) for row in summary_rows)
    total_r1 = sum(int(row["R1"]) for row in summary_rows)
    total_r2 = sum(int(row["R2"]) for row in summary_rows)
    total_r3 = sum(int(row["R3"]) for row in summary_rows)
    total_t1 = sum(int(row["Record qualificanti T1"]) for row in summary_rows)
    total_t2 = sum(int(row["Record qualificanti T2"]) for row in summary_rows)

    return {
        "ASL Residenza": "TOTALE",
        "Popolazione residente >=75": total_denominator,
        "Record qualificanti T1": total_t1,
        "Record qualificanti T2": total_t2,
        "Assistiti unici >=75": total_all,
        "R1": total_r1,
        "R2": total_r2,
        "R3": total_r3,
        "Indicatore Totale x1000": indicator_value(total_all, total_denominator),
        "Indicatore R1 x1000": indicator_value(total_r1, total_denominator),
        "Indicatore R2 x1000": indicator_value(total_r2, total_denominator),
        "Indicatore R3 x1000": indicator_value(total_r3, total_denominator),
    }


def build_d33za_report(
    xml_files: list[FarXmlFileInfo],
    analysis_year: int,
    anagrafe_dir: Path,
) -> tuple[
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
    dict[str, object],
]:
    anagrafe_lookup, population_rows, anagrafe_stats = load_anagrafe_data(anagrafe_dir, analysis_year)
    reference_date = date(analysis_year, 12, 31)

    far_records: list[FarRecord] = []
    stats: dict[str, object] = {
        "track1_records": 0,
        "track2_records": 0,
        "matched_track1": 0,
        "matched_track2": 0,
        "qualifying_track1": 0,
        "qualifying_track2": 0,
        "selected_patients": 0,
    }
    qualifying_by_track_and_level: Counter[tuple[int, str]] = Counter()

    for file_info in xml_files:
        root = ET.parse(file_info.path).getroot()
        expected_name = TRACK1_RECORD if file_info.track == 1 else TRACK2_RECORD
        for record in iter_track_records(root, expected_name):
            parsed = (
                parse_track1_record(record, file_info.relative_path)
                if file_info.track == 1
                else parse_track2_record(record, file_info.relative_path)
            )
            if parsed is None:
                continue
            far_records.append(parsed)
            if parsed.track == 1:
                stats["track1_records"] = int(stats["track1_records"]) + 1
            else:
                stats["track2_records"] = int(stats["track2_records"]) + 1

    detail_rows: list[dict[str, object]] = []
    patient_candidates: dict[str, list[dict[str, object]]] = defaultdict(list)
    raw_qualifying_by_asl_t1: Counter[str] = Counter()
    raw_qualifying_by_asl_t2: Counter[str] = Counter()

    for record in far_records:
        anagrafe_record, matched_key = resolve_anagrafe_record(record, anagrafe_lookup)
        if anagrafe_record is not None:
            if record.track == 1:
                stats["matched_track1"] = int(stats["matched_track1"]) + 1
            else:
                stats["matched_track2"] = int(stats["matched_track2"]) + 1

        resolved_asl = (
            anagrafe_record.resident_asl
            if anagrafe_record is not None and anagrafe_record.resident_asl
            else record.resident_asl_xml
        )
        resolved_region = (
            anagrafe_record.resident_region
            if anagrafe_record is not None and anagrafe_record.resident_region
            else record.resident_region_xml
        )

        age_value: int | None = None
        age_source = ""
        birth_date_text = ""
        if anagrafe_record is not None and anagrafe_record.birth_date is not None:
            age_value = age_on_reference_date(anagrafe_record.birth_date, reference_date)
            age_source = "anagrafe"
            birth_date_text = anagrafe_record.birth_date.isoformat()
        elif record.anno_nascita_xml is not None:
            age_value = analysis_year - record.anno_nascita_xml
            age_source = "xml"

        person_key = (
            anagrafe_record.primary_code
            if anagrafe_record is not None
            else normalize_code(record.cuni) or extract_last_16(record.id_rec)
        )

        qualifying = True
        note = "Record qualificante per D33Za."

        if record.tipo_prestazione not in ALLOWED_LEVELS:
            qualifying = False
            note = "Escluso: tipologia non prevista da D33Za."
        elif not resolved_asl:
            qualifying = False
            note = "Escluso: ASL di residenza non disponibile ne da FAR ne da anagrafe."
        elif age_value is None:
            qualifying = False
            note = "Escluso: eta non calcolabile."
        elif age_value < 75:
            qualifying = False
            note = "Escluso: eta inferiore a 75 anni al 31/12."

        detail_row = {
            "ASL Residenza": resolved_asl,
            "Regione Residenza": resolved_region,
            "ASL Erogatrice": record.erogatore_asl,
            "Tracciato": record.track,
            "Trimestre": record.quarter,
            "File XML": record.source_file,
            "Data evento": record.event_date.isoformat(),
            "Data evento oggetto": record.event_date,
            "ID_REC": record.id_rec,
            "Codice collegato": matched_key or person_key,
            "Chiave paziente": person_key,
            "CUNI": record.cuni,
            "Tipo prestazione": record.tipo_prestazione,
            "Data nascita anagrafe": birth_date_text,
            "Anno nascita XML": record.anno_nascita_xml,
            "Eta al 31/12": age_value,
            "Fonte eta": age_source or "",
            "Match anagrafe": "SI" if anagrafe_record is not None else "NO",
            "Fonte residenza": "anagrafe" if anagrafe_record is not None and anagrafe_record.resident_asl else "xml",
            "Record qualificante": "SI" if qualifying else "NO",
            "Incluso nel D33Za": "NO",
            "Note": note,
        }
        detail_rows.append(detail_row)

        if not qualifying:
            continue

        qualifying_by_track_and_level[(record.track, record.tipo_prestazione)] += 1
        if record.track == 1:
            stats["qualifying_track1"] = int(stats["qualifying_track1"]) + 1
            raw_qualifying_by_asl_t1[resolved_asl] += 1
        else:
            stats["qualifying_track2"] = int(stats["qualifying_track2"]) + 1
            raw_qualifying_by_asl_t2[resolved_asl] += 1

        patient_candidates[person_key].append(detail_row)

    selected_rows: list[dict[str, object]] = []
    selected_by_asl: dict[str, list[dict[str, object]]] = defaultdict(list)

    for patient_key, rows in sorted(patient_candidates.items()):
        winner = choose_winner(rows)
        winner["Incluso nel D33Za"] = "SI"
        if len(rows) == 1:
            winner["Note"] = "Incluso nel D33Za."
        else:
            winner["Note"] = (
                "Incluso: paziente presente in piu record/tracciati nell'anno; "
                f"selezionato il livello massimo {winner['Tipo prestazione']}."
            )

        for row in rows:
            if row is winner:
                continue
            row["Note"] = (
                "Escluso: stesso paziente gia contato una sola volta "
                f"sul livello {winner['Tipo prestazione']}."
            )

        selected_rows.append(
            {
                "ASL Residenza": winner["ASL Residenza"],
                "Regione Residenza": winner["Regione Residenza"],
                "Chiave paziente": patient_key,
                "Tipo selezionato": winner["Tipo prestazione"],
                "Tracciato selezionato": winner["Tracciato"],
                "Data record usato": winner["Data evento"],
                "ID_REC usato": winner["ID_REC"],
                "ASL Erogatrice": winner["ASL Erogatrice"],
                "Match anagrafe": winner["Match anagrafe"],
                "File XML": winner["File XML"],
                "Numero record qualificanti paziente": len(rows),
            }
        )
        selected_by_asl[str(winner["ASL Residenza"])].append(winner)

    population_over75_by_asl = {
        str(row["ASL Residenza"]): int(row["Popolazione residente >=75"])
        for row in population_rows
    }

    summary_rows: list[dict[str, object]] = []
    all_asls = sorted(set(population_over75_by_asl) | set(selected_by_asl))
    for asl in all_asls:
        selected = selected_by_asl.get(asl, [])
        level_counts = Counter(str(row["Tipo prestazione"]) for row in selected)
        denominator = population_over75_by_asl.get(asl, 0)
        numerator_total = len(selected)
        numerator_r1 = level_counts.get("R1", 0)
        numerator_r2 = level_counts.get("R2", 0)
        numerator_r3 = level_counts.get("R3", 0)
        summary_rows.append(
            {
                "ASL Residenza": asl,
                "Popolazione residente >=75": denominator,
                "Record qualificanti T1": raw_qualifying_by_asl_t1.get(asl, 0),
                "Record qualificanti T2": raw_qualifying_by_asl_t2.get(asl, 0),
                "Assistiti unici >=75": numerator_total,
                "R1": numerator_r1,
                "R2": numerator_r2,
                "R3": numerator_r3,
                "Indicatore Totale x1000": indicator_value(numerator_total, denominator),
                "Indicatore R1 x1000": indicator_value(numerator_r1, denominator),
                "Indicatore R2 x1000": indicator_value(numerator_r2, denominator),
                "Indicatore R3 x1000": indicator_value(numerator_r3, denominator),
            }
        )

    track_rows: list[dict[str, object]] = []
    for track in (1, 2):
        track_rows.append(
            {
                "Tracciato": track,
                "Record letti": int(stats[f"track{track}_records"]),
                "Record con match anagrafe": int(stats[f"matched_track{track}"]),
                "Record qualificanti": int(stats[f"qualifying_track{track}"]),
                "R1": qualifying_by_track_and_level.get((track, "R1"), 0),
                "R2": qualifying_by_track_and_level.get((track, "R2"), 0),
                "R3": qualifying_by_track_and_level.get((track, "R3"), 0),
            }
        )

    detail_rows.sort(
        key=lambda row: (
            str(row["ASL Residenza"]),
            str(row["Chiave paziente"]),
            str(row["Data evento"]),
            str(row["ID_REC"]),
        )
    )
    selected_rows.sort(
        key=lambda row: (
            str(row["ASL Residenza"]),
            str(row["Tipo selezionato"]),
            str(row["Chiave paziente"]),
        )
    )

    stats["selected_patients"] = len(selected_rows)
    stats["population_total_over75"] = int(anagrafe_stats["population_total_over75"])
    stats["anagrafe_files"] = int(anagrafe_stats["anagrafe_files"])
    stats["anagrafe_rows"] = int(anagrafe_stats["anagrafe_rows"])
    stats["anagrafe_lookup_keys"] = int(anagrafe_stats["anagrafe_lookup_keys"])
    stats["anagrafe_alias_conflicts"] = int(anagrafe_stats["anagrafe_alias_conflicts"])
    return summary_rows, detail_rows, selected_rows, population_rows, track_rows, stats


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL_ODD = PatternFill(fill_type="solid", fgColor="F7FBFF")
ALT_FILL_EVEN = PatternFill(fill_type="solid", fgColor="EAF2F8")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
TOTAL_FONT = Font(bold=True, color="1F1F1F")
CENTER_ALIGNMENT = Alignment(vertical="center")


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 45)


def style_worksheet(ws, has_total_row: bool = False) -> None:
    max_row = ws.max_row
    max_col = ws.max_column

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT

    data_end_row = max_row - 1 if has_total_row and max_row >= 2 else max_row
    for row_idx in range(2, data_end_row + 1):
        fill = ALT_FILL_ODD if row_idx % 2 == 0 else ALT_FILL_EVEN
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = CENTER_ALIGNMENT

    if has_total_row and max_row >= 2:
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=max_row, column=col_idx)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.alignment = CENTER_ALIGNMENT

    autosize_columns(ws)


def write_sheet(ws, rows: list[dict[str, object]], fallback_headers: list[str], total_row: bool = False) -> None:
    headers = list(rows[0].keys()) if rows else fallback_headers
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    style_worksheet(ws, has_total_row=total_row and bool(rows))


def save_workbook(
    output_path: Path,
    summary_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
    selected_rows: list[dict[str, object]],
    population_rows: list[dict[str, object]],
    track_rows: list[dict[str, object]],
) -> None:
    workbook = Workbook()

    ws_summary = workbook.active
    ws_summary.title = "D33Za_Riepilogo"
    summary_with_total = list(summary_rows)
    total_row = build_summary_total_row(summary_rows)
    if total_row is not None:
        summary_with_total.append(total_row)
    write_sheet(
        ws_summary,
        summary_with_total,
        [
            "ASL Residenza",
            "Popolazione residente >=75",
            "Record qualificanti T1",
            "Record qualificanti T2",
            "Assistiti unici >=75",
            "R1",
            "R2",
            "R3",
            "Indicatore Totale x1000",
            "Indicatore R1 x1000",
            "Indicatore R2 x1000",
            "Indicatore R3 x1000",
        ],
        total_row=True,
    )

    ws_detail = workbook.create_sheet("Dettaglio")
    write_sheet(
        ws_detail,
        detail_rows,
        [
            "ASL Residenza",
            "Regione Residenza",
            "ASL Erogatrice",
            "Tracciato",
            "Trimestre",
            "File XML",
            "Data evento",
            "ID_REC",
            "Codice collegato",
            "Chiave paziente",
            "CUNI",
            "Tipo prestazione",
            "Data nascita anagrafe",
            "Anno nascita XML",
            "Eta al 31/12",
            "Fonte eta",
            "Match anagrafe",
            "Fonte residenza",
            "Record qualificante",
            "Incluso nel D33Za",
            "Note",
        ],
    )

    ws_selected = workbook.create_sheet("Assistiti_Selezionati")
    write_sheet(
        ws_selected,
        selected_rows,
        [
            "ASL Residenza",
            "Regione Residenza",
            "Chiave paziente",
            "Tipo selezionato",
            "Tracciato selezionato",
            "Data record usato",
            "ID_REC usato",
            "ASL Erogatrice",
            "Match anagrafe",
            "File XML",
            "Numero record qualificanti paziente",
        ],
    )

    ws_population = workbook.create_sheet("Popolazione_Anagrafe")
    write_sheet(
        ws_population,
        population_rows,
        ["ASL Residenza", "File TXT", "Popolazione residente >=75"],
    )

    ws_tracks = workbook.create_sheet("Verifica_Tracciati")
    write_sheet(
        ws_tracks,
        track_rows,
        ["Tracciato", "Record letti", "Record con match anagrafe", "Record qualificanti", "R1", "R2", "R3"],
    )

    workbook.save(output_path)


class FarD33ZaApp:
    SPINNER_FRAMES = ["[...]", "[ ..]", "[  .]", "[ ..]"]

    def __init__(
        self,
        root: Any,
        tk_module: Any,
        ttk_module: Any,
        filedialog_module: Any,
        messagebox_module: Any,
        scrolledtext_module: Any,
        *,
        parent: Any | None = None,
        embed_mode: bool = False,
    ) -> None:
        self.root = root
        self.parent = parent or root
        self.embed_mode = embed_mode
        self.tk = tk_module
        self.ttk = ttk_module
        self.filedialog = filedialog_module
        self.messagebox = messagebox_module
        self.scrolledtext = scrolledtext_module

        if not self.embed_mode:
            self.root.title("FAR D33Za")
            self.root.geometry("1220x820")
            self.root.minsize(1080, 720)

        self.xml_dir_var = self.tk.StringVar(value="")
        self.anagrafe_dir_var = self.tk.StringVar(value="")
        self.output_file_var = self.tk.StringVar(value="")
        self.analysis_year_var = self.tk.StringVar(value="2025")
        self.status_var = self.tk.StringVar(
            value="Seleziona la cartella FAR, la cartella anagrafe TXT e genera il report dell'indicatore D33Za."
        )
        self.spinner_var = self.tk.StringVar(value="")

        self._busy = False
        self._spinner_index = 0
        self._spinner_job: str | None = None
        self._last_output: Path | None = None

        self.xml_files: list[FarXmlFileInfo] = []
        self.summary_rows: list[dict[str, object]] = []

        self._build_ui()
        self._load_state()

    def _build_ui(self) -> None:
        if not self.embed_mode:
            self._build_menu()

        style = self.ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        frame = self.ttk.Frame(self.parent, padding=12)
        frame.pack(fill="both", expand=True)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(3, weight=1)

        ribbon = self.ttk.Frame(frame)
        ribbon.grid(row=0, column=0, sticky="ew")
        ribbon.columnconfigure(0, weight=1)
        ribbon.columnconfigure(1, weight=1)
        ribbon.columnconfigure(2, weight=1)

        self._build_ribbon_group(
            ribbon,
            0,
            "Percorsi",
            [
                ("Cartella FAR", self.choose_xml_dir),
                ("Cartella anagrafe", self.choose_anagrafe_dir),
                ("File output", self.choose_output_file),
            ],
        )
        self._build_ribbon_group(
            ribbon,
            1,
            "Esecuzione",
            [
                ("Scansiona XML", self.scan_files),
                ("Genera report", self.generate_report),
                ("Pulisci log", self.clear_views),
            ],
        )
        self._build_ribbon_group(
            ribbon,
            2,
            "Report",
            [
                ("Apri Excel", self.open_output_file),
                ("Apri cartella", self.open_output_dir),
            ],
        )

        config_frame = self.ttk.LabelFrame(frame, text="Configurazione", padding=12)
        config_frame.grid(row=1, column=0, sticky="ew", pady=(12, 8))
        config_frame.columnconfigure(1, weight=1)

        self._add_path_row(config_frame, 0, "Cartella XML FAR", self.xml_dir_var, self.choose_xml_dir)
        self._add_path_row(config_frame, 1, "Cartella anagrafe TXT", self.anagrafe_dir_var, self.choose_anagrafe_dir)
        self._add_path_row(config_frame, 2, "File Excel output", self.output_file_var, self.choose_output_file)

        params = self.ttk.Frame(config_frame)
        params.grid(row=3, column=0, columnspan=3, sticky="w", pady=(8, 0))
        self.ttk.Label(params, text="Anno di analisi:").pack(side="left")
        self.ttk.Entry(params, textvariable=self.analysis_year_var, width=8).pack(side="left", padx=(6, 0))

        hint = (
            "D33Za: assistiti residenti con eta >=75 anni in trattamento residenziale R1/R2/R3, "
            "conteggiati una sola volta nell'anno sul livello di intensita piu elevato. "
            "Il report usa i TXT anagrafici per denominatore popolazione >=75 e per derivare "
            "eta/residenza quando necessario, incluso il Tracciato 2."
        )
        self.ttk.Label(frame, text=hint, wraplength=1160, justify="left").grid(row=2, column=0, sticky="ew", pady=(0, 8))

        self.notebook = self.ttk.Notebook(frame)
        self.notebook.grid(row=3, column=0, sticky="nsew")

        files_tab = self.ttk.Frame(self.notebook, padding=8)
        summary_tab = self.ttk.Frame(self.notebook, padding=8)
        output_tab = self.ttk.Frame(self.notebook, padding=8)
        self.notebook.add(files_tab, text="File XML")
        self.notebook.add(summary_tab, text="Riepilogo")
        self.notebook.add(output_tab, text="Output")

        files_tab.columnconfigure(0, weight=1)
        files_tab.rowconfigure(0, weight=1)
        summary_tab.columnconfigure(0, weight=1)
        summary_tab.rowconfigure(0, weight=1)
        output_tab.columnconfigure(0, weight=1)
        output_tab.rowconfigure(0, weight=1)

        file_columns = ("row_number", "track", "line_count", "size_mb", "relative_path")
        self.file_tree = self.ttk.Treeview(files_tab, columns=file_columns, show="headings", height=20)
        self.file_tree.heading("row_number", text="#")
        self.file_tree.heading("track", text="Tracciato")
        self.file_tree.heading("line_count", text="Righe")
        self.file_tree.heading("size_mb", text="MB")
        self.file_tree.heading("relative_path", text="File XML")
        self.file_tree.column("row_number", width=60, anchor="center", stretch=False)
        self.file_tree.column("track", width=100, anchor="center", stretch=False)
        self.file_tree.column("line_count", width=90, anchor="e", stretch=False)
        self.file_tree.column("size_mb", width=90, anchor="e", stretch=False)
        self.file_tree.column("relative_path", width=760, anchor="w")
        self.file_tree.grid(row=0, column=0, sticky="nsew")
        file_scrollbar = self.ttk.Scrollbar(files_tab, orient="vertical", command=self.file_tree.yview)
        file_scrollbar.grid(row=0, column=1, sticky="ns")
        self.file_tree.configure(yscrollcommand=file_scrollbar.set)

        self.summary_tree = self.ttk.Treeview(summary_tab, show="headings", height=20)
        self.summary_tree.grid(row=0, column=0, sticky="nsew")
        summary_scrollbar = self.ttk.Scrollbar(summary_tab, orient="vertical", command=self.summary_tree.yview)
        summary_scrollbar.grid(row=0, column=1, sticky="ns")
        summary_x_scrollbar = self.ttk.Scrollbar(summary_tab, orient="horizontal", command=self.summary_tree.xview)
        summary_x_scrollbar.grid(row=1, column=0, sticky="ew")
        self.summary_tree.configure(yscrollcommand=summary_scrollbar.set, xscrollcommand=summary_x_scrollbar.set)

        self.console_text = self.scrolledtext.ScrolledText(output_tab, wrap="word", font=("Menlo", 11))
        self.console_text.grid(row=0, column=0, sticky="nsew")
        self.console_text.configure(state="disabled")
        self.console_text.tag_configure("info", foreground="#1f1f1f")
        self.console_text.tag_configure("warn", foreground="#9c6500")
        self.console_text.tag_configure("error", foreground="#9c0006")
        self.console_text.tag_configure("success", foreground="#0b6b2f")
        self.console_text.tag_configure("header", foreground="#1f4e78", font=("Menlo", 11, "bold"))

        status_bar = self.ttk.Frame(frame)
        status_bar.grid(row=4, column=0, sticky="ew", pady=(8, 0))
        status_bar.columnconfigure(2, weight=1)
        self.ttk.Label(status_bar, textvariable=self.spinner_var, width=5, anchor="w").grid(row=0, column=0, sticky="w")
        self.progress = self.ttk.Progressbar(status_bar, mode="indeterminate", length=180)
        self.progress.grid(row=0, column=1, sticky="w", padx=(0, 12))
        self.ttk.Label(status_bar, textvariable=self.status_var, anchor="w").grid(row=0, column=2, sticky="ew")

    def _build_menu(self) -> None:
        menu = self.tk.Menu(self.root)

        file_menu = self.tk.Menu(menu, tearoff=False)
        file_menu.add_command(label="Seleziona cartella FAR", command=self.choose_xml_dir)
        file_menu.add_command(label="Seleziona cartella anagrafe", command=self.choose_anagrafe_dir)
        file_menu.add_command(label="Seleziona file output", command=self.choose_output_file)
        file_menu.add_separator()
        file_menu.add_command(label="Esci", command=self.root.destroy)
        menu.add_cascade(label="File", menu=file_menu)

        run_menu = self.tk.Menu(menu, tearoff=False)
        run_menu.add_command(label="Scansiona XML", command=self.scan_files)
        run_menu.add_command(label="Genera report", command=self.generate_report)
        run_menu.add_command(label="Pulisci log", command=self.clear_views)
        menu.add_cascade(label="Esecuzione", menu=run_menu)

        report_menu = self.tk.Menu(menu, tearoff=False)
        report_menu.add_command(label="Apri Excel", command=self.open_output_file)
        report_menu.add_command(label="Apri cartella output", command=self.open_output_dir)
        menu.add_cascade(label="Report", menu=report_menu)

        self.root.config(menu=menu)

    def _build_ribbon_group(self, parent: Any, column: int, title: str, buttons: list[tuple[str, Any]]) -> None:
        group = self.ttk.LabelFrame(parent, text=title, padding=8)
        group.grid(row=0, column=column, sticky="nsew", padx=(0, 8) if column < 2 else 0)
        for index, (label, command) in enumerate(buttons):
            self.ttk.Button(group, text=label, command=command).grid(row=0, column=index, padx=(0, 8), pady=2)

    def _add_path_row(self, parent: Any, row: int, label: str, variable: Any, command: Any) -> None:
        self.ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=4)
        self.ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=8, pady=4)
        self.ttk.Button(parent, text="Sfoglia...", command=command).grid(row=row, column=2, sticky="e", pady=4)

    def choose_xml_dir(self) -> None:
        path = self.filedialog.askdirectory(initialdir=self.xml_dir_var.get() or str(Path.home()))
        if not path:
            return
        xml_dir = Path(path)
        self.xml_dir_var.set(str(xml_dir))
        suggested_anagrafe_dir = xml_dir / "anagrafe"
        if suggested_anagrafe_dir.is_dir() and not self.anagrafe_dir_var.get().strip():
            self.anagrafe_dir_var.set(str(suggested_anagrafe_dir))
        if not self.output_file_var.get().strip():
            self.output_file_var.set(str(next_available_report_path(xml_dir)))
        self._save_state()

    def choose_anagrafe_dir(self) -> None:
        path = self.filedialog.askdirectory(initialdir=self.anagrafe_dir_var.get() or self.xml_dir_var.get() or str(Path.home()))
        if path:
            self.anagrafe_dir_var.set(path)
            self._save_state()

    def choose_output_file(self) -> None:
        current = self.output_file_var.get().strip()
        init_dir = str(Path(current).parent) if current else str(Path.home())
        path = self.filedialog.asksaveasfilename(
            initialdir=init_dir,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.output_file_var.set(path)
            self._save_state()

    def clear_views(self) -> None:
        self._set_text(self.console_text, "")
        self.status_var.set("Log ripuliti.")

    def scan_files(self) -> None:
        if self._busy:
            return
        xml_dir = Path(self.xml_dir_var.get().strip()).expanduser()
        if not xml_dir.is_dir():
            self.messagebox.showerror("Errore", "Seleziona una cartella FAR valida.")
            return
        self._save_state()
        self._set_busy(True, "Scansione XML FAR in corso...")
        threading.Thread(target=self._scan_worker, args=(xml_dir,), daemon=True).start()

    def _scan_worker(self, xml_dir: Path) -> None:
        try:
            xml_files = scan_xml_files(xml_dir)
        except Exception as exc:
            self.root.after(0, lambda: self._scan_failed(exc))
            return
        self.root.after(0, lambda: self._scan_completed(xml_dir, xml_files))

    def _scan_failed(self, exc: Exception) -> None:
        self._set_busy(False, "Scansione fallita.")
        self.append_console("Errore durante la scansione XML.", "error")
        self.append_console("".join(traceback.format_exception(exc)), "error")
        self.messagebox.showerror("Errore", f"Scansione XML fallita:\n{exc}")

    def _scan_completed(self, xml_dir: Path, xml_files: list[FarXmlFileInfo]) -> None:
        self.xml_files = xml_files
        self.populate_file_tree()
        if not self.output_file_var.get().strip():
            self.output_file_var.set(str(next_available_report_path(xml_dir)))

        track_counts = Counter(info.track for info in xml_files)
        self.clear_views()
        self.append_console(f"Cartella FAR: {xml_dir}", "header")
        self.append_console(f"File XML trovati: {len(xml_files)}", "info")
        self.append_console(f"Tracciato 1: {track_counts.get(1, 0)} file", "info")
        self.append_console(f"Tracciato 2: {track_counts.get(2, 0)} file", "info")
        self._set_busy(False, f"Scansione completata: {len(xml_files)} file XML.")

    def generate_report(self) -> None:
        if self._busy:
            return

        xml_dir = Path(self.xml_dir_var.get().strip()).expanduser()
        anagrafe_dir = Path(self.anagrafe_dir_var.get().strip()).expanduser()
        output_path = Path(self.output_file_var.get().strip()).expanduser()
        try:
            analysis_year = int(self.analysis_year_var.get().strip())
            if analysis_year < 2000 or analysis_year > 2100:
                raise ValueError
        except ValueError:
            self.messagebox.showerror("Errore", "L'anno di analisi deve essere un intero valido.")
            return

        if not xml_dir.is_dir():
            self.messagebox.showerror("Errore", "Seleziona una cartella FAR valida.")
            return
        if not anagrafe_dir.is_dir():
            self.messagebox.showerror("Errore", "Seleziona una cartella anagrafe valida.")
            return
        if not list(anagrafe_dir.glob("Export_Assistiti_*.txt")):
            self.messagebox.showerror("Errore", "Nella cartella anagrafe non ci sono file 'Export_Assistiti_*.txt'.")
            return
        if not output_path.name:
            self.messagebox.showerror("Errore", "Indica il file Excel di output.")
            return
        if not self.xml_files:
            try:
                self.xml_files = scan_xml_files(xml_dir)
            except Exception as exc:
                self.messagebox.showerror("Errore", f"Impossibile scansionare i file FAR:\n{exc}")
                return
            self.populate_file_tree()

        self._save_state()
        self.clear_views()
        self.append_console(f"Cartella FAR: {xml_dir}", "header")
        self.append_console(f"Cartella anagrafe: {anagrafe_dir}", "header")
        self.append_console(f"Anno di analisi: {analysis_year}", "header")
        self.append_console(f"Output: {output_path}", "header")
        self.append_console("", "info")

        self._set_busy(True, "Generazione report D33Za in corso...")
        threading.Thread(
            target=self._run_worker,
            args=(analysis_year, anagrafe_dir, output_path),
            daemon=True,
        ).start()

    def _run_worker(self, analysis_year: int, anagrafe_dir: Path, output_path: Path) -> None:
        try:
            summary_rows, detail_rows, selected_rows, population_rows, track_rows, stats = build_d33za_report(
                self.xml_files,
                analysis_year,
                anagrafe_dir,
            )
            output_path.parent.mkdir(parents=True, exist_ok=True)
            save_workbook(output_path, summary_rows, detail_rows, selected_rows, population_rows, track_rows)
        except Exception as exc:
            self.root.after(0, lambda: self._run_failed(exc))
            return
        self.root.after(
            0,
            lambda: self._run_completed(output_path, summary_rows, stats),
        )

    def _run_failed(self, exc: Exception) -> None:
        self._set_busy(False, "Generazione report fallita.")
        self.append_console("Errore durante la generazione del report D33Za.", "error")
        self.append_console("".join(traceback.format_exception(exc)), "error")
        self.messagebox.showerror("Errore", f"Generazione report fallita:\n{exc}")

    def _run_completed(self, output_path: Path, summary_rows: list[dict[str, object]], stats: dict[str, object]) -> None:
        self.summary_rows = summary_rows
        self.populate_summary_tree()
        self._last_output = output_path
        self.append_console(
            f"TXT anagrafe: {stats['anagrafe_files']} file | righe lette: {stats['anagrafe_rows']} | chiavi lookup: {stats['anagrafe_lookup_keys']}",
            "info",
        )
        self.append_console(
            f"Popolazione residente >=75 da anagrafe: {stats['population_total_over75']}",
            "info",
        )
        self.append_console(
            f"Record FAR letti T1: {stats['track1_records']} | T2: {stats['track2_records']}",
            "info",
        )
        self.append_console(
            f"Match anagrafe T1: {stats['matched_track1']} | T2: {stats['matched_track2']}",
            "info",
        )
        self.append_console(
            f"Record qualificanti T1: {stats['qualifying_track1']} | T2: {stats['qualifying_track2']}",
            "success",
        )
        self.append_console(
            f"Assistiti selezionati finali: {stats['selected_patients']}",
            "success",
        )
        self.append_console(f"Report salvato in: {output_path}", "success")
        self.notebook.select(1)
        self._set_busy(False, f"Report D33Za completato: {output_path.name}")

    def populate_file_tree(self) -> None:
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        for index, info in enumerate(self.xml_files, start=1):
            self.file_tree.insert(
                "",
                "end",
                values=(index, info.track, info.line_count, f"{info.size_mb:.2f}", info.relative_path),
            )

    def populate_summary_tree(self) -> None:
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        self.summary_tree["columns"] = ()
        if not self.summary_rows:
            return

        headers = list(self.summary_rows[0].keys())
        self.summary_tree["columns"] = headers
        for header in headers:
            self.summary_tree.heading(header, text=header)
            anchor = "center" if header == "ASL Residenza" else "e"
            width = 150 if header == "ASL Residenza" else 130
            self.summary_tree.column(header, width=width, anchor=anchor, stretch=True)
        for row in self.summary_rows:
            self.summary_tree.insert("", "end", values=[row.get(header) for header in headers])

    def append_console(self, text: str, tag: str = "info") -> None:
        self.console_text.configure(state="normal")
        self.console_text.insert("end", text + "\n", tag)
        self.console_text.see("end")
        self.console_text.configure(state="disabled")

    def _set_text(self, widget: Any, value: str) -> None:
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        if value:
            widget.insert("1.0", value)
        widget.configure(state="disabled")

    def _set_busy(self, busy: bool, message: str) -> None:
        self._busy = busy
        self.status_var.set(message)
        if busy:
            self.progress.start(12)
            self._animate_spinner()
        else:
            self.progress.stop()
            if self._spinner_job is not None:
                self.root.after_cancel(self._spinner_job)
                self._spinner_job = None
            self.spinner_var.set("")

    def _animate_spinner(self) -> None:
        if not self._busy:
            return
        self.spinner_var.set(self.SPINNER_FRAMES[self._spinner_index % len(self.SPINNER_FRAMES)])
        self._spinner_index += 1
        self._spinner_job = self.root.after(250, self._animate_spinner)

    def _load_state(self) -> None:
        if not STATE_FILE.is_file():
            return
        try:
            data = json.loads(STATE_FILE.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return
        if not isinstance(data, dict):
            return
        self.xml_dir_var.set(str(data.get("xml_dir", "") or ""))
        self.anagrafe_dir_var.set(str(data.get("anagrafe_dir", "") or ""))
        self.output_file_var.set(str(data.get("output_file", "") or ""))
        self.analysis_year_var.set(str(data.get("analysis_year", "2025") or "2025"))

    def _save_state(self) -> None:
        data = {
            "xml_dir": self.xml_dir_var.get().strip(),
            "anagrafe_dir": self.anagrafe_dir_var.get().strip(),
            "output_file": self.output_file_var.get().strip(),
            "analysis_year": self.analysis_year_var.get().strip(),
        }
        try:
            STATE_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")
        except OSError:
            pass

    def open_output_file(self) -> None:
        output_path = self._last_output or Path(self.output_file_var.get().strip()).expanduser()
        if not output_path.exists():
            self.messagebox.showerror("Errore", "Il file Excel di output non esiste ancora.")
            return
        self._open_path(output_path)

    def open_output_dir(self) -> None:
        output_path = self._last_output or Path(self.output_file_var.get().strip()).expanduser()
        target_dir = output_path.parent if output_path.name else Path(self.xml_dir_var.get().strip()).expanduser()
        if not target_dir.exists():
            self.messagebox.showerror("Errore", "La cartella di output non esiste.")
            return
        self._open_path(target_dir)

    def _open_path(self, path: Path) -> None:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
            return
        if sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
            return
        subprocess.Popen(["xdg-open", str(path)])


def main() -> None:
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext, ttk

    root = tk.Tk()
    FarD33ZaApp(
        root,
        tk,
        ttk,
        filedialog,
        messagebox,
        scrolledtext,
    )
    root.mainloop()


if __name__ == "__main__":
    main()
