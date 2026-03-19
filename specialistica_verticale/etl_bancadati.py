#!/usr/bin/env python3

from __future__ import annotations

import csv
import os
import re
import subprocess
import unicodedata
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
BANCADATI_DIR = ROOT / "bancadati"
ENV_MAPPING_FILE = os.environ.get("SPECIALISTICA_MAPPING_FILE", "").strip()
MAPPING_FILE = Path(ENV_MAPPING_FILE).expanduser() if ENV_MAPPING_FILE else ROOT / "BRANCA_Codici regionali-Codici SSN.xlsx"
TEMPLATE_FILE = ROOT / "FORMAT_ACQUAVIVA.xlsx"
OUTPUT_DIR = ROOT / "output"

PLACEHOLDER_BRANCH = "CODICE DELLA BRANCA SPECIALISTICA COME DA DPCM 2017 E SMI"
PLACEHOLDER_CODE = (
    "AVVALORARE SOLO SE NON TUTTE LE PRESTAZIONI DELLA BRANCA SONO CONTRATTUALIZZATE"
)

HEADER_STS = "CODICE STS11 (*)"
HEADER_DENOM = "DENOMINAZIONE STRUTTURA"
HEADER_BRANCH = "BRANCA SPECIALISTICA (*)"
HEADER_DPCM = "CODICE BRANCA DPCM 2017 (*)"
HEADER_CODE = "CODICE PRESTAZIONE (*)"
HEADER_CODE_REGIONALE = "CODICE PRESTAZIONE (Regionale)"
HEADER_CODE_NAZIONALE = "CODICE PRESTAZIONE (Nazionale)"

CODE_RE = re.compile(r"(?<![\d.])\d{2}(?:\.[0-9A-Z]{1,4})+(?![0-9A-Z])")
STS_RE = re.compile(r"\b\d{4,6}\b", re.IGNORECASE)

MANUAL_BRANCH_HINTS = {
    "LABORATOR": "11",
    "ANALISI CLINIC": "11",
    "CHIMICA CLINIC": "11",
    "MICROBIOLOG": "11",
    "DIAGNOSTICA PER IMMAGINI": "08",
    "RADIOLOG": "08",
    "MEDICINA FISICA": "12",
    "RIABILIT": "12",
    "ODONTO": "17",
    "STOMATO": "17",
    "CARDIOLOG": "02",
    "ORTOPEDIA": "19",
}

BRANCH_ONLY_MARKERS = (
    "TUTTE LE PRESTAZIONI",
    "CONTRATTUALIZZATE",
    "COD. REGIONALE",
    "CODICE REGIONALE",
    "SETTORI",
    "BRANCA",
)


@dataclass(frozen=True)
class MappingEntry:
    branches: tuple[str, ...]
    regional: str
    ssn: str


@dataclass
class Issue:
    source_file: str
    row: str
    issue: str
    details: str


@dataclass
class RawRow:
    row_number: int
    sts_text: str
    denom_text: str
    branch_text: str
    dpcm_text: str
    regional_text: str
    national_text: str


@dataclass
class ParsedFile:
    path: Path
    relative_path: str
    sts: str = ""
    denom: str = ""
    raw_rows: list[RawRow] = field(default_factory=list)
    issues: list[Issue] = field(default_factory=list)


@dataclass
class ProcessedFile:
    path: Path
    relative_path: str
    sts: str
    denom: str
    output_rows: list[tuple[str, str, str, str]]
    issues: list[Issue]
    explicit_matches: list[tuple[int, str, str, str, str, str, str]]


@dataclass
class SourceParseResult:
    parsed: ParsedFile | None
    supported: bool


class MappingStore:
    def __init__(self, entries: list[MappingEntry]) -> None:
        self.entries = entries
        self.by_code: dict[str, list[MappingEntry]] = defaultdict(list)
        self.by_ssn: dict[str, list[MappingEntry]] = defaultdict(list)
        self.by_branch: dict[str, list[MappingEntry]] = defaultdict(list)
        self.code_to_branches: dict[str, set[str]] = defaultdict(set)
        self.ssn_to_branches: dict[str, set[str]] = defaultdict(set)
        for entry in entries:
            self.by_code[entry.regional].append(entry)
            self.by_ssn[entry.ssn].append(entry)
            for branch in entry.branches:
                self.by_branch[branch].append(entry)
                self.code_to_branches[entry.regional].add(branch)
                self.ssn_to_branches[entry.ssn].add(branch)


def collapse_spaces(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def ascii_fold(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return normalized.encode("ascii", "ignore").decode("ascii")


def branch_text_key(value: str) -> str:
    text = ascii_fold(collapse_spaces(value).upper())
    if not text:
        return ""
    if PLACEHOLDER_BRANCH in text:
        return ""
    text = re.sub(r"[^A-Z ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def ordered_unique(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            ordered.append(value)
    return ordered


def normalize_branch_code(value: str) -> str:
    return str(int(value)).zfill(2)


def extract_branch_codes_from_text(value: str) -> list[str]:
    text = collapse_spaces(value).upper()
    if not text or PLACEHOLDER_BRANCH in text:
        return []

    matches: list[str] = []

    if re.fullmatch(r"\d{1,2}", text):
        matches.append(normalize_branch_code(text))

    matches.extend(
        normalize_branch_code(match)
        for match in re.findall(r"(?<!\d)(\d{1,2})(?=\s*[-.])", text)
    )
    matches.extend(
        normalize_branch_code(match)
        for match in re.findall(r"\bN[.°O]*\s*(\d{1,2})\b", text)
    )
    matches.extend(
        normalize_branch_code(match)
        for match in re.findall(r"\bBRANCA\s*(\d{1,2})\b", text)
    )

    filtered = [code for code in matches if 1 <= int(code) <= 26]
    return ordered_unique(filtered)


def infer_branch_codes(
    branch_text: str,
    dpcm_text: str,
    learned_descriptions: dict[str, str],
) -> list[str]:
    direct = ordered_unique(
        extract_branch_codes_from_text(branch_text)
        + extract_branch_codes_from_text(dpcm_text)
    )
    if direct:
        return direct

    for raw_text in (branch_text, dpcm_text):
        key = branch_text_key(raw_text)
        if key and key in learned_descriptions:
            return [learned_descriptions[key]]

    combined_key = " ".join(
        key for key in (branch_text_key(branch_text), branch_text_key(dpcm_text)) if key
    )
    for haystack in filter(None, [combined_key, branch_text_key(branch_text), branch_text_key(dpcm_text)]):
        for needle, code in MANUAL_BRANCH_HINTS.items():
            if needle in haystack:
                return [code]

    return []


def normalize_sts(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, int):
        return str(value).zfill(6)

    text = collapse_spaces(value).upper()
    if not text:
        return ""

    compact = re.sub(r"\s+", "", text)
    match = re.search(r"SSA\d{3}", compact)
    if match:
        return match.group(0)

    match = re.search(r"\bAMBULATORIO\s+([A-Z]\d{2,3})\b", text)
    if match:
        return match.group(1)

    match = re.search(r"\b\d{6}\b", text)
    if match:
        return match.group(0)

    match = STS_RE.search(text)
    if match:
        return match.group(0).upper().zfill(6)

    generic_tokens = re.findall(r"\b[A-Z]{1,3}\s*\d{2,3}\b", text)
    for token in generic_tokens:
        normalized = re.sub(r"\s+", "", token)
        if normalized.startswith(("STS", "RC", "ASP")):
            continue
        return normalized

    if "CODICE STS11" in text:
        return ""

    return ""


def normalize_denom(value: object) -> str:
    text = collapse_spaces(value)
    if not text:
        return ""
    upper = text.upper()
    if "DENOMINAZIONE DELLA STRUTTURA SANITARIA" in upper:
        return ""
    return text


def prepare_code_text(value: str) -> str:
    text = collapse_spaces(value).upper()
    text = text.replace("‐", "-").replace("–", "-").replace("—", "-")
    text = re.sub(r"\s*\.\s*", ".", text)
    text = re.sub(r"([A-Z])\s+(\d{1,2})(?!\d)", r"\1\2", text)
    text = re.sub(r"(\d)\s+([A-Z])(?=\d)", r"\1\2", text)
    text = re.sub(r"([A-Z]\d)\s+([A-Z])", r"\1\2", text)
    return text


def normalize_code(value: str) -> str:
    return prepare_code_text(value).strip(" ,;:-")


def extract_code_tokens(value: str) -> list[str]:
    prepared = prepare_code_text(value)
    return ordered_unique(normalize_code(match.group(0)) for match in CODE_RE.finditer(prepared))


def exact_mapping_ssns(code: str, branches: list[str], mapping: MappingStore) -> list[str]:
    exact = unique_entries_by_ssn(filter_entries_by_branches(mapping.by_code.get(code, []), branches))
    if not exact:
        exact = unique_entries_by_ssn(mapping.by_code.get(code, []))
    return ordered_unique(entry.ssn for entry in exact)


def explicit_pair_catalog_status(regional: str, ssn: str, branches: list[str], mapping: MappingStore) -> str:
    expected_ssns = exact_mapping_ssns(regional, branches, mapping)
    if not expected_ssns:
        return "catalog_missing"
    if ssn in expected_ssns:
        return "catalog_match"
    return f"catalog_differs:{','.join(expected_ssns)}"


def should_use_mapping_for_explicit_pairs(parsed: ParsedFile) -> bool:
    return parsed.path.parent.name.upper() == "RONTGEN"


def resolve_explicit_output_ssn(
    parsed: ParsedFile,
    regional: str,
    explicit_ssn: str,
    branches: list[str],
    mapping: MappingStore,
    catalog_status: str,
) -> tuple[str, str]:
    if should_use_mapping_for_explicit_pairs(parsed):
        expected_ssns = exact_mapping_ssns(regional, branches, mapping)
        if len(expected_ssns) == 1:
            output_ssn = expected_ssns[0]
            if output_ssn == explicit_ssn:
                return output_ssn, "catalog_table_confirmed"
            return output_ssn, "catalog_table_override"
        if len(expected_ssns) > 1:
            return "", "catalog_ambiguous_blank"
        return "", "catalog_missing_blank"

    if catalog_status.startswith("catalog_differs:"):
        return "", "catalog_conflict_blank"
    return explicit_ssn, "input_explicit"


def extract_explicit_ssn_pairs(
    value: str,
    branches: list[str],
    mapping: MappingStore,
) -> tuple[list[tuple[str, str, str, str]], list[str]]:
    explicit_pairs: list[tuple[str, str, str, str]] = []
    prepared = prepare_code_text(value)
    consumed_ranges: list[tuple[int, int]] = []
    segment_pattern = re.compile(rf"(?P<outer>{CODE_RE.pattern})?\s*(?P<groups>(?:\([^)]*\)\s*)+)")

    for match in segment_pattern.finditer(prepared):
        outside_ssn = normalize_code(match.group("outer") or "")
        group_text = match.group("groups")
        inner_tokens = ordered_unique(
            code
            for group in re.findall(r"\(([^)]*)\)", group_text)
            for code in extract_code_tokens(group)
        )

        if inner_tokens and outside_ssn and is_known_ssn_for_branches(outside_ssn, branches, mapping):
            source_segment = collapse_spaces(match.group(0))
            for regional in inner_tokens:
                explicit_pairs.append(
                    (
                        regional,
                        outside_ssn,
                        explicit_pair_catalog_status(regional, outside_ssn, branches, mapping),
                        source_segment,
                    )
                )
            consumed_ranges.append(match.span())

    if consumed_ranges:
        remainder_chars = list(prepared)
        for start, end in consumed_ranges:
            for index in range(start, end):
                remainder_chars[index] = " "
        remainder_text = "".join(remainder_chars)
    else:
        remainder_text = prepared

    return explicit_pairs, extract_code_tokens(remainder_text)


def is_branch_only_cell(value: str) -> bool:
    text = collapse_spaces(value).upper()
    if not text:
        return True
    if PLACEHOLDER_CODE in text:
        return True
    if text in {"X", "COD. REGIONALE", "CODICE REGIONALE"}:
        return True
    if "CONTRATTUALIZZATE" in text or "TUTTE LE PRESTAZIONI" in text:
        return True
    if "SETTORI" in text:
        return True
    return not extract_code_tokens(text)


def make_issue(parsed: ParsedFile | ProcessedFile | None, row: int | str, issue: str, details: str) -> Issue:
    source = parsed.relative_path if parsed else ""
    return Issue(source_file=source, row=str(row), issue=issue, details=details)


def load_template_headers() -> list[str]:
    workbook = load_workbook(TEMPLATE_FILE, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [collapse_spaces(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if headers != ["Codice STS.11", "BRANCA", "CODICE REGIONALE", "CODICE SSN"]:
        raise RuntimeError("Intestazioni template non attese.")
    return headers


def load_mapping() -> MappingStore:
    workbook = load_workbook(MAPPING_FILE, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    entries: list[MappingEntry] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        branches = tuple(
            normalize_branch_code(collapse_spaces(value))
            for value in row[:4]
            if collapse_spaces(value)
        )
        regional = normalize_code(collapse_spaces(row[4]))
        ssn = normalize_code(collapse_spaces(row[5]))
        if regional and ssn:
            entries.append(MappingEntry(branches=branches, regional=regional, ssn=ssn))
    return MappingStore(entries)


def find_header(sheet) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        normalized = [collapse_spaces(value) for value in row]
        if HEADER_STS in normalized:
            indices = {value: index for index, value in enumerate(normalized) if value}
            return row_number, indices
    raise RuntimeError("Header non trovato.")


def cell_text(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return collapse_spaces(row[index])


def source_relative_path(path: Path) -> str:
    for base_dir in (BANCADATI_DIR, ROOT):
        try:
            return str(path.relative_to(base_dir))
        except ValueError:
            continue
    return path.name


def parse_source_file(path: Path) -> ParsedFile:
    relative_path = source_relative_path(path)
    parsed = ParsedFile(path=path, relative_path=relative_path)

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row, indices = find_header(sheet)

    code_index = indices.get(HEADER_CODE)
    if code_index is None:
        code_index = indices.get(HEADER_CODE_REGIONALE)

    national_index = indices.get(HEADER_CODE_NAZIONALE)

    empty_streak = 0
    seen_non_empty = False
    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        raw_row = RawRow(
            row_number=row_number,
            sts_text=cell_text(row, indices.get(HEADER_STS)),
            denom_text=cell_text(row, indices.get(HEADER_DENOM)),
            branch_text=cell_text(row, indices.get(HEADER_BRANCH)),
            dpcm_text=cell_text(row, indices.get(HEADER_DPCM)),
            regional_text=cell_text(row, code_index),
            national_text=cell_text(row, national_index),
        )
        if any(
            [
                raw_row.sts_text,
                raw_row.denom_text,
                raw_row.branch_text,
                raw_row.dpcm_text,
                raw_row.regional_text,
                raw_row.national_text,
            ]
        ):
            parsed.raw_rows.append(raw_row)
            seen_non_empty = True
            empty_streak = 0
            if not parsed.sts:
                parsed.sts = normalize_sts(raw_row.sts_text)
            if not parsed.denom:
                parsed.denom = normalize_denom(raw_row.denom_text)
        else:
            empty_streak += 1
            if seen_non_empty and empty_streak >= 50:
                break

    if not parsed.sts:
        for raw_row in parsed.raw_rows:
            parsed.sts = normalize_sts(raw_row.sts_text)
            if parsed.sts:
                break
    if not parsed.denom:
        for raw_row in parsed.raw_rows:
            parsed.denom = normalize_denom(raw_row.denom_text)
            if parsed.denom:
                break

    return parsed


def extract_pdf_text(path: Path, *, raw: bool) -> str:
    command = ["pdftotext"]
    if raw:
        command.append("-raw")
    else:
        command.append("-layout")
    command.extend([str(path), "-"])
    result = subprocess.run(command, check=True, capture_output=True, text=True)
    return result.stdout


def parse_pdf_denominazione(text: str, path: Path) -> str:
    patterns = [
        r"(ECORAD\s+SRL\s*-\s*STUDIO\s+DI\s+RADIOLOGIA\s+ED\s+ECOGRAFIA)",
        r"(STUDIO\s+RADIOLOGICO\s+DOTT\.\s+VINCENZO\s+ARCERI\s+SAS)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return collapse_spaces(match.group(1))
    return path.parent.name


def parse_pdf_branch(text: str) -> str:
    patterns = [
        r"\b(0?\d|1\d|2[0-6])\s*-\s*([A-Z][A-Z ]{3,})",
        r"\b(0?\d|1\d|2[0-6])\s+([A-Z][A-Z ]{3,})\s+\d{2}/\d{2}/\d{4}",
        r"\bBRANCA\s+(0?\d|1\d|2[0-6])\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        code = normalize_branch_code(match.group(1))
        if len(match.groups()) > 1:
            desc = collapse_spaces(match.group(2))
            return f"{code} - {desc}"
        return code
    return ""


def parse_pdf_source_file(path: Path) -> SourceParseResult:
    relative_path = source_relative_path(path)
    raw_text = extract_pdf_text(path, raw=True)
    layout_text = extract_pdf_text(path, raw=False)
    compact_upper = re.sub(r"\s+", " ", raw_text.upper())
    if "CODICE STS11" not in compact_upper or "CODICE PRESTAZIONE" not in compact_upper:
        return SourceParseResult(parsed=None, supported=False)

    parsed = ParsedFile(path=path, relative_path=relative_path)
    parsed.sts = normalize_sts(raw_text)
    parsed.denom = parse_pdf_denominazione(raw_text, path)
    branch_text = parse_pdf_branch(layout_text)
    codes = ordered_unique(
        code
        for line in layout_text.splitlines()
        for code in extract_code_tokens(line)
    )
    if not codes:
        parsed.issues.append(make_issue(parsed, "-", "pdf_senza_codici", "Nessun codice prestazione estratto"))
    for row_number, code in enumerate(codes, start=1):
        parsed.raw_rows.append(
            RawRow(
                row_number=row_number,
                sts_text=parsed.sts,
                denom_text=parsed.denom,
                branch_text=branch_text,
                dpcm_text=branch_text,
                regional_text=code,
                national_text="",
            )
        )
    return SourceParseResult(parsed=parsed, supported=True)


def parse_source_path(path: Path) -> SourceParseResult:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return SourceParseResult(parsed=parse_source_file(path), supported=True)
    if suffix == ".pdf":
        return parse_pdf_source_file(path)
    return SourceParseResult(parsed=None, supported=False)


def build_learned_descriptions(parsed_files: list[ParsedFile]) -> dict[str, str]:
    counter: dict[str, Counter[str]] = defaultdict(Counter)
    for parsed in parsed_files:
        for raw_row in parsed.raw_rows:
            direct_codes = ordered_unique(
                extract_branch_codes_from_text(raw_row.branch_text)
                + extract_branch_codes_from_text(raw_row.dpcm_text)
            )
            if not direct_codes:
                continue
            for text in (raw_row.branch_text, raw_row.dpcm_text):
                key = branch_text_key(text)
                if not key:
                    continue
                for code in direct_codes:
                    counter[key][code] += 1

    learned: dict[str, str] = {}
    for key, code_counter in counter.items():
        if len(code_counter) == 1:
            learned[key] = next(iter(code_counter))
    return learned


def infer_missing_metadata(parsed_files: list[ParsedFile]) -> None:
    by_parent: dict[Path, list[ParsedFile]] = defaultdict(list)
    for parsed in parsed_files:
        by_parent[parsed.path.parent].append(parsed)

    for siblings in by_parent.values():
        known_sts = ordered_unique(parsed.sts for parsed in siblings if parsed.sts)
        known_denom = ordered_unique(parsed.denom for parsed in siblings if parsed.denom)

        if len(known_sts) == 1:
            for parsed in siblings:
                if not parsed.sts:
                    parsed.sts = known_sts[0]
                    parsed.issues.append(
                        make_issue(parsed, "-", "sts_inferito_da_cartella", f"Usato STS {known_sts[0]}")
                    )

        if len(known_denom) == 1:
            for parsed in siblings:
                if not parsed.denom:
                    parsed.denom = known_denom[0]
                    parsed.issues.append(
                        make_issue(
                            parsed,
                            "-",
                            "denominazione_inferita_da_cartella",
                            f"Usata denominazione {known_denom[0]}",
                        )
                    )


def filter_entries_by_branches(entries: Iterable[MappingEntry], branches: Iterable[str]) -> list[MappingEntry]:
    branch_set = {branch for branch in branches if branch}
    if not branch_set:
        return list(entries)
    return [
        entry
        for entry in entries
        if any(branch in branch_set for branch in entry.branches)
    ]


def unique_entries_by_ssn(entries: Iterable[MappingEntry]) -> list[MappingEntry]:
    by_key: dict[tuple[str, str], MappingEntry] = {}
    for entry in entries:
        by_key[(entry.regional, entry.ssn)] = entry
    return list(by_key.values())


def match_mapping(code: str, branches: list[str], mapping: MappingStore) -> tuple[MappingEntry | None, str]:
    exact = unique_entries_by_ssn(filter_entries_by_branches(mapping.by_code.get(code, []), branches))
    if not exact:
        exact = unique_entries_by_ssn(mapping.by_code.get(code, []))
    if len(exact) == 1:
        return exact[0], "exact"
    if len(exact) > 1:
        return None, "ambiguous_exact"

    prefix_candidates = [
        entry
        for entry in mapping.entries
        if entry.regional.startswith(f"{code}.")
    ]
    prefix_filtered = unique_entries_by_ssn(filter_entries_by_branches(prefix_candidates, branches))
    if not prefix_filtered:
        prefix_filtered = unique_entries_by_ssn(prefix_candidates)
    if len(prefix_filtered) == 1:
        return prefix_filtered[0], "prefix"
    if len(prefix_filtered) > 1:
        return None, "ambiguous_prefix"

    return None, "missing"


def is_known_ssn_for_branches(code: str, branches: list[str], mapping: MappingStore) -> bool:
    if not code or not branches:
        return False
    known_branches = mapping.ssn_to_branches.get(code, set())
    return any(branch in known_branches for branch in branches)


def infer_branches_from_codes(codes: list[str], mapping: MappingStore) -> list[str]:
    inferred: list[str] = []
    for code in codes:
        exact_branches = mapping.code_to_branches.get(code)
        if exact_branches and len(exact_branches) == 1:
            inferred.extend(exact_branches)
    return ordered_unique(inferred)


def choose_output_branch(resolved_entry: MappingEntry | None, requested_branches: list[str]) -> str:
    if resolved_entry and requested_branches:
        for branch in requested_branches:
            if branch in resolved_entry.branches:
                return branch
    if requested_branches:
        return requested_branches[0]
    if resolved_entry and resolved_entry.branches:
        return resolved_entry.branches[0]
    return ""


def sanitize_filename_component(value: str) -> str:
    folded = ascii_fold(value).strip()
    folded = re.sub(r"[\\/:*?\"<>|]+", " ", folded)
    folded = re.sub(r"\s+", "_", folded)
    folded = re.sub(r"_+", "_", folded).strip("_")
    return folded or "SENZA_DENOMINAZIONE"


def build_output_name(sts: str, denom: str, index: int, total: int) -> str:
    safe_denom = sanitize_filename_component(denom)
    if total > 1:
        return f"{sts}_{index:02d}_{safe_denom}.xlsx"
    return f"{sts}_{safe_denom}.xlsx"


def is_data_output_file(path: Path) -> bool:
    return path.suffix.lower() == ".xlsx" and not path.name.startswith("_")


def add_output_row(
    rows_by_key: dict[tuple[str, str], tuple[str, str, str, str]],
    sts: str,
    branch: str,
    regional: str,
    ssn: str,
) -> None:
    key = (regional, ssn)
    if key not in rows_by_key:
        rows_by_key[key] = (sts, branch, regional, ssn)


def remove_conflicting_rows(
    rows_by_key: dict[tuple[str, str], tuple[str, str, str, str]],
    regional: str,
    preferred_ssn: str,
) -> None:
    for key in [key for key in rows_by_key if key[0] == regional and key[1] != preferred_ssn]:
        del rows_by_key[key]


def is_blocked_by_explicit_override(regional: str, ssn: str, explicit_overrides: dict[str, str]) -> bool:
    if regional not in explicit_overrides:
        return False
    return explicit_overrides[regional] != ssn


def process_file(parsed: ParsedFile, mapping: MappingStore, learned_descriptions: dict[str, str]) -> ProcessedFile:
    issues = list(parsed.issues)
    rows_by_key: dict[tuple[str, str], tuple[str, str, str, str]] = {}
    explicit_matches: list[tuple[int, str, str, str, str, str, str]] = []
    explicit_overrides: dict[str, str] = {}

    if not parsed.sts:
        issues.append(make_issue(parsed, "-", "sts_mancante", "Impossibile determinare il codice STS"))
    if not parsed.denom:
        issues.append(make_issue(parsed, "-", "denominazione_mancante", "Impossibile determinare la denominazione"))

    file_branch_candidates: list[str] = []
    for raw_row in parsed.raw_rows:
        file_branch_candidates.extend(
            infer_branch_codes(raw_row.branch_text, raw_row.dpcm_text, learned_descriptions)
        )
        file_branch_candidates.extend(
            infer_branches_from_codes(extract_code_tokens(raw_row.regional_text), mapping)
        )
    file_default_branches = ordered_unique(file_branch_candidates)
    if len(file_default_branches) != 1:
        file_default_branches = []

    current_branches: list[str] = []
    for raw_row in parsed.raw_rows:
        row_branches = infer_branch_codes(raw_row.branch_text, raw_row.dpcm_text, learned_descriptions)
        branch_inference_tokens = extract_code_tokens(raw_row.regional_text)
        national_tokens = extract_code_tokens(raw_row.national_text)

        if not row_branches:
            row_branches = infer_branches_from_codes(branch_inference_tokens, mapping)
        if not row_branches:
            row_branches = current_branches or file_default_branches
        if row_branches:
            current_branches = row_branches

        explicit_pairs, regional_tokens = extract_explicit_ssn_pairs(raw_row.regional_text, row_branches, mapping)

        if not any([row_branches, regional_tokens, national_tokens, explicit_pairs]):
            continue

        branch_only = is_branch_only_cell(raw_row.regional_text) and not explicit_pairs

        if branch_only:
            if not row_branches:
                issues.append(
                    make_issue(
                        parsed,
                        raw_row.row_number,
                        "branca_non_risolta",
                        "Riga senza codice prestazione e senza branca risolvibile",
                    )
                )
                continue
            if raw_row.regional_text and collapse_spaces(raw_row.regional_text).upper() not in {"", "X"}:
                issues.append(
                    make_issue(
                        parsed,
                        raw_row.row_number,
                        "espansione_branca",
                        f"Usata espansione per branca da testo '{collapse_spaces(raw_row.regional_text)}'",
                    )
                )
            for branch in row_branches:
                for entry in mapping.by_branch.get(branch, []):
                    if is_blocked_by_explicit_override(entry.regional, entry.ssn, explicit_overrides):
                        continue
                    add_output_row(rows_by_key, parsed.sts, branch, entry.regional, entry.ssn)
            continue

        branch = row_branches[0] if row_branches else ""
        for regional, ssn, catalog_status, source_line in explicit_pairs:
            output_ssn, output_rule = resolve_explicit_output_ssn(
                parsed,
                regional,
                ssn,
                row_branches,
                mapping,
                catalog_status,
            )
            explicit_matches.append(
                (raw_row.row_number, regional, ssn, output_ssn, catalog_status, output_rule, source_line)
            )

            explicit_overrides[regional] = output_ssn
            remove_conflicting_rows(rows_by_key, regional, output_ssn)
            add_output_row(rows_by_key, parsed.sts, branch, regional, output_ssn)

            if catalog_status == "catalog_missing":
                issues.append(
                    make_issue(
                        parsed,
                        raw_row.row_number,
                        "mapping_parentesi_catalogo_mancante",
                        f"{regional} -> {ssn} da riga esplicita '{source_line}'",
                    )
                )
            elif output_rule == "catalog_ambiguous_blank":
                issues.append(
                    make_issue(
                        parsed,
                        raw_row.row_number,
                        "mapping_parentesi_catalogo_ambiguo",
                        f"{regional} ha piu codici SSN candidati per la branca {branch}",
                    )
                )
            elif catalog_status.startswith("catalog_differs:"):
                issues.append(
                    make_issue(
                        parsed,
                        raw_row.row_number,
                        "mapping_parentesi_catalogo_diverso",
                        f"{regional} -> {ssn} da riga esplicita '{source_line}' (catalogo {catalog_status.split(':', 1)[1]})",
                    )
                )

        filtered_tokens: list[str] = []
        last_ssn = ""
        for token in regional_tokens:
            if token == last_ssn:
                continue
            entry, strategy = match_mapping(token, row_branches, mapping)
            if entry is not None:
                filtered_tokens.append(token)
                last_ssn = entry.ssn
            else:
                filtered_tokens.append(token)
                last_ssn = ""

        for token in filtered_tokens:
            entry, strategy = match_mapping(token, row_branches, mapping)
            if entry is not None:
                branch = choose_output_branch(entry, row_branches)
                if is_blocked_by_explicit_override(token, entry.ssn, explicit_overrides):
                    continue
                add_output_row(rows_by_key, parsed.sts, branch, token, entry.ssn)
                if strategy == "prefix":
                    issues.append(
                        make_issue(
                            parsed,
                            raw_row.row_number,
                            "mapping_prefisso",
                            f"{token} risolto tramite {entry.regional} -> {entry.ssn}",
                        )
                    )
                continue

            ssn_fallback = national_tokens[0] if len(national_tokens) == 1 else ""
            if ssn_fallback:
                if is_blocked_by_explicit_override(token, ssn_fallback, explicit_overrides):
                    continue
                add_output_row(rows_by_key, parsed.sts, branch, token, ssn_fallback)
                issues.append(
                    make_issue(
                        parsed,
                        raw_row.row_number,
                        "mapping_fallback_nazionale",
                        f"{token} non trovato in anagrafica, usato SSN {ssn_fallback}",
                    )
                )
                continue

            if is_known_ssn_for_branches(token, row_branches, mapping):
                if is_blocked_by_explicit_override(token, token, explicit_overrides):
                    continue
                add_output_row(rows_by_key, parsed.sts, branch, token, token)
                issues.append(
                    make_issue(
                        parsed,
                        raw_row.row_number,
                        "mapping_codice_gia_ssn",
                        f"{token} gia presente come codice SSN per la branca {branch}",
                    )
                )
                continue

            if is_blocked_by_explicit_override(token, "", explicit_overrides):
                continue
            add_output_row(rows_by_key, parsed.sts, branch, token, "")
            issues.append(
                make_issue(
                    parsed,
                    raw_row.row_number,
                    "mapping_non_trovato",
                    f"{token} non risolto",
                )
            )

    return ProcessedFile(
        path=parsed.path,
        relative_path=parsed.relative_path,
        sts=parsed.sts,
        denom=parsed.denom or parsed.path.parent.name,
        output_rows=sorted(rows_by_key.values(), key=lambda item: (item[1], item[2], item[3])),
        issues=issues,
        explicit_matches=explicit_matches,
    )


def write_output_file(path: Path, headers: list[str], rows: list[tuple[str, str, str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Foglio1"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def write_report_file(path: Path, headers: list[str], rows: list[tuple[str, ...]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anomalie"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def main() -> None:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    headers = load_template_headers()
    mapping = load_mapping()

    candidate_files = sorted(
        path
        for path in BANCADATI_DIR.rglob("*")
        if path.suffix.lower() in {".xlsx", ".pdf"} and path.name != "~$" and not path.name.startswith("~$")
    )

    parse_results = [parse_source_path(path) for path in candidate_files]
    parsed_files = [result.parsed for result in parse_results if result.supported and result.parsed is not None]
    infer_missing_metadata(parsed_files)
    learned_descriptions = build_learned_descriptions(parsed_files)
    processed_files = [process_file(parsed, mapping, learned_descriptions) for parsed in parsed_files]

    OUTPUT_DIR.mkdir(exist_ok=True)
    for existing_output in OUTPUT_DIR.glob("*.xlsx"):
        existing_output.unlink()

    groups: dict[str, list[ProcessedFile]] = defaultdict(list)
    for processed in processed_files:
        groups[processed.sts].append(processed)

    all_issues: list[Issue] = []
    incremental_lines: list[str] = []
    manifest_lines: list[str] = []
    unresolved_report_lines: list[str] = []
    explicit_match_lines: list[str] = []
    unresolved_excel_rows: list[tuple[str, str, str]] = []

    for sts, group in sorted(groups.items(), key=lambda item: item[0]):
        sorted_group = sorted(group, key=lambda item: item.relative_path)
        total = len(sorted_group)
        for index, processed in enumerate(sorted_group, start=1):
            if not processed.sts:
                processed.issues.append(
                    make_issue(processed, "-", "output_non_generato", "STS assente")
                )
                all_issues.extend(processed.issues)
                continue

            output_name = build_output_name(processed.sts, processed.denom, index, total)
            output_path = OUTPUT_DIR / output_name
            write_output_file(output_path, headers, processed.output_rows)
            manifest_lines.append(f"{output_name}\t{processed.relative_path}\t{len(processed.output_rows)}")

            unresolved_codes = ordered_unique(row[2] for row in processed.output_rows if not row[3])
            if unresolved_codes:
                unresolved_excel_rows.append(
                    (
                        processed.sts,
                        f"{processed.path.parent.name}/{processed.path.name}",
                        ", ".join(unresolved_codes),
                    )
                )
                unresolved_report_lines.append(
                    "\t".join(
                        [
                            processed.path.parent.name,
                            processed.path.name,
                            output_name,
                            ", ".join(unresolved_codes),
                        ]
                    )
                )

            for row_number, regional, input_ssn, output_ssn, catalog_status, output_rule, source_line in processed.explicit_matches:
                explicit_match_lines.append(
                    "\t".join(
                        [
                            processed.path.parent.name,
                            processed.path.name,
                            output_name,
                            str(row_number),
                            regional,
                            input_ssn,
                            output_ssn,
                            catalog_status,
                            output_rule,
                            source_line,
                        ]
                    )
                )

            if total > 1:
                incremental_lines.append(f"{output_path.stem} | {processed.relative_path} ;")

            all_issues.extend(processed.issues)

    (OUTPUT_DIR / "_incrementali.txt").write_text(
        "\n".join(incremental_lines) + ("\n" if incremental_lines else ""),
        encoding="utf-8",
    )

    with (OUTPUT_DIR / "_anomalie.tsv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["source_file", "row", "issue", "details"])
        for issue in all_issues:
            writer.writerow([issue.source_file, issue.row, issue.issue, issue.details])

    (OUTPUT_DIR / "_manifest.tsv").write_text(
        "output_file\tsource_file\trows\n" + "\n".join(manifest_lines) + ("\n" if manifest_lines else ""),
        encoding="utf-8",
    )
    (OUTPUT_DIR / "_prestazioni_non_ricavabili.tsv").write_text(
        "folder_name\tsource_file\toutput_file\tprestazioni_non_ricavabili\n"
        + "\n".join(unresolved_report_lines)
        + ("\n" if unresolved_report_lines else ""),
        encoding="utf-8",
    )
    write_report_file(
        OUTPUT_DIR / "_anomalie_non_riconducibili_ssn.xlsx",
        ["CODICE STRUTTURA", "CARTELLA/FILE INPUT", "PRESTAZIONI NON RICONDUCIBILI A SSN"],
        unresolved_excel_rows,
    )
    (OUTPUT_DIR / "_match_parentesi_espliciti.tsv").write_text(
        "folder_name\tsource_file\toutput_file\trow\tcodice_regionale\tcodice_ssn_input\tcodice_ssn_output\tverifica_catalogo\tregola_output\tsource_line\n"
        + "\n".join(explicit_match_lines)
        + ("\n" if explicit_match_lines else ""),
        encoding="utf-8",
    )

    generated_outputs = sum(1 for path in OUTPUT_DIR.glob("*.xlsx") if is_data_output_file(path))
    print(f"File sorgente processati: {len(parsed_files)}")
    print(f"File output generati: {generated_outputs}")
    print(f"Anomalie registrate: {len(all_issues)}")
    print(f"Log incrementali: {OUTPUT_DIR / '_incrementali.txt'}")
    print(f"Log anomalie: {OUTPUT_DIR / '_anomalie.tsv'}")
    print(f"Report prestazioni non ricavabili: {OUTPUT_DIR / '_prestazioni_non_ricavabili.tsv'}")
    print(f"Report excel anomalie: {OUTPUT_DIR / '_anomalie_non_riconducibili_ssn.xlsx'}")
    print(f"Report match parentesi: {OUTPUT_DIR / '_match_parentesi_espliciti.tsv'}")


if __name__ == "__main__":
    main()
