#!/usr/bin/env python3

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

try:
    from .etl_bancadati import (
        BANCADATI_DIR,
        OUTPUT_DIR,
        build_learned_descriptions,
        build_output_name,
        infer_missing_metadata,
        load_mapping,
        load_template_headers,
        parse_source_path,
        process_file,
    )
except ImportError:
    from etl_bancadati import (
        BANCADATI_DIR,
        OUTPUT_DIR,
        build_learned_descriptions,
        build_output_name,
        infer_missing_metadata,
        load_mapping,
        load_template_headers,
        parse_source_path,
        process_file,
    )


@dataclass
class ValidationIssue:
    output_file: str
    source_file: str
    severity: str
    issue: str
    details: str


def normalized_output_rows(path: Path) -> tuple[list[str], list[tuple[str, str, str, str]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header = [("" if value is None else str(value).strip()) for value in rows[0][:4]]
    data = [
        tuple("" if value is None else str(value).strip() for value in row[:4])
        for row in rows[1:]
        if any(value not in (None, "") for value in row[:4])
    ]
    return header, data


def summarize_source_issues(issue_types: Counter[str]) -> str:
    if not issue_types:
        return ""
    return ", ".join(f"{issue}:{count}" for issue, count in issue_types.most_common())


def main() -> None:
    template_headers = load_template_headers()
    mapping = load_mapping()

    candidate_files = sorted(
        path
        for path in BANCADATI_DIR.rglob("*")
        if path.suffix.lower() in {".xlsx", ".pdf"} and path.name != "~$" and not path.name.startswith("~$")
    )

    parse_results = [parse_source_path(path) for path in candidate_files]
    parsed_files = [result.parsed for result in parse_results if result.supported and result.parsed is not None]

    infer_missing_metadata(parsed_files)
    learned_descriptions = build_learned_descriptions(parsed_files)
    processed_files = [process_file(parsed, mapping, learned_descriptions) for parsed in parsed_files]

    groups: dict[str, list] = defaultdict(list)
    for processed in processed_files:
        groups[processed.sts].append(processed)

    expected_outputs: dict[str, object] = {}
    for sts, group in sorted(groups.items(), key=lambda item: item[0]):
        sorted_group = sorted(group, key=lambda item: item.relative_path)
        total = len(sorted_group)
        for index, processed in enumerate(sorted_group, start=1):
            output_name = build_output_name(processed.sts, processed.denom, index, total)
            expected_outputs[output_name] = processed

    actual_output_files = {
        path.name: path
        for path in OUTPUT_DIR.glob("*.xlsx")
        if not path.name.startswith("_")
    }
    issues: list[ValidationIssue] = []
    per_file_rows: list[tuple[str, str, str, int, str, str]] = []

    extra_outputs = sorted(set(actual_output_files) - set(expected_outputs))
    for output_name in extra_outputs:
        issues.append(
            ValidationIssue(
                output_file=output_name,
                source_file="",
                severity="error",
                issue="output_extra",
                details="File presente in output ma non atteso dal processo ETL",
            )
        )

    missing_outputs = sorted(set(expected_outputs) - set(actual_output_files))
    for output_name in missing_outputs:
        processed = expected_outputs[output_name]
        issues.append(
            ValidationIssue(
                output_file=output_name,
                source_file=processed.relative_path,
                severity="error",
                issue="output_mancante",
                details="File atteso non trovato in output",
            )
        )

    for output_name, processed in sorted(expected_outputs.items()):
        output_path = actual_output_files.get(output_name)
        if output_path is None:
            continue

        header, actual_rows = normalized_output_rows(output_path)
        expected_rows = list(processed.output_rows)
        source_issue_types = Counter(issue.issue for issue in processed.issues)

        if header != template_headers:
            issues.append(
                ValidationIssue(
                    output_file=output_name,
                    source_file=processed.relative_path,
                    severity="error",
                    issue="header_mismatch",
                    details=f"Header atteso {template_headers}, trovato {header}",
                )
            )

        if actual_rows != expected_rows:
            details = f"righe attese={len(expected_rows)} trovate={len(actual_rows)}"
            for index, (expected_row, actual_row) in enumerate(zip(expected_rows, actual_rows), start=1):
                if expected_row != actual_row:
                    details += f"; prima differenza riga {index}: atteso={expected_row} trovato={actual_row}"
                    break
            else:
                if len(expected_rows) != len(actual_rows):
                    if len(expected_rows) > len(actual_rows):
                        details += f"; prima riga mancante attesa={expected_rows[len(actual_rows)]}"
                    else:
                        details += f"; prima riga extra trovata={actual_rows[len(expected_rows)]}"
            issues.append(
                ValidationIssue(
                    output_file=output_name,
                    source_file=processed.relative_path,
                    severity="error",
                    issue="contenuto_mismatch",
                    details=details,
                )
            )

        if len(actual_rows) != len(set(actual_rows)):
            issues.append(
                ValidationIssue(
                    output_file=output_name,
                    source_file=processed.relative_path,
                    severity="error",
                    issue="duplicati_output",
                    details="Output contiene righe duplicate",
                )
            )

        invalid_branch_rows = [
            row for row in actual_rows if row[1] and (not row[1].isdigit() or len(row[1]) != 2)
        ]
        if invalid_branch_rows:
            issues.append(
                ValidationIssue(
                    output_file=output_name,
                    source_file=processed.relative_path,
                    severity="error",
                    issue="branca_non_normalizzata",
                    details=f"Esempio: {invalid_branch_rows[0]}",
                )
            )

        invalid_sts_rows = [row for row in actual_rows if row[0] != processed.sts]
        if invalid_sts_rows:
            issues.append(
                ValidationIssue(
                    output_file=output_name,
                    source_file=processed.relative_path,
                    severity="error",
                    issue="sts_non_uniforme",
                    details=f"Esempio: {invalid_sts_rows[0]}",
                )
            )

        semantic_issue_types = Counter(
            {
                issue: count
                for issue, count in source_issue_types.items()
                if issue in {"mapping_non_trovato", "mapping_fallback_nazionale", "sts_inferito_da_cartella"}
            }
        )
        if semantic_issue_types:
            issues.append(
                ValidationIssue(
                    output_file=output_name,
                    source_file=processed.relative_path,
                    severity="warning",
                    issue="rischio_semantico",
                    details=summarize_source_issues(semantic_issue_types),
                )
            )

        per_file_rows.append(
            (
                output_name,
                processed.relative_path,
                "ok" if not [i for i in issues if i.output_file == output_name and i.severity == "error"] else "error",
                len(actual_rows),
                summarize_source_issues(source_issue_types),
                processed.sts,
            )
        )

    structural_errors = [issue for issue in issues if issue.severity == "error"]
    warnings = [issue for issue in issues if issue.severity == "warning"]

    report_path = OUTPUT_DIR / "_validation_report.tsv"
    with report_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["output_file", "source_file", "severity", "issue", "details"])
        for issue in issues:
            writer.writerow([issue.output_file, issue.source_file, issue.severity, issue.issue, issue.details])

    per_file_path = OUTPUT_DIR / "_validation_per_file.tsv"
    with per_file_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["output_file", "source_file", "status", "rows", "source_issues", "sts"])
        for row in per_file_rows:
            writer.writerow(row)

    summary_path = OUTPUT_DIR / "_validation_summary.txt"
    summary_lines = [
        f"source_supported={len(parsed_files)}",
        f"output_expected={len(expected_outputs)}",
        f"output_present={len(actual_output_files)}",
        f"structural_errors={len(structural_errors)}",
        f"warnings={len(warnings)}",
        f"report={report_path}",
        f"per_file={per_file_path}",
    ]
    summary_path.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")

    print(f"Sorgenti supportati: {len(parsed_files)}")
    print(f"Output attesi: {len(expected_outputs)}")
    print(f"Output presenti: {len(actual_output_files)}")
    print(f"Errori strutturali: {len(structural_errors)}")
    print(f"Warning semantici: {len(warnings)}")
    print(f"Report: {report_path}")
    print(f"Per file: {per_file_path}")


if __name__ == "__main__":
    main()
